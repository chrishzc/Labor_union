"""Read-only monthly accounts-payable export for bank transfers."""

from __future__ import annotations

from calendar import monthrange
from datetime import date, datetime, timedelta
from decimal import Decimal
from io import BytesIO
import re

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from services.db_service import get_connection


EXPORT_HEADERS = (
    "月份-銀行代碼-流水號",
    "銀行名稱",
    "客戶or服務人員姓名",
    "銀行帳號",
    "銀行代號(碼)",
    "金額",
    "身分證字號(匯款到永豐才要填)",
    "案件編號",
    "匯款日期",
)

OUTGOING_BANKS = {
    "31": "永豐銀行",
    "633": "台新銀行",
}

ACCOUNTS_PAYABLE_SUMMARY_HEADERS = (
    "案件編號",
    "客戶姓名",
    "完成日期",
    "服務人員付款日",
    "補助退款日",
    "應付薪資",
    "已付薪資",
    "薪資未付",
    "應付補助款",
    "已退補助款",
    "補助剩餘",
)


def _month_bounds(target_month: str) -> tuple[date, date]:
    if not isinstance(target_month, str) or not re.fullmatch(r"\d{4}-(0[1-9]|1[0-2])", target_month):
        raise ValueError("target_month 必須為 YYYY-MM")
    try:
        first_day = datetime.strptime(target_month, "%Y-%m").date().replace(day=1)
    except (TypeError, ValueError) as exc:
        raise ValueError("target_month 必須為 YYYY-MM") from exc

    last_day = date(
        first_day.year,
        first_day.month,
        monthrange(first_day.year, first_day.month)[1],
    )
    return first_day, last_day


def _as_decimal(value) -> Decimal:
    return Decimal(str(value or 0))


def _to_date(value) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if hasattr(value, "date"):
        return value.date()
    try:
        return datetime.strptime(str(value).split(" ")[0].strip(), "%Y-%m-%d").date()
    except:
        return None


def _safe_int(value) -> int:
    try:
        return int(value or 0)
    except:
        return 0


def _month_index(base_date: date, month_offset: int) -> date:
    month_index = base_date.year * 12 + (base_date.month - 1) + month_offset
    year = month_index // 12
    month = month_index % 12 + 1
    return date(year=year, month=month, day=15)


def _derive_service_end_date(row: dict) -> date | None:
    actual_end_date = _to_date(row.get("actual_end_date"))
    if actual_end_date:
        return actual_end_date

    actual_start_date = _to_date(row.get("actual_start_date"))
    service_days = _safe_int(row.get("service_days"))
    if not actual_start_date or not service_days:
        return None
    return actual_start_date + timedelta(days=max(service_days - 1, 0))


def _derive_staff_payment_date(row: dict) -> str:
    end_date = _derive_service_end_date(row)
    if not end_date:
        return ""

    identity_status = str(row.get("identity_status") or "").strip()
    month_offset = 2 if identity_status == "補助市民" else 1
    return _month_index(end_date, month_offset).isoformat()


def _derive_subsidy_refund_date(row: dict) -> str:
    end_date = _derive_service_end_date(row)
    identity_status = str(row.get("identity_status") or "").strip()
    if not end_date or identity_status == "非市民":
        return ""

    month_end_day = monthrange(end_date.year, end_date.month)[1]
    return (date(end_date.year, end_date.month, month_end_day) + timedelta(days=5)).isoformat()


def _fetch_payables(target_month: str) -> list[dict]:
    first_day, last_day = _month_bounds(target_month)
    conn = get_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute(
                """
                SELECT sp.id AS source_payment_id,
                       sp.staff_id,
                       sp.case_no,
                       COALESCE(od.salary_payment_date_1, sp.due_date) AS transfer_date,
                       sp.total_payable AS amount,
                       sp.amount_paid,
                       sp.payment_status,
                       s.name AS recipient_name,
                       s.identity_card,
                       sba.account_no,
                       sba.bank_code AS recipient_bank_code
                FROM staff_payments sp
                JOIN staff s ON s.id = sp.staff_id
                LEFT JOIN v_order_details od ON od.case_no = sp.case_no
                LEFT JOIN staff_bank_accounts sba ON sba.id = (
                    SELECT sba2.id
                    FROM staff_bank_accounts sba2
                    WHERE sba2.staff_id = s.id
                      AND sba2.is_primary = 1
                    ORDER BY sba2.id
                    LIMIT 1
                )
                WHERE COALESCE(od.salary_payment_date_1, sp.due_date) BETWEEN %s AND %s
                AND od.order_status IN ('訂單成立', '服務中', '訂單完成')
                ORDER BY sp.staff_id, COALESCE(od.salary_payment_date_1, sp.due_date), sp.id
                """,
                (first_day, last_day),
            )
            staff_rows = cursor.fetchall()

            cursor.execute(
                """
                SELECT cp.id AS source_payment_id,
                       cp.case_no,
                       COALESCE(od.govt_claim_date, cp.subsidy_return_due_date) AS transfer_date,
                       cp.subsidy_return_receivable AS amount,
                       cp.subsidy_return_refunded,
                       cp.subsidy_return_review_status,
                       cp.subsidy_return_review_reason,
                       c.name AS recipient_name,
                       br.refund_account_no AS account_no,
                       br.refund_bank_code AS recipient_bank_code
                FROM client_payments cp
                JOIN clients c ON c.case_no = cp.case_no
                LEFT JOIN v_order_details od ON od.case_no = cp.case_no
                LEFT JOIN beclass_records br ON br.query_no = cp.case_no
                WHERE COALESCE(od.govt_claim_date, cp.subsidy_return_due_date) BETWEEN %s AND %s
                AND od.order_status IN ('訂單成立', '服務中', '訂單完成')
                AND cp.subsidy_return_receivable > 0
                ORDER BY COALESCE(od.govt_claim_date, cp.subsidy_return_due_date), cp.id
                """,
                (first_day, last_day),
            )
            client_rows = cursor.fetchall()

    finally:
        conn.close()

    staff_groups = {}
    for row in staff_rows:
        staff_id = int(row["staff_id"])
        group = staff_groups.setdefault(staff_id, {
            "source_payment_ids": [],
            "case_nos": set(),
            "transfer_dates": [],
            "amount": Decimal("0"),
            "payment_statuses": set(),
            "recipient_name": row.get("recipient_name"),
            "identity_card": row.get("identity_card"),
            "account_no": row.get("account_no"),
            "recipient_bank_code": row.get("recipient_bank_code"),
        })
        group["source_payment_ids"].append(int(row["source_payment_id"]))
        group["case_nos"].add(str(row["case_no"]))
        group["transfer_dates"].append(row["transfer_date"])
        group["amount"] += _as_decimal(row["amount"])
        if row.get("payment_status"):
            group["payment_statuses"].add(str(row["payment_status"]))

    rows = []
    for group in staff_groups.values():
        transfer_date = min(
            group["transfer_dates"],
            key=lambda value: _to_date(value) or date.max,
        )
        rows.append({
            "source_payment_id": min(group["source_payment_ids"]),
            "case_no": ",".join(sorted(group["case_nos"])),
            "transfer_date": transfer_date,
            "amount": group["amount"],
            "recipient_name": group.get("recipient_name"),
            "identity_card": group.get("identity_card"),
            "account_no": group.get("account_no"),
            "recipient_bank_code": group.get("recipient_bank_code"),
            "outgoing_bank_code": "31",
            "outgoing_bank_name": OUTGOING_BANKS["31"],
            "source_type": "staff_payment",
            "payment_status": ",".join(sorted(group["payment_statuses"])),
            "review_status": "",
            "review_reason": "",
        })
    for row in client_rows:
        receivable = _as_decimal(row["amount"])
        if receivable <= 0:
            continue
        refunded = _as_decimal(row.get("subsidy_return_refunded"))
        if refunded <= 0:
            payment_status = "pending"
        elif refunded < receivable:
            payment_status = "partially_paid"
        else:
            payment_status = "paid"
        rows.append({
            **row,
            "amount": receivable,
            "identity_card": "",
            "outgoing_bank_code": "633",
            "outgoing_bank_name": OUTGOING_BANKS["633"],
            "source_type": "client_subsidy_return",
            "payment_status": payment_status,
            "review_status": row.get("subsidy_return_review_status") or "",
            "review_reason": row.get("subsidy_return_review_reason") or "",
        })
    rows.sort(key=lambda row: (
        row["outgoing_bank_code"],
        _to_date(row["transfer_date"]) or date.max,
        row["source_payment_id"],
    ))
    return rows


def _build_workbook(payable_rows: list[dict], bank_totals: dict[str, Decimal]) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "應付匯款清單"
    worksheet.append(EXPORT_HEADERS)

    yellow_fill = PatternFill(fill_type="solid", fgColor="FFFF00")
    for cell in worksheet[1]:
        cell.fill = yellow_fill
        cell.font = Font(bold=True)

    for row in payable_rows:
        worksheet.append(tuple(row[header] for header in EXPORT_HEADERS))

    worksheet.append(())
    for code in ("31", "633"):
        worksheet.append((f"{OUTGOING_BANKS[code]}總額", "", "", "", "", float(bank_totals[code])))

    worksheet.freeze_panes = "A2"
    worksheet.column_dimensions["A"].width = 24
    worksheet.column_dimensions["B"].width = 14
    worksheet.column_dimensions["C"].width = 22
    worksheet.column_dimensions["D"].width = 22
    worksheet.column_dimensions["E"].width = 16
    worksheet.column_dimensions["F"].width = 14
    worksheet.column_dimensions["G"].width = 30
    worksheet.column_dimensions["H"].width = 16
    worksheet.column_dimensions["I"].width = 14

    for row_index in range(2, 2 + len(payable_rows)):
        worksheet.cell(row=row_index, column=4).number_format = "@"
        worksheet.cell(row=row_index, column=5).number_format = "@"
        worksheet.cell(row=row_index, column=9).number_format = "yyyy-mm-dd"

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_accounts_payable_export(target_month: str) -> dict:
    """Build a read-only monthly payable list and in-memory XLSX workbook."""
    source_rows = _fetch_payables(target_month)
    serials = {"31": 0, "633": 0}
    bank_totals = {"31": Decimal("0"), "633": Decimal("0")}
    payable_rows = []

    for source_row in source_rows:
        bank_code = source_row["outgoing_bank_code"]
        serials[bank_code] += 1
        amount = _as_decimal(source_row["amount"])
        bank_totals[bank_code] += amount
        payable_rows.append({
            EXPORT_HEADERS[0]: f"{int(target_month[5:7])}-{bank_code}-{serials[bank_code]}",
            EXPORT_HEADERS[1]: source_row["outgoing_bank_name"],
            EXPORT_HEADERS[2]: source_row.get("recipient_name") or "",
            EXPORT_HEADERS[3]: str(source_row.get("account_no") or ""),
            EXPORT_HEADERS[4]: str(source_row.get("recipient_bank_code") or ""),
            EXPORT_HEADERS[5]: amount,
            EXPORT_HEADERS[6]: source_row.get("identity_card") or "",
            EXPORT_HEADERS[7]: str(source_row["case_no"]),
            EXPORT_HEADERS[8]: source_row["transfer_date"],
            "預定付款／退款日期": source_row["transfer_date"],
            "付款狀態": source_row.get("payment_status") or "",
            "覆核狀態": source_row.get("review_status") or "",
            "覆核原因": source_row.get("review_reason") or "",
        })

    return {
        "payable_rows": payable_rows,
        "xlsx_bytes": _build_workbook(payable_rows, bank_totals),
        "bank_totals": bank_totals,
    }


def build_completed_order_payables_summary(target_month: str) -> dict:
    """Build a read-only summary for completed cases in a specific completion month."""
    first_day, last_day = _month_bounds(target_month)
    conn = get_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute(
                """
                SELECT o.case_no,
                       c.name AS client_name,
                       COALESCE(
                           o.actual_end_date,
                           CASE
                               WHEN o.actual_start_date IS NOT NULL AND o.service_days IS NOT NULL
                               THEN DATE_ADD(o.actual_start_date, INTERVAL o.service_days - 1 DAY)
                               ELSE NULL
                           END
                       ) AS completion_date,
                       o.actual_start_date,
                       o.service_days,
                       c.identity_status,
                       COALESCE(sp_agg.total_payable, 0) AS payable_salary,
                       COALESCE(sp_agg.amount_paid, 0) AS paid_salary,
                       cp.subsidy_return_at,
                       cp.subsidy_return_receivable,
                       cp.subsidy_return_refunded
                FROM orders o
                JOIN clients c ON c.id = o.client_id
                LEFT JOIN client_payments cp ON cp.case_no = o.case_no
                LEFT JOIN (
                    SELECT case_no,
                           SUM(total_payable) AS total_payable,
                           SUM(amount_paid) AS amount_paid
                    FROM staff_payments
                    WHERE payment_status <> 'cancelled'
                    GROUP BY case_no
                ) sp_agg ON sp_agg.case_no = o.case_no
                WHERE o.status = '訂單完成'
                  AND COALESCE(
                      o.actual_end_date,
                      CASE
                          WHEN o.actual_start_date IS NOT NULL AND o.service_days IS NOT NULL
                          THEN DATE_ADD(o.actual_start_date, INTERVAL o.service_days - 1 DAY)
                          ELSE NULL
                      END
                  ) BETWEEN %s AND %s
                ORDER BY o.case_no
                """,
                (first_day, last_day),
            )
            rows = cursor.fetchall()
    finally:
        conn.close()

    summary_rows = []
    total_payable_salary = Decimal("0")
    total_paid_salary = Decimal("0")
    total_salary_outstanding = Decimal("0")
    total_subsidy_receivable = Decimal("0")
    total_subsidy_refunded = Decimal("0")
    total_subsidy_remaining = Decimal("0")

    for row in rows:
        payable_salary = _as_decimal(row["payable_salary"])
        paid_salary = _as_decimal(row["paid_salary"])
        subsidy_receivable = _as_decimal(row["subsidy_return_receivable"])
        subsidy_refunded = _as_decimal(row["subsidy_return_refunded"])
        salary_outstanding = payable_salary - paid_salary
        subsidy_remaining = subsidy_receivable - subsidy_refunded

        summary_rows.append({
            ACCOUNTS_PAYABLE_SUMMARY_HEADERS[0]: row["case_no"],
            ACCOUNTS_PAYABLE_SUMMARY_HEADERS[1]: row["client_name"] or "",
            ACCOUNTS_PAYABLE_SUMMARY_HEADERS[2]: row["completion_date"],
            ACCOUNTS_PAYABLE_SUMMARY_HEADERS[3]: _derive_staff_payment_date(row),
            ACCOUNTS_PAYABLE_SUMMARY_HEADERS[4]: _derive_subsidy_refund_date(row),
            ACCOUNTS_PAYABLE_SUMMARY_HEADERS[5]: payable_salary,
            ACCOUNTS_PAYABLE_SUMMARY_HEADERS[6]: paid_salary,
            ACCOUNTS_PAYABLE_SUMMARY_HEADERS[7]: salary_outstanding,
            ACCOUNTS_PAYABLE_SUMMARY_HEADERS[8]: subsidy_receivable,
            ACCOUNTS_PAYABLE_SUMMARY_HEADERS[9]: subsidy_refunded,
            ACCOUNTS_PAYABLE_SUMMARY_HEADERS[10]: subsidy_remaining,
        })

        total_payable_salary += payable_salary
        total_paid_salary += paid_salary
        total_salary_outstanding += salary_outstanding
        total_subsidy_receivable += subsidy_receivable
        total_subsidy_refunded += subsidy_refunded
        total_subsidy_remaining += subsidy_remaining

    return {
        "summary_rows": summary_rows,
        "totals": {
            "payable_salary": total_payable_salary,
            "paid_salary": total_paid_salary,
            "salary_outstanding": total_salary_outstanding,
            "subsidy_receivable": total_subsidy_receivable,
            "subsidy_refunded": total_subsidy_refunded,
            "subsidy_remaining": total_subsidy_remaining,
        },
    }
