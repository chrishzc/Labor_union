"""Read-only subsidy reconciliation registers and annual summaries."""

from __future__ import annotations

import json
from datetime import date, datetime
from decimal import Decimal, ROUND_HALF_UP
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from services.db_service import get_connection


QUARTERLY_HEADERS = (
    "\u5e8f\u865f", "\u5e02\u5e9c\u8a02\u55ae\u865f\u78bc", "\u88dc\u52a9\u8cc7\u683c", "\u670d\u52d9\u958b\u59cb", "\u670d\u52d9\u7d50\u675f",
    "\u88dc\u52a9\u6642\u6578", "\u88dc\u52a9\u5929\u6578", "\u670d\u52d9\u5929\u6578", "\u88dc\u52a9\u6b3e\u91d1\u984d", "\u55ae\u50f9",
    "\u96c7\u4e3b", "\u670d\u52d9\u4eba\u54e1", "\u8eab\u5206\u8b49\u5b57\u865f", "\u5730\u5740", "\u7c3d\u9818",
)
ANNUAL_HEADERS = (
    "\u5e8f\u865f", "\u5e02\u5e9c\u8a02\u55ae\u865f\u78bc", "\u88dc\u52a9\u8cc7\u683c", "\u670d\u52d9\u958b\u59cb", "\u670d\u52d9\u7d50\u675f",
    "\u670d\u52d9\u5929\u6578", "\u88dc\u52a9\u6b3e\u91d1\u984d", "\u55ae\u50f9", "\u96c7\u4e3b", "\u670d\u52d9\u4eba\u54e1",
)

GENERAL_CITIZEN = "\u4e00\u822c\u5e02\u6c11"
SUBSIDIZED_CITIZEN = "\u88dc\u52a9\u5e02\u6c11"
IDENTITY_CARD_KEY = "\u8eab\u5206\u8b49\u5b57\u865f"


def _as_date(value) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        try:
            return datetime.strptime(value[:10], "%Y-%m-%d").date()
        except ValueError:
            return None
    return None


def _validate_year_and_quarter(application_year: int, quarter: int) -> None:
    if not isinstance(application_year, int) or application_year < 1912:
        raise ValueError("application_year must be a Gregorian year")
    if quarter not in (1, 2, 3, 4):
        raise ValueError("quarter must be 1, 2, 3, or 4")


def _completion_period(application_year: int, claim_quarter: int) -> tuple[int, int, int]:
    """Return the completion-year months for the selected reconciliation quarter."""
    _validate_year_and_quarter(application_year, claim_quarter)
    start_month = (claim_quarter - 1) * 3 + 1
    return application_year, start_month, start_month + 2


def _decode_legacy_key(key: object) -> str:
    text = str(key)
    for encoding in ("latin1", "cp1252"):
        try:
            decoded = text.encode(encoding).decode("utf-8")
        except (UnicodeDecodeError, UnicodeEncodeError):
            continue
        if IDENTITY_CARD_KEY in decoded:
            return decoded
    return text


def extract_employer_identity_card(survey_details: object) -> str:
    """Read the employer identity card from current or legacy-encoded survey JSON."""
    if isinstance(survey_details, str):
        try:
            survey_details = json.loads(survey_details)
        except json.JSONDecodeError:
            return ""
    if not isinstance(survey_details, dict):
        return ""
    for key, value in survey_details.items():
        if IDENTITY_CARD_KEY in _decode_legacy_key(key):
            return str(value or "").strip()
    return ""


def _fetch_completed_cases() -> list[dict]:
    conn = get_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute(
                """
                SELECT o.case_no, c.identity_status, o.actual_start_date,
                       o.actual_end_date, o.service_days, o.service_hours_per_day,
                       c.name AS employer_name, c.address AS employer_address,
                       s.name AS staff_name, br.survey_details
                FROM orders o
                JOIN clients c ON c.id = o.client_id
                LEFT JOIN staff s ON s.id = o.staff_id
                LEFT JOIN beclass_records br ON br.query_no = o.case_no
                WHERE o.actual_end_date IS NOT NULL
                  AND c.identity_status IN (%s, %s)
                ORDER BY o.case_no
                """,
                (GENERAL_CITIZEN, SUBSIDIZED_CITIZEN),
            )
            return cursor.fetchall()
    finally:
        conn.close()


def _subsidy_terms(eligibility: str, total_service_hours: Decimal) -> tuple[Decimal, Decimal]:
    if eligibility == GENERAL_CITIZEN:
        return min(Decimal("40"), total_service_hours), Decimal("300")
    if eligibility == SUBSIDIZED_CITIZEN:
        return min(Decimal("120"), total_service_hours), Decimal("350")
    return Decimal("0"), Decimal("0")


def _to_register_row(source: dict) -> dict | None:
    actual_start = _as_date(source.get("actual_start_date"))
    actual_end = _as_date(source.get("actual_end_date"))
    daily_hours = Decimal(str(source.get("service_hours_per_day") or 0))
    service_days = Decimal(str(source.get("service_days") or 0))
    subsidy_hours, unit_price = _subsidy_terms(
        source.get("identity_status"), service_days * daily_hours,
    )
    if not actual_start or not actual_end or subsidy_hours <= 0 or daily_hours <= 0:
        return None

    return {
        "\u5e02\u5e9c\u8a02\u55ae\u865f\u78bc": str(source["case_no"]),
        "\u88dc\u52a9\u8cc7\u683c": source["identity_status"],
        "\u670d\u52d9\u958b\u59cb": actual_start,
        "\u670d\u52d9\u7d50\u675f": actual_end,
        "\u88dc\u52a9\u6642\u6578": subsidy_hours,
        "\u88dc\u52a9\u5929\u6578": (subsidy_hours / daily_hours).quantize(
            Decimal("0.01"), rounding=ROUND_HALF_UP
        ),
        "\u670d\u52d9\u5929\u6578": source.get("service_days") or 0,
        "\u88dc\u52a9\u6b3e\u91d1\u984d": subsidy_hours * unit_price,
        "\u55ae\u50f9": unit_price,
        "\u96c7\u4e3b": source.get("employer_name") or "",
        "\u670d\u52d9\u4eba\u54e1": source.get("staff_name") or "",
        "\u8eab\u5206\u8b49\u5b57\u865f": extract_employer_identity_card(source.get("survey_details")),
        "\u5730\u5740": source.get("employer_address") or "",
        "\u7c3d\u9818": "",
    }


def _partition_rows(rows: list[dict]) -> tuple[list[dict], list[dict]]:
    general = []
    subsidized = []
    for row in rows:
        (subsidized if row["\u88dc\u52a9\u8cc7\u683c"] == SUBSIDIZED_CITIZEN else general).append(row)
    return general, subsidized


def _with_serials(rows: list[dict]) -> list[dict]:
    return [{"\u5e8f\u865f": index, **row} for index, row in enumerate(rows, start=1)]


def _build_workbook(headers: tuple[str, ...], general_rows: list[dict], subsidized_rows: list[dict], title: str) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = title
    yellow_fill = PatternFill(fill_type="solid", fgColor="FFFF00")

    def append_section(rows: list[dict], label: str) -> None:
        worksheet.append((label,))
        worksheet.append(headers)
        header_row = worksheet.max_row
        for cell in worksheet[header_row]:
            cell.fill = yellow_fill
            cell.font = Font(bold=True)
        for row in rows:
            worksheet.append(tuple(row.get(header, "") for header in headers))
            if "\u88dc\u52a9\u5929\u6578" in headers:
                worksheet.cell(worksheet.max_row, headers.index("\u88dc\u52a9\u5929\u6578") + 1).number_format = "0.00"

    append_section(general_rows, GENERAL_CITIZEN)
    if subsidized_rows:
        worksheet.append(())
        append_section(subsidized_rows, SUBSIDIZED_CITIZEN)

    worksheet.freeze_panes = "A3"
    for column in ("D", "E"):
        worksheet.column_dimensions[column].width = 14
    for column in ("B", "K", "L", "M", "N", "O"):
        worksheet.column_dimensions[column].width = 18
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _filtered_rows(application_year: int, claim_quarter: int | None) -> tuple[list[dict], list[dict]]:
    if claim_quarter is not None:
        completion_year, start_month, end_month = _completion_period(application_year, claim_quarter)
    rows = []
    for source in _fetch_completed_cases():
        row = _to_register_row(source)
        if not row:
            continue
        completion = row["\u670d\u52d9\u7d50\u675f"]
        if claim_quarter is not None:
            if completion.year != completion_year or not start_month <= completion.month <= end_month:
                continue
        elif completion.year != application_year:
            continue
        rows.append(row)
    rows.sort(key=lambda row: row["\u5e02\u5e9c\u8a02\u55ae\u865f\u78bc"])
    general, subsidized = _partition_rows(rows)
    return _with_serials(general), _with_serials(subsidized)


def build_quarterly_subsidy_register(application_year: int, quarter: int) -> dict:
    """Build a register for the selected completion-year quarter."""
    _validate_year_and_quarter(application_year, quarter)
    general_rows, subsidized_rows = _filtered_rows(application_year, quarter)
    return {
        "general_citizen_rows": general_rows,
        "subsidized_citizen_rows": subsidized_rows,
        "xlsx_bytes": _build_workbook(QUARTERLY_HEADERS, general_rows, subsidized_rows, "\u5206\u5b63\u6838\u92b7"),
    }


def build_annual_subsidy_summary(application_year: int) -> dict:
    """Build a completion-year annual summary, separated by subsidy eligibility."""
    if not isinstance(application_year, int) or application_year < 1912:
        raise ValueError("application_year must be a Gregorian year")
    general_rows, subsidized_rows = _filtered_rows(application_year, None)
    return {
        "general_citizen_rows": general_rows,
        "subsidized_citizen_rows": subsidized_rows,
        "xlsx_bytes": _build_workbook(ANNUAL_HEADERS, general_rows, subsidized_rows, "\u5e74\u5ea6\u7e3d\u8868"),
    }
