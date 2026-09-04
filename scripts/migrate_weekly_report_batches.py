"""
File: migrate_weekly_report_batches.py
Description: 營運週報結算批次與案件封存表 (方案 C) 之資料庫升級與初始資料匯入腳本。
執行方式: python scripts/migrate_weekly_report_batches.py
"""

from __future__ import annotations

import os
import sys
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

import openpyxl
from infrastructure.mysql.mysql_adapter import get_connection

DDL_SQL = """
CREATE TABLE IF NOT EXISTS `weekly_report_batches` (
    `id` INT AUTO_INCREMENT PRIMARY KEY,
    `year` INT NOT NULL,
    `week_code` VARCHAR(32) NOT NULL,
    `cutoff_at` DATETIME NOT NULL,
    `promotion_count` INT NOT NULL DEFAULT 0,
    `inquiry_count` INT NOT NULL DEFAULT 0,
    `notes` TEXT NULL,
    `created_at` DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
    `updated_at` DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
    UNIQUE KEY `uk_year_week` (`year`, `week_code`),
    INDEX `idx_year` (`year`)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci;

CREATE TABLE IF NOT EXISTS `weekly_report_batch_cases` (
    `case_no` VARCHAR(64) NOT NULL PRIMARY KEY,
    `batch_id` INT NOT NULL,
    `bound_at` DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
    INDEX `idx_batch_id` (`batch_id`),
    CONSTRAINT `fk_batch_cases_batch` FOREIGN KEY (`batch_id`) REFERENCES `weekly_report_batches` (`id`) ON DELETE CASCADE
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci;
"""


def seed_history_from_template(conn, year: int = 2026) -> int:
    template_path = PROJECT_ROOT / "document" / "管理端UI" / "表格需求模板" / "週報.xlsx"
    if not template_path.exists():
        print(f"[WARN] 模板檔案不存在: {template_path}，跳過歷史資料種子初始化。")
        return 0

    wb = openpyxl.load_workbook(template_path, data_only=True)
    ws = wb.worksheets[0]

    # 解析第一分頁中的歷史週次與指標
    weekly_metrics: dict[str, tuple[int, int]] = {}
    for r in range(5, ws.max_row + 1):
        w_code = ws.cell(row=r, column=4).value
        p_val = ws.cell(row=r, column=6).value
        i_val = ws.cell(row=r, column=7).value
        if w_code and isinstance(w_code, str) and "-" in w_code:
            w_str = w_code.strip()
            p_int = int(p_val) if p_val is not None and str(p_val).isdigit() else 0
            i_int = int(i_val) if i_val is not None and str(i_val).isdigit() else 0
            if w_str not in weekly_metrics or (p_int > 0 or i_int > 0):
                weekly_metrics[w_str] = (p_int, i_int)

    inserted = 0
    with conn.cursor() as cur:
        for w_code, (p_count, i_count) in sorted(weekly_metrics.items()):
            cur.execute(
                """
                INSERT IGNORE INTO weekly_report_batches 
                (year, week_code, cutoff_at, promotion_count, inquiry_count, notes)
                VALUES (%s, %s, %s, %s, %s, %s)
                """,
                (
                    year,
                    w_code,
                    f"{year}-06-30 12:00:00",
                    p_count,
                    i_count,
                    "從 Excel 模板歷史資料初始化",
                ),
            )
            inserted += cur.rowcount
    conn.commit()
    return inserted


def main() -> int:
    print("=" * 60)
    print("🚀 開始執行週報批次管理資料庫升級 (weekly_report_batches)")
    print("=" * 60)

    try:
        conn = get_connection()
    except Exception as e:
        print(f"[ERROR] 無法連線至 MySQL 資料庫: {e}", file=sys.stderr)
        return 1

    try:
        with conn.cursor() as cur:
            print("[INFO] 正在建立 weekly_report_batches 與 weekly_report_batch_cases 表格...")
            for statement in DDL_SQL.strip().split(";"):
                stmt = statement.strip()
                if stmt:
                    cur.execute(stmt)
        conn.commit()
        print("[SUCCESS] 資料表建立/驗證完成！")

        # 檢查並初始化歷史資料
        with conn.cursor() as cur:
            cur.execute("SELECT COUNT(*) AS cnt FROM weekly_report_batches WHERE year = 2026")
            row = cur.fetchone()
            count = row["cnt"] if isinstance(row, dict) else row[0]

        if count == 0:
            print("[INFO] 偵測到 2026 年度尚無歷史週次，正在自 Excel 模板匯入歷史指標...")
            imported = seed_history_from_template(conn, year=2026)
            print(f"[SUCCESS] 成功初始化匯入 {imported} 週歷史指標！")
        else:
            print(f"[INFO] 2026 年度既有批次筆數: {count}，維持既有資料不覆蓋。")

        print("=" * 60)
        print("✅ 週報批次資料庫升級腳本執行完畢！")
        print("=" * 60)
        return 0
    except Exception as e:
        print(f"[ERROR] 升級腳本執行失敗: {e}", file=sys.stderr)
        return 1
    finally:
        conn.close()


if __name__ == "__main__":
    sys.exit(main())
