import os
import sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

excel_path = os.path.join("db", "templates", "contracts", "contract_client_copy.xlsx")
wb = openpyxl.load_workbook(excel_path, data_only=False)
ws = wb.active

print(f"📊 檢測 [{excel_path}] 裡面原生勾選 wrap_text (自動換行) 的儲存格：\n")

wrap_cells = []
for r in range(1, ws.max_row + 1):
    for c in range(1, min(ws.max_column + 1, 10)):
        cell = ws.cell(row=r, column=c)
        if cell.alignment and cell.alignment.wrap_text:
            val_str = str(cell.value).strip() if cell.value is not None else ""
            wrap_cells.append((cell.coordinate, val_str[:25]))

print(f"✅ 共發現 {len(wrap_cells)} 個原生勾選自動換行的儲存格/區域：")
for coord, txt in wrap_cells:
    print(f"   座標 [{coord:5s}]: '{txt}'")
