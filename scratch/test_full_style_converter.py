import os
import sys
import openpyxl
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')

excel_path = os.path.join("db", "templates", "contracts", "contract_client_copy.xlsx")
wb = openpyxl.load_workbook(excel_path, data_only=True)
ws = wb.active

def argb_to_hex(color_obj):
    if not color_obj:
        return None
    rgb = getattr(color_obj, 'rgb', None)
    if not rgb:
        return None
    rgb_str = str(rgb).upper()
    if len(rgb_str) == 8: # AARRGGBB
        return f"#{rgb_str[2:]}"
    elif len(rgb_str) == 6: # RRGGBB
        return f"#{rgb_str}"
    return None

print(f"📊 解析 [{excel_path}] 樣式與欄寬中...")
col_widths = {}
total_width = 0
for c in range(1, min(ws.max_column + 1, 10)):
    col_letter = get_column_letter(c)
    w = ws.column_dimensions[col_letter].width
    w_val = float(w) if w else 12.0
    col_widths[c] = w_val
    total_width += w_val

print("欄位寬度比例:", col_widths)

print("\n前 10 行儲存格樣式範例:")
for r in range(1, 12):
    row_str = []
    for c in range(1, min(ws.max_column + 1, 10)):
        cell = ws.cell(row=r, column=c)
        val = str(cell.value).strip() if cell.value is not None else ""
        bg = argb_to_hex(cell.fill.start_color) if cell.fill else None
        fg = argb_to_hex(cell.font.color) if cell.font else None
        bold = cell.font.bold if cell.font else False
        align = cell.alignment.horizontal if cell.alignment else None
        
        row_str.append(f"[{get_column_letter(c)}{r}: '{val[:8]}' bg={bg} fg={fg} B={bold} A={align}]")
    print(" ".join(row_str[:4]))
