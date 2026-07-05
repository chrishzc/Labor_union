import os
import sys
import openpyxl
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')

excel_path = os.path.join("db", "templates", "contracts", "contract_client_copy.xlsx")
wb = openpyxl.load_workbook(excel_path, data_only=True)
ws = wb.active

def get_html_hex_color(color_obj, default=None):
    if not color_obj:
        return default
    rgb = getattr(color_obj, 'rgb', None)
    if not rgb:
        return default
    rgb_str = str(rgb).upper()
    if len(rgb_str) == 8: # AARRGGBB
        hex_val = rgb_str[2:]
        if hex_val == "000000": # 透明或無色
            return default
        return f"#{hex_val}"
    elif len(rgb_str) == 6:
        return f"#{rgb_str}"
    return default

# 1. 欄寬計算
col_widths = {}
total_w = 0
max_cols = min(ws.max_column, 10)

for c in range(1, max_cols + 1):
    col_letter = get_column_letter(c)
    w = ws.column_dimensions[col_letter].width
    w_val = float(w) if w else 12.0
    col_widths[c] = w_val
    total_w += w_val

col_styles = ""
for c in range(1, max_cols + 1):
    pct = round((col_widths[c] / total_w) * 100, 2)
    col_styles += f"<col style='width: {pct}%;'>\n"

print("✅ 生成 colgroup 欄寬 HTML:")
print(col_styles)
