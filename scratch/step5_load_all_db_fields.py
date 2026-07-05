import os
import sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

db_source_file = "document/資料庫、資料處理/資料庫來源表.xlsx"
wb = openpyxl.load_workbook(db_source_file, data_only=True)

print("==========================================================")
print("📊 掃描資料庫來源表.xlsx 中的全量欄位 (Full Database Schema Scan)")
print("==========================================================\n")

full_schema = {}

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    # 取前兩列標題
    headers = []
    for r in range(1, 3):
        for c in range(1, ws.max_column + 1):
            val = str(ws.cell(row=r, column=c).value).strip() if ws.cell(row=r, column=c).value else ""
            if val and val != "None" and not val.isdigit() and val not in headers:
                headers.append(val)
    full_schema[sheet_name] = headers
    print(f"📋 資料表 [{sheet_name}] (共 {len(headers)} 個欄位):")
    for idx, h in enumerate(headers, 1):
        print(f"   {idx:02d}. {h}")
    print()

