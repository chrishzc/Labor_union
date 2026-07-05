import os
import sys
import openpyxl
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')

excel_path = os.path.join("db", "templates", "contracts", "contract_client_copy.xlsx")
wb = openpyxl.load_workbook(excel_path, data_only=False)
ws = wb.active

def is_yellow_fill(cell):
    if not cell.fill or not cell.fill.fill_type:
        return False
    fill_color = getattr(cell.fill.start_color, 'rgb', None) or getattr(cell.fill.fgColor, 'rgb', None)
    if not fill_color:
        return False
    fill_color_str = str(fill_color).upper()
    yellow_patterns = ["FFFF00", "FFFFCC", "FFF2CC", "FFE599", "FFFF99", "FFFF0000"]
    return any(p in fill_color_str for p in yellow_patterns)

yellow_cells = []
for r in range(1, ws.max_row + 1):
    for c in range(1, min(ws.max_column + 1, 10)):
        cell = ws.cell(row=r, column=c)
        val_str = str(cell.value).strip() if cell.value is not None else ""
        
        # 過濾黃底儲存格，且排除空文字或純公式無定義者
        if is_yellow_fill(cell) or ("(" in val_str and ")" in val_str and ("市府" in val_str or "訂單" in val_str or "BECLASS" in val_str or "綁定" in val_str)):
            # 去除極為明顯的非填空單元格
            if val_val := val_str:
                yellow_cells.append({
                    "coord": cell.coordinate,
                    "row": r,
                    "col": c,
                    "orig_text": val_val
                })

# 去除完全重複之座標與表頭類
dedup_cells = []
seen = set()
for item in yellow_cells:
    coord = item["coord"]
    if coord not in seen and item["orig_text"] not in ["款項", "金額", "匯款日期", "新竹市月子照顧服務人員職業工會"]:
        seen.add(coord)
        dedup_cells.append(item)

print(f"📊 經全面審查，共篩選出 {len(dedup_cells)} 個需連動之黃底填空點：\n")

# 編列 P1 ~ PN 標籤清單
param_list = []
for idx, item in enumerate(dedup_cells, 1):
    p_tag = f"P{idx}"
    item["param_tag"] = p_tag
    param_list.append(item)
    print(f"{p_tag:4s} | 座標: {item['coord']:5s} | 原始內容: '{item['orig_text']}'")

# 輸出為便利對照的 json 檔與對照清單
import json
with open("scratch/param_list_step3.json", "w", encoding="utf-8") as f:
    json.dump(param_list, f, ensure_ascii=False, indent=2)

print("\n✅ 已將 P1~PN 完整編號列表匯出至 scratch/param_list_step3.json")
