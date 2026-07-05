import os
import sys
import openpyxl
import glob

sys.stdout.reconfigure(encoding='utf-8')

print("==========================================================")
print("🔍 ADAD 優先任務資料庫與欄位歸屬深度稽核 (Priority Task Audit)")
print("==========================================================\n")

# 1. 尋找資料庫來源表.xlsx 與 假資料_範例.xlsx
db_source_file = "document/資料庫、資料處理/資料庫來源表.xlsx"
fake_data_file = "document/資料庫、資料處理/假資料_範例.xlsx"

# 稽核 任務 6 & 7 & 10: 檢查資料庫來源表.xlsx 中的所有 Sheet 與標題列 (Header Columns)
if os.path.exists(db_source_file):
    print(f"📂 開啟資料庫來源表: {db_source_file}")
    wb = openpyxl.load_workbook(db_source_file, data_only=True)
    print("  現有 Sheets:", wb.sheetnames)
    for sname in wb.sheetnames:
        ws = wb[sname]
        headers = [str(ws.cell(row=r, column=c).value).strip() for r in range(1, 4) for c in range(1, ws.max_column+1) if ws.cell(row=r, column=c).value]
        # 過濾不重複
        clean_headers = list(dict.fromkeys([h for h in headers if h and h != "None"]))
        print(f"\n  📋 分頁 [{sname}] 欄位一覽 (前 25 個):")
        print("     ", clean_headers[:25])

if os.path.exists(fake_data_file):
    print(f"\n📂 開啟假資料_範例: {fake_data_file}")
    wb_fake = openpyxl.load_workbook(fake_data_file, data_only=True)
    print("  現有 Sheets:", wb_fake.sheetnames)
    for sname in wb_fake.sheetnames:
        ws = wb_fake[sname]
        headers = [str(ws.cell(row=1, column=c).value).strip() for c in range(1, ws.max_column+1) if ws.cell(row=1, column=c).value]
        print(f"\n  📋 分頁 [{sname}] 第一列標題 (Header Columns):")
        print("     ", headers)

