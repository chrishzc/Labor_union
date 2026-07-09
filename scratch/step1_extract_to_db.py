import os
import sys
import glob
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

excel_path = None
for f in glob.glob("document/**/*.xlsx", recursive=True):
    if "所需表格.xlsx" in f and not os.path.basename(f).startswith("~$"):
        excel_path = f
        break

if not excel_path:
    print("❌ 未找到 所需表格.xlsx 檔案！")
    sys.exit(1)

print(f"📂 讀取 Excel 檔案: {excel_path}")
wb_src = openpyxl.load_workbook(excel_path)

if "客戶契約" not in wb_src.sheetnames:
    print("❌ 未找到 '客戶契約' 分頁！")
    sys.exit(1)

# 1. 建立獨立的 db/templates/contracts/contract_client_copy.xlsx
target_dir = os.path.join("db", "templates", "contracts")
os.makedirs(target_dir, exist_ok=True)
standalone_path = os.path.join(target_dir, "contract_client_copy.xlsx")

# 建立全新只有單一 Sheet 的 Workbook
wb_new = openpyxl.Workbook()
ws_new = wb_new.active
ws_new.title = "客戶契約"

ws_src = wb_src["客戶契約"]

# 複製內容、格式、寬度與儲存格合併
for row in ws_src.iter_rows():
    for cell in row:
        new_cell = ws_new.cell(row=cell.row, column=cell.column, value=cell.value)
        if cell.has_style:
            new_cell.font = openpyxl.styles.Font(
                name=cell.font.name, size=cell.font.size, bold=cell.font.bold,
                italic=cell.font.italic, color=cell.font.color
            )
            new_cell.border = openpyxl.styles.Border(
                left=cell.border.left, right=cell.border.right,
                top=cell.border.top, bottom=cell.border.bottom
            )
            if cell.fill and cell.fill.fill_type:
                new_cell.fill = openpyxl.styles.PatternFill(
                    fill_type=cell.fill.fill_type,
                    start_color=cell.fill.start_color,
                    end_color=cell.fill.end_color
                )
            new_cell.number_format = cell.number_format
            new_cell.alignment = openpyxl.styles.Alignment(
                horizontal=cell.alignment.horizontal,
                vertical=cell.alignment.vertical,
                wrap_text=cell.alignment.wrap_text
            )

# 複製欄寬
for col in ws_src.column_dimensions:
    ws_new.column_dimensions[col].width = ws_src.column_dimensions[col].width

# 複製合併儲存格
for range_item in ws_src.merged_cells.ranges:
    ws_new.merge_cells(str(range_item))

wb_new.save(standalone_path)
print(f"✅ 成功將 '客戶契約' 抽出為獨立極速範本檔: {standalone_path}")

# 2. 清理原始 所需表格.xlsx 中的臨時 '客戶契約(複製版)' 分頁，保持原始檔 100% 乾淨
if "客戶契約(複製版)" in wb_src.sheetnames:
    del wb_src["客戶契約(複製版)"]
    wb_src.save(excel_path)
    print(f"🧹 已自動清理 '所需表格.xlsx' 中的臨時測試分頁，原始檔案恢復 100% 乾淨！")
