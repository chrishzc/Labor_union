import os
import sys
import glob
import openpyxl
from openpyxl.styles import PatternFill

sys.stdout.reconfigure(encoding='utf-8')

# 尋找所需表格.xlsx 檔案
excel_path = None
for f in glob.glob("document/**/*.xlsx", recursive=True):
    if "所需表格.xlsx" in f and not os.path.basename(f).startswith("~$"):
        excel_path = f
        break

if not excel_path:
    print("❌ 未找到 所需表格.xlsx 檔案！")
    sys.exit(1)

print(f"📂 正在開啟 Excel 檔案: {excel_path}")
wb = openpyxl.load_workbook(excel_path)

if "客戶契約" not in wb.sheetnames:
    print(f"❌ 分頁 '客戶契約' 不存在，現有分頁: {wb.sheetnames}")
    sys.exit(1)

# 1. 如果已經有 '客戶契約(複製版)'，先刪除以重新複製
if "客戶契約(複製版)" in wb.sheetnames:
    del wb["客戶契約(複製版)"]

source_sheet = wb["客戶契約"]
target_sheet = wb.copy_worksheet(source_sheet)
target_sheet.title = "客戶契約(複製版)"

print(f"✅ 成功複製分頁為 '客戶契約(複製版)'！現有分頁: {wb.sheetnames}")

# 2. 檢測並掃描 '客戶契約(複製版)' 中所有黃底的欄位 (Yellow Fill Background)
yellow_cells = []

def is_yellow_fill(cell):
    if not cell.fill or not cell.fill.fill_type:
        return False
    fill_color = getattr(cell.fill.start_color, 'rgb', None) or getattr(cell.fill.fgColor, 'rgb', None)
    if not fill_color:
        return False
    fill_color_str = str(fill_color).upper()
    # 常見的黃色 RGB 色碼: FFFF00, FFFFFF00, FFFFCC, FFF2CC, FFE599, FFFF99 等
    yellow_patterns = ["FFFF00", "FFFFCC", "FFF2CC", "FFE599", "FFFF99", "FFFF0000"]
    return any(p in fill_color_str for p in yellow_patterns)

for row in target_sheet.iter_rows():
    for cell in row:
        if is_yellow_fill(cell) or (cell.value and "{" in str(cell.value)):
            yellow_cells.append({
                "coord": cell.coordinate,
                "value": str(cell.value).strip() if cell.value is not None else "",
                "color": cell.fill.start_color.rgb if cell.fill else "Unknown"
            })

print(f"\n🔍 掃描到的黃底需連動欄位總數: {len(yellow_cells)} 個")
for idx, item in enumerate(yellow_cells, 1):
    print(f"  {idx:02d}. 座標 [{item['coord']}]: 原本內容 -> '{item['value']}' (色碼: {item['color']})")

# 儲存 Excel 檔案
wb.save(excel_path)
print(f"\n🎉 步驟 1 完成！修改已安全儲存至檔案中。")
