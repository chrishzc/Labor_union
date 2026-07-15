"""
================================================================================
檔案名稱: ui/pages/05_form_management.py
功能說明: 表單管理與動態表單建立沙盒 (5:5 雙視窗 Side-by-Side + 拖拉排序 + 二次確認刪除)
專案名稱: Lobar Union - 服務人員與訂單管理系統
建立日期: 2026-07-04
架構規範: ADAD Version 33.1 (validated)
================================================================================
職責與業務規則:
1. 實施刪除二次確認防呆 Modal (INV-UI-FORM-08)。
2. 實施欄位上下順序平移/拖拉排序 (INV-UI-FORM-09)。
3. 實施 5:5 左右雙視窗 Side-by-Side 實時同步 Preview 視窗 (INV-UI-FORM-10)。
4. 實施 Draft Buffer 編輯草稿隔離機制，點擊取消時 100% 丟棄草稿 (INV-UI-FORM-06)。
5. 導覽約束: 本檔案末尾嚴禁包含頂層 show() 呼叫，由 ui/app.py 動態載入。
================================================================================
"""

import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
import json
import os
import time
from datetime import datetime, date
import math
import importlib
from urllib.error import HTTPError, URLError
from urllib.parse import urlencode
from urllib.request import urlopen
from services import db_service
importlib.reload(db_service)

import uuid

title = "📋 表單與履歷問卷管理"

JSON_TPL_PATH = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), "db", "form_templates.json")
TEMPLATES_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), "db", "templates")
CONTRACTS_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), "db", "templates", "contracts")
API_BASE_URL = os.getenv("API_BASE_URL", "http://localhost:8000").rstrip("/")


def fetch_staff_contract_context(case_no: str, assignment_id: int | None = None) -> dict:
    """Read staff-contract facts from FastAPI without writing the workbook."""
    query = urlencode({"assignment_id": assignment_id}) if assignment_id else ""
    url = f"{API_BASE_URL}/api/v1/contracts/staff/{case_no}"
    if query:
        url = f"{url}?{query}"
    try:
        with urlopen(url, timeout=5) as response:
            payload = json.loads(response.read().decode("utf-8"))
    except HTTPError as error:
        detail = error.read().decode("utf-8", errors="replace")
        raise ValueError(f"契約資料 API 回應 {error.code}: {detail}") from error
    except URLError as error:
        raise ValueError(f"無法連線契約資料 API ({API_BASE_URL}): {error.reason}") from error

    if not payload.get("success", False) or not isinstance(payload.get("data"), dict):
        raise ValueError(payload.get("error") or payload.get("message") or "契約資料 API 未回傳資料")
    return payload["data"]


def flatten_staff_contract_context(context: dict) -> dict:
    """Flatten the read-only API payload into existing Excel mapping keys."""
    order = context.get("order") or {}
    client = context.get("client") or {}
    assignment = context.get("assignment") or {}
    staff = context.get("staff") or {}

    flat = {
        **{key: value for key, value in order.items() if value is not None},
        **{key: value for key, value in assignment.items() if value is not None},
        "client_name": client.get("name"),
        "client_phone": client.get("phone"),
        "city": client.get("city"),
        "address": client.get("address"),
        "service_type": client.get("service_type"),
        "service_time": client.get("service_time"),
        "staff_name": staff.get("name"),
        "staff_phone": staff.get("phone"),
    }
    return {key: value for key, value in flat.items() if value is not None}

def generate_field_id() -> str:
    """產生全域絕對不重複的 field_id (INV-UI-FORM-07 防碰撞防線)"""
    return f"f_{uuid.uuid4().hex[:10]}_{int(time.time() * 1000)}"

import openpyxl
from openpyxl.utils import get_column_letter

def get_html_hex_color(color_obj, default=None):
    """將 openpyxl Color 物件轉為標準 HTML HEX 色碼 (#RRGGBB)"""
    if not color_obj:
        return default
    rgb = getattr(color_obj, 'rgb', None)
    if not rgb:
        return default
    rgb_str = str(rgb).upper()
    if len(rgb_str) == 8: # AARRGGBB
        hex_val = rgb_str[2:]
        if hex_val == "000000" or hex_val == "FFFFFF":
            return default
        return f"#{hex_val}"
    elif len(rgb_str) == 6:
        return f"#{rgb_str}"
    return default

def get_border_style(cell_border):
    """精確解析 Excel 單元格顯式設定之邊框，空白單元格不繪製多餘灰框，確保列印 PDF 時 100% 保留實心框線 (INV-UI-FORM-29)"""
    if not cell_border:
        return "border: none !important;"
    
    b_left = "1px solid #111111 !important;" if (cell_border.left and cell_border.left.style) else "none !important;"
    b_right = "1px solid #111111 !important;" if (cell_border.right and cell_border.right.style) else "none !important;"
    b_top = "1px solid #111111 !important;" if (cell_border.top and cell_border.top.style) else "none !important;"
    b_bottom = "1px solid #111111 !important;" if (cell_border.bottom and cell_border.bottom.style) else "none !important;"
    
    return f"border-left: {b_left} border-right: {b_right} border-top: {b_top} border-bottom: {b_bottom}"

def render_excel_contract_mirror(contract_config: dict, target_order: dict, global_stats: dict) -> str:
    """實時 1:1 解析與高精度鏡像渲染 Excel 實體範本檔 (含背景色、欄寬、粗體與原廠邊框)"""
    tpl_filename = contract_config.get('template_filename', 'contract_client_copy.xlsx')
    excel_file_path = os.path.join(CONTRACTS_DIR, tpl_filename)
    
    if not os.path.exists(excel_file_path):
        return f"<div style='color:red;'>❌ 未找到 Excel 範本檔案: {tpl_filename}</div>"

    wb = openpyxl.load_workbook(excel_file_path, data_only=True)
    ws = wb.active

    mappings = contract_config.get('param_mappings', {})

    # 1. 計算欄位寬度比例 (<colgroup>)
    max_cols = min(ws.max_column, 10)
    col_widths = {}
    total_w = 0
    for c in range(1, max_cols + 1):
        col_letter = get_column_letter(c)
        w = ws.column_dimensions[col_letter].width
        w_val = float(w) if w else 12.0
        col_widths[c] = w_val
        total_w += w_val

    colgroup_html = "<colgroup>\n"
    for c in range(1, max_cols + 1):
        pct = round((col_widths[c] / total_w) * 100, 2)
        colgroup_html += f"  <col style='width: {pct}%;'>\n"
    colgroup_html += "</colgroup>\n"

    # 2. 建立合併儲存格涵蓋集合
    merged_spans = {}
    for rng in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = rng.min_col, rng.min_row, rng.max_col, rng.max_row
        top_left_coord = get_column_letter(min_col) + str(min_row)
        colspan = max_col - min_col + 1
        rowspan = max_row - min_row + 1
        
        merged_spans[top_left_coord] = (colspan, rowspan)
        
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                cell_coord = get_column_letter(c) + str(r)
                if cell_coord != top_left_coord:
                    merged_spans[cell_coord] = None

    # 3. 逐列解析單元格內容與高精度 CSS 樣式
    table_rows_html = ""
    for r in range(1, min(ws.max_row + 1, 190)):
        row_html = ""
        for c in range(1, max_cols + 1):
            coord = get_column_letter(c) + str(r)
            
            if coord in merged_spans and merged_spans[coord] is None:
                continue

            cell = ws.cell(row=r, column=c)
            cell_val = str(cell.value).strip() if cell.value is not None else ""

            span_attr = ""
            if coord in merged_spans and merged_spans[coord] is not None:
                c_span, r_span = merged_spans[coord]
                if c_span > 1: span_attr += f" colspan='{c_span}'"
                if r_span > 1: span_attr += f" rowspan='{r_span}'"

            # 解析字體、顏色、粗體、背景色、對齊與邊框
            bg_color = get_html_hex_color(cell.fill.start_color if cell.fill else None, default="transparent")
            text_color = get_html_hex_color(cell.font.color if cell.font else None, default="#111111")
            is_bold = cell.font.bold if cell.font else False
            font_weight_css = "bold" if is_bold else "normal"
            border_css = get_border_style(cell.border)
            
            h_align = cell.alignment.horizontal if (cell.alignment and cell.alignment.horizontal) else "left"
            if h_align == "general": h_align = "left"

            # 4. 100% 原汁原味複製重現 Excel 原生 alignment.wrap_text 設定 (Native Excel Wrap Alignment)
            is_excel_wrap = bool(cell.alignment and cell.alignment.wrap_text)
            wrap_css = "white-space:normal !important; word-break:break-word !important;" if is_excel_wrap else "white-space:nowrap; overflow:visible;"

            # 檢查是否為映射連動欄位 (若成功抓到數據，背景自動清洗為乾淨白底/透明)
            if coord in mappings:
                p_info = mappings[coord]
                db_k = p_info.get('db_key', '')
                
                # 成功填入數據後，自動清掉黃底標記，呈獻乾淨紙本質感
                bg_color = "transparent" if target_order else (bg_color if bg_color != "transparent" else "#FFFF00")
                text_color = "#0D47A1" if not target_order else "#111111"
                font_weight_css = "bold"
                
                if p_info.get('status') == 'pending':
                    cell_val = "暫不連動"
                    bg_color = "#FFF3CD"
                    text_color = "#9C6500"
                elif db_k == '__today__':
                    cell_val = date.today().strftime('%Y-%m-%d')
                elif db_k in global_stats:
                    cell_val = format_db_value(db_k, global_stats[db_k])
                elif target_order and db_k in target_order:
                    cell_val = format_db_value(db_k, target_order.get(db_k, ''))
                elif not cell_val or cell_val == "None":
                    cell_val = f"[{p_info.get('label', coord)}]"

            style_str = f"background:{bg_color}; color:{text_color}; font-weight:{font_weight_css}; text-align:{h_align}; vertical-align:middle; padding:3px 6px; {border_css} font-size:12px; {wrap_css}"
            row_html += f"<td {span_attr} style='{style_str}'>{cell_val}</td>"
        
        table_rows_html += f"<tr>{row_html}</tr>"

    mirror_html = f"""
    <div id="excel-mirror-container" style="font-family:'Segoe UI', Microsoft JhengHei, sans-serif; padding:15px; background:#FAFAFA; border-radius:8px;">
        <style>
            @page {{
                size: A4 portrait;
                margin: 8mm 10mm;
            }}
            @media print {{
                .no-print {{ display: none !important; }}
                #excel-mirror-container {{ padding: 0 !important; background: white !important; border-radius: 0 !important; margin: 0 !important; }}
                body, html {{ background: white !important; margin: 0 !important; padding: 0 !important; }}
                table, td, th {{
                    -webkit-print-color-adjust: exact !important;
                    print-color-adjust: exact !important;
                    color-adjust: exact !important;
                }}
            }}
        </style>
        <div style="text-align:center; margin-bottom:10px; color:#1565C0; font-weight:bold; font-size:15px;" class="no-print">
            📊 《{contract_config.get('name')}》 1:1 原汁原味 Excel 樣式鏡像渲染視窗 (檔案: {tpl_filename})
            <div style="font-size:12px; color:#666; font-weight:normal; margin-top:4px;">
                💡 提示：點擊下方列印按鈕，另存 PDF 時可在預覽視窗右側取消勾選【頁首和頁尾】即可阻斷頂部日期與 localhost 網址！
            </div>
        </div>
        <table style="border-collapse:collapse; width:100%; table-layout:fixed; background:#FFFFFF; margin:auto;">
            {colgroup_html}
            <tbody>
                {table_rows_html}
            </tbody>
        </table>
        <div style="text-align:center; margin-top:15px;" class="no-print">
            <button onclick="window.print()" style="background-color:#2E7D32; color:white; border:none; padding:8px 20px; font-size:14px; font-weight:bold; border-radius:6px; cursor:pointer;">
                🖨️ 點擊另存為 PDF / 輸出列印 (Print / Save as PDF)
            </button>
        </div>
    </div>
    """
    return mirror_html

def load_contract_templates():
    """從 db/templates/contracts/ 讀取所有制式契約與變數代理設定 (INV-UI-FORM-16/17)"""
    os.makedirs(CONTRACTS_DIR, exist_ok=True)
    c_list = []
    json_files = [f for f in os.listdir(CONTRACTS_DIR) if f.endswith('.json')]
    for fname in sorted(json_files):
        fpath = os.path.join(CONTRACTS_DIR, fname)
        try:
            with open(fpath, "r", encoding="utf-8") as f:
                c_list.append(json.load(f))
        except Exception:
            pass
    return c_list

def save_contract_template(contract_data: dict):
    """持久化保存契約變數對照設定至 db/templates/contracts/*.json"""
    os.makedirs(CONTRACTS_DIR, exist_ok=True)
    cid = contract_data.get('id', 'contract_hsinchu_v1')
    fpath = os.path.join(CONTRACTS_DIR, f"{cid}.json")
    with open(fpath, "w", encoding="utf-8") as f:
        json.dump(contract_data, f, ensure_ascii=False, indent=2)

def load_json_templates():
    """從 db/templates/ 目錄讀取所有獨立模板 JSON 檔案 (INV-UI-FORM-13 獨立模組架構 + 去重防護)"""
    os.makedirs(TEMPLATES_DIR, exist_ok=True)
    tpls = []
    
    # 遍歷 db/templates/*.json
    json_files = [f for f in os.listdir(TEMPLATES_DIR) if f.endswith('.json')]
    if json_files:
        for fname in sorted(json_files):
            fpath = os.path.join(TEMPLATES_DIR, fname)
            try:
                with open(fpath, "r", encoding="utf-8") as f:
                    t = json.load(f)
                    seen_ids = set()
                    for idx, field in enumerate(t.get('fields', [])):
                        fid = field.get('id')
                        if not fid or fid in seen_ids:
                            field['id'] = generate_field_id()
                        seen_ids.add(field['id'])
                    tpls.append(t)
            except Exception:
                pass
        return tpls

    # 向下相容備用搬遷：若是舊專案只有 db/form_templates.json
    if os.path.exists(JSON_TPL_PATH):
        try:
            with open(JSON_TPL_PATH, "r", encoding="utf-8") as f:
                old_tpls = json.load(f)
                for t in old_tpls:
                    save_single_template(t)
                return old_tpls
        except Exception:
            pass

    return []

def save_single_template(template: dict):
    """將單一模板寫入獨立的 db/templates/tpl_xx.json 檔案 (INV-UI-FORM-13 按需獨立寫入)"""
    os.makedirs(TEMPLATES_DIR, exist_ok=True)
    tpl_id = template.get('id') or f"tpl_{int(time.time())}"
    template['id'] = tpl_id
    fpath = os.path.join(TEMPLATES_DIR, f"{tpl_id}.json")
    with open(fpath, "w", encoding="utf-8") as f:
        json.dump(template, f, ensure_ascii=False, indent=2)

def delete_single_template(tpl_id: str):
    """單獨刪除指定的 db/templates/tpl_xx.json 檔案 (INV-UI-FORM-13 獨立檔案刪除)"""
    fpath = os.path.join(TEMPLATES_DIR, f"{tpl_id}.json")
    if os.path.exists(fpath):
        try:
            os.remove(fpath)
        except Exception:
            pass

def save_json_templates(templates):
    """批次寫入所有獨立模板檔案 (向下相容)"""
    for t in templates:
        save_single_template(t)

def safe_int(val) -> int:
    if val is None:
        return 0
    try:
        f = float(val)
        if math.isnan(f) or math.isinf(f):
            return 0
        return int(round(f))
    except:
        return 0

def format_db_value(db_k: str, val_raw) -> str:
    """精確格式化 DB 36 欄位與 15 大照護細節數值 (INV-UI-FORM-12 精確單詞防護 Guardrail)"""
    if val_raw is None or str(val_raw).strip() == "":
        return "—"
    
    val_str_clean = str(val_raw).strip()
    db_k_lower = str(db_k).lower()
    
    # 1. 優先判斷日期欄位 (防止 date 被誤判為金額)
    if "date" in db_k_lower or "due" in db_k_lower:
        if isinstance(val_raw, (date, datetime)):
            return val_raw.strftime("%Y-%m-%d")
        return val_str_clean.split(" ")[0].strip()
    
    # 2. 金額欄位單詞精確比對 (Exact Word Matching)，防止 breastfeeding 等包含 fee 的單詞被誤判！
    money_words = {'amount', 'fee', 'payable', 'rate', 'salary', 'deposit'}
    key_tokens = set(db_k_lower.split('_'))
    
    is_money_field = bool(key_tokens.intersection(money_words)) and not ('days' in db_k_lower or 'hours' in db_k_lower)
    
    if is_money_field:
        # 雙重驗證：只有當 val_raw 確實是有效數字時才加上「元」
        try:
            f = float(val_raw)
            if not (math.isnan(f) or math.isinf(f)):
                return f"{safe_int(f):,} 元"
        except (ValueError, TypeError):
            pass
        
    return val_str_clean


# 按 SQL 資料庫資料表來源原生歸屬分層組織全系統欄位 (INV-UI-FORM-14/15/25 全量資料庫欄位 100% 完整開載公理)
DB_TABLE_FIELDS = {
    "orders (訂單主表 - 36 大業務與金額 calculations)": {
        "case_no": "案件編號 (case_no)",
        "start_date": "預期服務開始日 (start_date)",
        "actual_start_date": "服務開始 (actual_start_date)",
        "actual_end_date": "服務結束 (actual_end_date)",
        "service_days": "希望服務天數 (service_days)",
        "service_hours_per_day": "每日服務時數 (service_hours_per_day)",
        "service_time": "服務時段 (service_time)",
        "service_mode": "服務方式 (service_mode)",
        "subsidy_eligibility": "身分補助資格 (subsidy_eligibility)",
        "total_hours": "服務總時數 (total_hours)",
        "subsidy_hours": "補助時數 (subsidy_hours)",
        "self_pay_hours": "自費時數 (self_pay_hours)",
        "claim_total_days": "請款總日數 (claim_total_days)",
        "floor_fee": "樓層費用 (floor_fee)",
        "employer_hourly_rate": "雇主單價 (employer_hourly_rate)",
        "deposit_days": "訂金天數 (deposit_days)",
        "deposit_amount": "訂金金額 (deposit_amount)",
        "deposit_date": "訂金入帳日 (deposit_date)",
        "first_payment_days": "第一期款天數 (first_payment_days)",
        "first_payment_amount": "第一期金額 (first_payment_amount)",
        "first_payment_date": "第一期款入帳日 (first_payment_date)",
        "second_payment_days": "第二期款天數 (second_payment_days)",
        "second_payment_amount": "第二期金額 (second_payment_amount)",
        "second_payment_date": "第二期款入帳日 (second_payment_date)",
        "final_payment_date": "尾款日期 (final_payment_date)",
        "total_employer_self_pay_payable": "雇主自費合計金額 (total_employer_self_pay_payable)",
        "order_status": "訂單成立狀態 (order_status)",
        "staff_name": "服務人員 (staff_name)",
        "caregiver_rate": "服務單價 (caregiver_rate)",
        "special_holidays": "特殊休假 (special_holidays)",
        "service_salary": "服務薪資 (service_salary)",
        "salary_payment_date_1": "預計發薪日 (salary_payment_date_1)",
        "subsidy_salary": "補助薪資金額 (subsidy_salary)",
        "total_caregiver_salary": "總薪資 (total_caregiver_salary)",
        "govt_claim_date": "市府請款日 (govt_claim_date)"
    },
    "clients (客戶主表 - 個人基本資料與市府申請表)": {
        "client_name": "客戶名稱 (client_name)",
        "phone": "聯絡電話 (phone)",
        "address": "通訊與服務地址 (address)",
        "due_date": "預產期 (due_date)",
        "email": "Email 電子郵件 (email)",
        "bank_code": "雇主退款銀行代號/分行 (bank_code)",
        "bank_account": "雇主退款銀行帳號 (bank_account)",
        "service_mode": "服務方式 (service_mode)",
        "baby_info": "寶寶資訊 (baby_info)",
        "delivery_mode": "生產方式 (delivery_mode)",
        "residence_type": "居住型態 (residence_type)",
        "notes": "其他備註事項 (notes)",
        "id_card": "身分證字號 (id_card)",
        "line_id": "LINE ID (line_id)"
    },
    "staff (服務人員 - 月嫂主表與撥款帳號)": {
        "staff_name": "月嫂姓名 (staff_name)",
        "staff_phone": "月嫂電話 (staff_phone)",
        "staff_bank_code": "月嫂銀行代號/分行 (staff_bank_code)",
        "staff_bank_account": "月嫂銀行帳號 (staff_bank_account)",
        "id_card": "月嫂身分證號 (id_card)",
        "transportation": "服務交通工具 (transportation)",
        "service_region": "可承接案件區域 (service_region)"
    },
    "beclass_records (BeClass 報名表 - 15 大產婦照顧與飲食環境細節)": {
        "dietary_habits": "月子餐點調理喜好/飲食習慣",
        "vegetarian_preference": "無葷食時是否可接受蛋奶素餐食",
        "alcohol_ratio": "餐飲含酒比例 (全酒/半酒/無酒)",
        "cooking_oil_type": "料理用油 (苦茶油/麻油/橄欖油)",
        "maternal_allergy": "媽咪有無過敏體質",
        "special_care_notes": "特殊照護時應注意事項",
        "meal_preferences": "餐點喜忌備註",
        "cooking_tools": "現有烹煮工具 (炒菜鍋/電鍋/烤箱)",
        "bath_water_prep": "洗澡水準備方式 (熱水/薑水/中藥包)",
        "breastfeeding_method": "哺乳方式 (母乳/配方奶)",
        "holiday_pricing_terms": "特殊計費: 國定節日加班費",
        "multi_birth_count": "特殊計費: 胎數",
        "stair_floor_fee_mode": "透天服務樓層方式 (樓層加收費)",
        "parking_space_provided": "提供服務人員轎車停車位",
        "other_babies_present": "服務時間內是否有其他寶寶",
        "bank_code": "補助款退款:銀行代號+分行代號",
        "bank_account": "補助款退款:銀行帳號"
    },
    "global_stats (全域與多案件營運統計視圖)": {
        "global_active_orders_count": "📊 本週服務中案件總數",
        "global_active_staff_count": "📊 本週出勤月嫂總人數",
        "global_subsidy_orders_count": "📊 補助身份案件統計總數",
        "global_total_receivable_sum": "📊 系統應收自費總金額彙總",
        "global_govt_claim_count": "📊 本月市府請款案件總數"
    }
}

# 打平索引（向下相容與繪製視圖用）
ALL_DB_FIELDS = {}
for _tbl, _fmap in DB_TABLE_FIELDS.items():
    ALL_DB_FIELDS.update(_fmap)

def get_table_for_key(db_key: str) -> str:
    """自動反查 db_key 隸屬哪一個 SQL 資料表」"""
    for tbl_name, fmap in DB_TABLE_FIELDS.items():
        if db_key in fmap:
            return tbl_name
    return list(DB_TABLE_FIELDS.keys())[0]


def render_html_document(tpl_data: dict, target_order: dict, global_stats: dict) -> str:
    """渲染雙欄 CSS Grid 高質感 HTML 實體單據"""
    html_items = ""
    for f in tpl_data.get('fields', []):
        lbl = f.get('label', '未命名欄位')
        f_type = f.get('type', 'text')
        f_width = f.get('width', 'half')
        val_str = "___________"
        
        if f_type == "db_link":
            db_k = f.get('db_key', 'client_name')
            if db_k in global_stats:
                val_raw = global_stats[db_k]
                val_str = format_db_value(db_k, val_raw)
            elif target_order:
                val_raw = target_order.get(db_k, '')
                val_str = format_db_value(db_k, val_raw)
        
        is_full = (f_width == "full" or f_type == "textarea")
        grid_span = "grid-column: span 2;" if is_full else ""
        
        html_items += f"""
        <div style="background:#FFFFFF; border:1px solid #E0E0E0; border-radius:6px; padding:10px 14px; {grid_span}">
            <div style="font-size:12px; color:#616161; font-weight:600; margin-bottom:4px;">• {lbl}</div>
            <div style="font-size:15px; color:#1565C0; font-weight:bold; word-break:break-word;">{val_str}</div>
        </div>
        """

    print_doc_html = f"""
    <div id="printable-area" style="font-family:'Segoe UI', Microsoft JhengHei, sans-serif; padding:25px; border:2px solid #1565C0; background:#FAFAFA; border-radius:8px; max-width:850px; margin:auto;">
        <div style="text-align:center; border-bottom:2px double #1565C0; padding-bottom:12px; margin-bottom:20px;">
            <h2 style="margin:0; color:#1565C0; font-size:22px;">新竹市月子照顧服務人員職業工會</h2>
            <h3 style="margin:6px 0 0 0; color:#333; font-size:17px;">【 {tpl_data.get('name', '表單單據')} 】</h3>
            <p style="margin:4px 0 0 0; color:#757575; font-size:11px;">單據編號：TPL-{tpl_data.get('id', '00')} | 列印日期：{date.today().strftime('%Y-%m-%d')}</p>
        </div>
        
        <div style="display:grid; grid-template-columns: 1fr 1fr; gap:10px 14px;">
            {html_items}
        </div>
    </div>

    <div style="text-align:center; margin-top:15px;">
        <button onclick="window.print()" style="background-color:#2E7D32; color:white; border:none; padding:10px 24px; font-size:15px; font-weight:bold; border-radius:6px; cursor:pointer; box-shadow:0 2px 4px rgba(0,0,0,0.2);">
            🖨️ 點擊另存為 PDF / 輸出圖片 (Print / Save as PDF)
        </button>
    </div>
    """
    return print_doc_html


def show():
    """FormManagementUI 進入點 (5:5 雙視窗 Side-by-Side + 拖拉排序 + 二次確認刪除)"""
    st.title("📋 表單與履歷問卷管理專區")

    try:
        orders_data = db_service.get_order_details()
    except Exception as e:
        st.error(f"讀取資料庫失敗: {e}")
        orders_data = []

    global_stats = {
        "global_active_orders_count": len([o for o in orders_data if o.get('order_status') in ['服務中', '訂單成立']]),
        "global_active_staff_count": len(set([o['staff_name'] for o in orders_data if o.get('staff_name')])),
        "global_subsidy_orders_count": len([o for o in orders_data if o.get('subsidy_eligibility') != '一般身分']),
        "global_total_receivable_sum": sum([safe_int(o.get('total_employer_self_pay_payable')) for o in orders_data]),
        "global_govt_claim_count": len([o for o in orders_data if o.get('govt_claim_date')])
    }

    # 頂部：選擇連動模式與作用域
    col_scope, col_order = st.columns([1.5, 2.5])
    with col_scope:
        scope_mode = st.radio("⚙️ 選擇表單連動作用域", ["🎯 特定單筆案件 (契約/個人單據)", "📊 全域/多案件統計模式 (週報/統計表)"], horizontal=True, key="sbs_scope_mode")
    
    target_order = None
    with col_order:
        if "特定單筆案件" in scope_mode and orders_data:
            order_opts = {
                f"案件 #{o['case_no']} - 客戶: {o['client_name']} [{o['order_status']}] (月嫂: {o.get('staff_name') or '尚未指派'})": o['case_no']
                for o in orders_data
            }
            sel_label = st.selectbox("🎯 選擇連動測試的訂單案件", list(order_opts.keys()), key="sbs_order_picker")
            target_case_no = order_opts[sel_label]
            target_order = next((o for o in orders_data if o['case_no'] == target_case_no), None)
            if target_order:
                try:
                    client_rows = db_service.get_table_data('clients')
                    client_row = next((row for row in client_rows if row.get('case_no') == target_case_no), None)
                    if client_row:
                        target_order = {
                            **target_order,
                            **{
                                key: client_row.get(key, '')
                                for key in ('service_time', 'service_type', 'delivery_type', 'residence_type', 'city')
                            },
                        }
                except Exception:
                    pass
        else:
                st.info("💡 目前切換為「全域/多案件統計模式」，無須鎖定單一訂單。")

    # 讀取 JSON 持久化表單模板 (INV-UI-FORM-02)
    st.session_state['custom_form_templates'] = load_json_templates()

    st.markdown("---")

    tab1, tab2, tab3 = st.tabs([
            "➕ 1. 手動創建與設計新表單 (UX 實驗室)", 
            "🗄️ 2. 自訂表單模板庫與 5:5 雙視窗線上編輯預覽",
            "📜 3. 制式定型化契約管理 (EPPP 變數代理引擎)"
    ])

    field_widths = {
            "half": "半寬 (50% 雙欄並排)",
            "full": "全寬 (100% 單獨一列)"
    }
    field_types = {
            "text": "單行文字輸入",
            "textarea": "多行備註區域",
            "number": "數字/金額數值",
            "date": "日期選擇器",
            "db_link": "⚡ 連動 DB 欄位 (支援單筆與全域統計)"
    }

    # =========================================================================
    # TAB 1: 手動創建與設計新表單 (UX 實驗室)
    # =========================================================================
    with tab1:
            st.markdown("### 🛠️ 步驟一：表單基本資訊與用途設定")
            c_title, c_desc = st.columns([1.5, 2.5])
            with c_title:
                builder_title = st.text_input("表單名稱", value="自訂母嬰照顧合約證明", key="sbs_title_input")
            with c_desc:
                builder_desc = st.text_input("表單用途說明", value="供客戶申報補助與工會備查之標準單據", key="sbs_desc_input")

            st.markdown("---")
            st.markdown("### ⚙️ 步驟二：動態新增與設計表單欄位 (支援 [⬆️上移] [⬇️下移] 順序平移)")

            if 'builder_fields' not in st.session_state:
                st.session_state['builder_fields'] = [
                    {"id": generate_field_id(), "label": "客戶姓名", "type": "db_link", "db_key": "client_name", "width": "half"},
                    {"id": generate_field_id(), "label": "案件編號", "type": "db_link", "db_key": "case_no", "width": "half"},
                    {"id": generate_field_id(), "label": "服務薪資", "type": "db_link", "db_key": "service_salary", "width": "half"},
                    {"id": generate_field_id(), "label": "樓層費", "type": "db_link", "db_key": "floor_fee", "width": "half"},
                    {"id": generate_field_id(), "label": "預計發薪日", "type": "db_link", "db_key": "salary_payment_date_1", "width": "half"},
                    {"id": generate_field_id(), "label": "服務地址", "type": "db_link", "db_key": "address", "width": "full"},
                    {"id": generate_field_id(), "label": "注意事項與簽名聲明", "type": "textarea", "db_key": "", "width": "full"}
                ]

            col_add, _ = st.columns([1, 4])
            with col_add:
                if st.button("➕ 新增一個新欄位", key="btn_add_sbs_field"):
                    st.session_state['builder_fields'].append({
                        "id": generate_field_id(),
                        "label": f"新自訂欄位 {len(st.session_state['builder_fields']) + 1}",
                        "type": "text",
                        "db_key": "client_name",
                        "width": "half"
                    })
                    st.rerun()

            # 確保 Tab 1 builder_fields 中的 id 全域唯一 (INV-UI-FORM-07)
            seen_b_ids = set()
            for idx, f in enumerate(st.session_state['builder_fields']):
                fid = f.get('id')
                if not fid or fid in seen_b_ids:
                    f['id'] = generate_field_id()
                seen_b_ids.add(f['id'])

            # 渲染 Tab 1 欄位建造器（含順序上下平移）
            for idx, f in enumerate(st.session_state['builder_fields']):
                fid = f['id']
                with st.container(border=True):
                    fc1, fc2, fc3, fc4, fc_up, fc_dn, fc_del = st.columns([2, 1.8, 2.2, 1.6, 0.5, 0.5, 0.6])
                    with fc1:
                        f['label'] = st.text_input(f"欄位 #{idx+1} 名稱", value=f['label'], key=f"sbs_fl_lbl_{fid}")
                    with fc2:
                        type_keys = list(field_types.keys())
                        curr_t_idx = type_keys.index(f['type']) if f['type'] in type_keys else 0
                        f['type'] = st.selectbox(f"資料型態", type_keys, index=curr_t_idx, format_func=lambda x: field_types[x], key=f"sbs_fl_type_{fid}")
                    with fc3:
                        if f['type'] == "db_link":
                            curr_db_k = f.get('db_key', 'client_name')
                            curr_tbl = get_table_for_key(curr_db_k)
                            tbl_list = list(DB_TABLE_FIELDS.keys())
                            c_t_idx = tbl_list.index(curr_tbl) if curr_tbl in tbl_list else 0
                            
                            sel_tbl = st.selectbox("1️⃣ 資料表來源", tbl_list, index=c_t_idx, key=f"sbs_tbl_t1_{fid}")
                            
                            tbl_fmap = DB_TABLE_FIELDS[sel_tbl]
                            f_keys = list(tbl_fmap.keys())
                            c_k_idx = f_keys.index(curr_db_k) if curr_db_k in f_keys else 0
                            f['db_key'] = st.selectbox("2️⃣ 綁定目標欄位", f_keys, index=c_k_idx, format_func=lambda x: tbl_fmap[x], key=f"sbs_fl_db_t1_{fid}")
                        else:
                            st.caption("（手動填寫欄位）")
                    with fc4:
                        w_keys = list(field_widths.keys())
                        f_w = f.get('width', 'half')
                        curr_w_idx = w_keys.index(f_w) if f_w in w_keys else 0
                        f['width'] = st.selectbox("排版寬度", w_keys, index=curr_w_idx, format_func=lambda x: field_widths[x], key=f"sbs_fl_w_{fid}")
                    
                    # 順序平移按鈕 (INV-UI-FORM-09)
                    with fc_up:
                        st.write("")
                        if st.button("⬆️", key=f"btn_up_t1_{fid}", disabled=(idx == 0)):
                            st.session_state['builder_fields'][idx], st.session_state['builder_fields'][idx-1] = st.session_state['builder_fields'][idx-1], st.session_state['builder_fields'][idx]
                            st.rerun()
                    with fc_dn:
                        st.write("")
                        if st.button("⬇️", key=f"btn_dn_t1_{fid}", disabled=(idx == len(st.session_state['builder_fields']) - 1)):
                            st.session_state['builder_fields'][idx], st.session_state['builder_fields'][idx+1] = st.session_state['builder_fields'][idx+1], st.session_state['builder_fields'][idx]
                            st.rerun()
                    with fc_del:
                        st.write("")
                        if st.button("🗑️", key=f"btn_del_t1_{fid}"):
                            st.session_state['builder_fields'] = [x for x in st.session_state['builder_fields'] if x.get('id') != fid]
                            st.rerun()

            st.markdown("---")
            st.markdown("### 👁️ 步驟三：實時 UI 渲染預覽與 UX 測試區")

            with st.container(border=True):
                st.markdown(f"## 📋 【預覽】{builder_title}")
                st.caption(f"📝 說明：{builder_desc}")
                st.markdown("---")

                prev_cols = st.columns(2)
                for i, f in enumerate(st.session_state['builder_fields']):
                    with prev_cols[i % 2]:
                        lbl = f['label']
                        f_type = f['type']
                        
                        if f_type == "db_link":
                            db_k = f.get('db_key', 'client_name')
                            if db_k in global_stats:
                                val_raw = global_stats[db_k]
                                val_disp = format_db_value(db_k, val_raw)
                                st.text_input(f"⚡ {lbl} (全域連動)", value=val_disp, disabled=True, key=f"pv_sbs_gdb_{i}")
                            else:
                                val_raw = target_order.get(db_k, '—') if target_order else '— (需選取單筆案件)'
                                val_disp = format_db_value(db_k, val_raw) if target_order else val_raw
                                st.text_input(f"⚡ {lbl} (單筆 DB 連動)", value=val_disp, disabled=True, key=f"pv_sbs_sdb_{i}")
                        elif f_type == "text":
                            st.text_input(lbl, value="", key=f"pv_sbs_txt_{i}")
                        elif f_type == "textarea":
                            st.text_area(lbl, value="", key=f"pv_sbs_area_{i}")
                        elif f_type == "number":
                            st.number_input(lbl, value=0, step=1, key=f"pv_sbs_num_{i}")
                        elif f_type == "date":
                            st.date_input(lbl, value=date.today(), key=f"pv_sbs_date_{i}")

                st.markdown("---")
                if st.button("💾 確定儲存為新表單模板", key="btn_save_sbs_tpl", type="primary"):
                    if not builder_title.strip():
                        st.error("請輸入表單名稱！")
                    else:
                        new_tpl = {
                            "id": f"tpl_{len(st.session_state['custom_form_templates'])+1:02d}_{int(time.time())}",
                            "name": builder_title.strip(),
                            "desc": builder_desc.strip(),
                            "fields": json.loads(json.dumps(st.session_state['builder_fields']))
                        }
                        st.session_state['custom_form_templates'].append(new_tpl)
                        save_single_template(new_tpl)
                        st.success(f"🎉 新表單模板【{builder_title}】已成功寫入 `db/templates/{new_tpl['id']}.json` 保存！")

    # =========================================================================
    # TAB 2: 自訂表單模板庫 (5:5 雙視窗 Side-by-Side + 拖拉排序 + 二次確認刪除)
    # =========================================================================
    with tab2:
            st.markdown("### 🗄️ 所有已建立之表單模板庫 (支援 5:5 雙視窗實時預覽、順序平移與二次刪除確認)")
            
            if not st.session_state['custom_form_templates']:
                st.info("目前尚無任何自訂表單模板。請至 Tab 1 手動創建新表單。")
            else:
                tpl_names = {t['name']: t['id'] for t in st.session_state['custom_form_templates']}
                sel_tpl_name = st.selectbox("選取要檢視或修改的表單模板", list(tpl_names.keys()), key="sbs_tpl_picker")
                curr_tpl_idx = next((i for i, t in enumerate(st.session_state['custom_form_templates']) if t['name'] == sel_tpl_name), 0)
                curr_tpl = st.session_state['custom_form_templates'][curr_tpl_idx]

                btn_col1, btn_col2, _ = st.columns([1.5, 1.5, 3])
                edit_key = f"editing_mode_{curr_tpl['id']}"
                if edit_key not in st.session_state:
                    st.session_state[edit_key] = False

                draft_key = f"edit_draft_tpl_{curr_tpl['id']}"
                confirm_del_key = f"confirm_del_mode_{curr_tpl['id']}"
                if confirm_del_key not in st.session_state:
                    st.session_state[confirm_del_key] = False
                    
                with btn_col1:
                    if st.button("✏️ 編輯修改此模板與欄位順序", key=f"btn_toggle_sbs_{curr_tpl['id']}", type="primary"):
                        st.session_state[edit_key] = not st.session_state[edit_key]
                        if st.session_state[edit_key]:
                            d_copy = json.loads(json.dumps(curr_tpl))
                            for idx, f in enumerate(d_copy.get('fields', [])):
                                if not f.get('id'):
                                    f['id'] = generate_field_id()
                            st.session_state[draft_key] = d_copy
                        else:
                            st.session_state.pop(draft_key, None)
                        st.rerun()

                with btn_col2:
                    # 刪除按鈕 (觸發二次確認彈窗，INV-UI-FORM-08)
                    if st.button("🗑️ 刪除此表單模板", key=f"btn_trigger_del_{curr_tpl['id']}"):
                        st.session_state[confirm_del_key] = True
                        st.rerun()

                # -----------------------------------------------------------------
                # 🚨 刪除二次確認對話框 (Delete Confirmation Modal Guardrail)
                # -----------------------------------------------------------------
                if st.session_state[confirm_del_key]:
                    with st.container(border=True):
                        st.warning(f"⚠️ **確定要永久刪除【{curr_tpl['name']}】表單模板嗎？** 此動作無法復原！")
                        c_del1, c_del2, _ = st.columns([1.5, 1.5, 4])
                        with c_del1:
                            if st.button("💥 確定永久刪除", key=f"btn_do_del_{curr_tpl['id']}", type="primary"):
                                del_target_id = curr_tpl['id']
                                st.session_state['custom_form_templates'].pop(curr_tpl_idx)
                                delete_single_template(del_target_id)
                                st.session_state[confirm_del_key] = False
                                st.success(f"已成功永久刪除模板實體檔案：{curr_tpl['name']}")
                                st.rerun()
                        with c_del2:
                            if st.button("✖️ 取消刪除", key=f"btn_cancel_del_{curr_tpl['id']}"):
                                st.session_state[confirm_del_key] = False
                                st.rerun()

                # =================================================================
                # 5:5 左右雙視窗 Side-by-Side 線上編輯與實時同步預覽器 (INV-UI-FORM-10)
                # =================================================================
                if st.session_state[edit_key] and draft_key in st.session_state:
                    st.markdown("---")
                    draft_tpl = st.session_state[draft_key]
                    st.markdown(f"### ✏️ 5:5 雙視窗線上編輯【{draft_tpl['name']}】草稿 (左側編輯，右側 0.1秒實時同步預覽)")
                    
                    col_left, col_right = st.columns([1, 1])

                    # -------------------------------------------------------------
                    # 左側 50%: 實時編輯與拖拉排序面板
                    # -------------------------------------------------------------
                    with col_left:
                        st.markdown("#### ⚙️ 左側：編輯名稱與欄位順序平移")
                        ed_name = st.text_input("表單名稱", value=draft_tpl['name'], key=f"ed_sbs_name_{draft_tpl['id']}")
                        ed_desc = st.text_input("用途說明", value=draft_tpl['desc'], key=f"ed_sbs_desc_{draft_tpl['id']}")
                        draft_tpl['name'] = ed_name
                        draft_tpl['desc'] = ed_desc
                        
                        col_add_e, _ = st.columns([1, 2])
                        with col_add_e:
                            if st.button("➕ 追加一個新欄位", key=f"btn_add_sbs_e_{draft_tpl['id']}"):
                                draft_tpl['fields'].append({
                                    "id": generate_field_id(),
                                    "label": f"新自訂欄位 {len(draft_tpl['fields']) + 1}",
                                    "type": "text",
                                    "db_key": "",
                                    "width": "half"
                                })
                                st.rerun()

                        # 確保 Tab 2 草稿中的 id 全域唯一 (INV-UI-FORM-07)
                        seen_d_ids = set()
                        for idx, f in enumerate(draft_tpl['fields']):
                            fid = f.get('id')
                            if not fid or fid in seen_d_ids:
                                f['id'] = generate_field_id()
                            seen_d_ids.add(f['id'])

                        # 渲染編輯卡片與順序調控按鈕 (INV-UI-FORM-09/10)
                        for idx, f in enumerate(draft_tpl['fields']):
                            fid = f['id']
                            with st.container(border=True):
                                st.caption(f"📌 欄位 #{idx+1}")
                                e_c1, e_c2 = st.columns([1.5, 1.5])
                                with e_c1:
                                    f['label'] = st.text_input("欄位名稱", value=f['label'], key=f"se_lbl_{draft_tpl['id']}_{fid}")
                                with e_c2:
                                    type_keys = list(field_types.keys())
                                    c_t_idx = type_keys.index(f['type']) if f['type'] in type_keys else 0
                                    f['type'] = st.selectbox("資料型態", type_keys, index=c_t_idx, format_func=lambda x: field_types[x], key=f"se_type_{draft_tpl['id']}_{fid}")
                                
                                e_c3, e_c4 = st.columns([1.5, 1.5])
                                with e_c3:
                                    if f['type'] == "db_link":
                                        curr_db_k = f.get('db_key', 'client_name')
                                        curr_tbl = get_table_for_key(curr_db_k)
                                        tbl_list = list(DB_TABLE_FIELDS.keys())
                                        c_t_idx = tbl_list.index(curr_tbl) if curr_tbl in tbl_list else 0
                                        
                                        sel_tbl = st.selectbox("1️⃣ 資料表來源", tbl_list, index=c_t_idx, key=f"se_tbl_{draft_tpl['id']}_{fid}")
                                        
                                        tbl_fmap = DB_TABLE_FIELDS[sel_tbl]
                                        f_keys = list(tbl_fmap.keys())
                                        c_k_idx = f_keys.index(curr_db_k) if curr_db_k in f_keys else 0
                                        f['db_key'] = st.selectbox("2️⃣ 綁定目標欄位", f_keys, index=c_k_idx, format_func=lambda x: tbl_fmap[x], key=f"se_db_{draft_tpl['id']}_{fid}")
                                    else:
                                        st.caption("（手動填寫）")
                                with e_c4:
                                    w_keys = list(field_widths.keys())
                                    f_w = f.get('width', 'half')
                                    c_w_idx = w_keys.index(f_w) if f_w in w_keys else 0
                                    f['width'] = st.selectbox("排版寬度", w_keys, index=c_w_idx, format_func=lambda x: field_widths[x], key=f"se_w_{draft_tpl['id']}_{fid}")

                                # 控制按鈕列: 上移 / 下移 / 刪除
                                b_u, b_d, b_x, _ = st.columns([1, 1, 1, 3])
                                with b_u:
                                    if st.button("⬆️ 上移", key=f"b_up_{draft_tpl['id']}_{fid}", disabled=(idx == 0)):
                                        draft_tpl['fields'][idx], draft_tpl['fields'][idx-1] = draft_tpl['fields'][idx-1], draft_tpl['fields'][idx]
                                        st.rerun()
                                with b_d:
                                    if st.button("⬇️ 下移", key=f"b_dn_{draft_tpl['id']}_{fid}", disabled=(idx == len(draft_tpl['fields']) - 1)):
                                        draft_tpl['fields'][idx], draft_tpl['fields'][idx+1] = draft_tpl['fields'][idx+1], draft_tpl['fields'][idx]
                                        st.rerun()
                                with b_x:
                                    if st.button("🗑️ 刪除", key=f"b_del_{draft_tpl['id']}_{fid}"):
                                        draft_tpl['fields'] = [x for x in draft_tpl['fields'] if x.get('id') != fid]
                                        st.rerun()

                        st.markdown("---")
                        es1, es2 = st.columns([1.5, 2])
                        with es1:
                            if st.button("💾 確定更新此模板 (寫入獨立 JSON 檔)", key=f"btn_save_sbs_tpl_chg_{draft_tpl['id']}", type="primary"):
                                if not ed_name.strip():
                                    st.error("表單名稱不能為空！")
                                else:
                                    updated_tpl = json.loads(json.dumps(draft_tpl))
                                    st.session_state['custom_form_templates'][curr_tpl_idx] = updated_tpl
                                    save_single_template(updated_tpl)
                                    st.session_state[edit_key] = False
                                    st.session_state.pop(draft_key, None)
                                    st.success(f"🎉 模板【{ed_name}】已成功更新寫入 `db/templates/{updated_tpl['id']}.json`！")
                                    st.rerun()
                        with es2:
                            if st.button("✖️ 取消編輯 (丟棄草稿)", key=f"btn_cancel_sbs_edit_{draft_tpl['id']}"):
                                st.session_state[edit_key] = False
                                st.session_state.pop(draft_key, None)
                                st.info("已取消編輯，草稿已丟棄。")
                                st.rerun()

                    # -------------------------------------------------------------
                    # 右側 50%: 實時同步單據與 Print-to-PDF 視窗
                    # -------------------------------------------------------------
                    with col_right:
                        st.markdown("#### 👁️ 右側：實時單據與 PDF 即時預覽")
                        live_html = render_html_document(draft_tpl, target_order, global_stats)
                        st.iframe(live_html, height=620)

                # =================================================================
                # 實體單據預覽與 PDF / 圖片導出 (標準檢視視窗)
                # =================================================================
                else:
                    st.markdown("---")
                    doc_html = render_html_document(curr_tpl, target_order, global_stats)
                    st.iframe(doc_html, height=580)

    # =========================================================================
    # TAB 3: 制式定型化契約管理 (EPPP 變數代理引擎 - INV-UI-FORM-16/17/18)
    # =========================================================================
    with tab3:
        st.markdown("### 📜 定型化契約變數代理管理引擎 (EPPP Engine)")
        st.caption("輕鬆將 Excel 範本與系統 SQL 資料庫欄位動態綁定，支援 1:1 CSS A4 紙本鏡像視窗、全滿版預覽與全套範本生命週期管理。")
        st.markdown("---")

        contracts = load_contract_templates()
        if not contracts:
            st.warning("目前尚無任何定型化契約範本。已自動為您建立預設標準契約！")
            st.rerun()

        c_names = {c['name']: c['id'] for c in contracts}
        
        # 契約頂部導覽列、視角切換器與操作按鈕 (INV-UI-FORM-27/28)
        c_pick_col, c_view_mode_col, c_btn_edit, c_btn_del = st.columns([2.2, 1.8, 1, 1])
        with c_pick_col:
            sel_c_name = st.selectbox("選取要檢視與設定的定型化契約範本", list(c_names.keys()), key="eppp_contract_picker")
            curr_contract = next((c for c in contracts if c['name'] == sel_c_name), contracts[0])
            curr_cid = curr_contract['id']

        contract_target_order = target_order
        if curr_cid == "contract_staff_service":
            if not target_order:
                contract_target_order = None
                st.info("請先選擇案件，再載入服務人員契約資料。")
            else:
                assignment_text = st.text_input(
                    "服務人員指派 ID（同案有多位月嫂時必填）",
                    key=f"staff_contract_assignment_{target_order['case_no']}",
                ).strip()
                try:
                    assignment_id = int(assignment_text) if assignment_text else None
                    if assignment_id is not None and assignment_id < 1:
                        raise ValueError("指派 ID 必須為正整數")
                    contract_target_order = flatten_staff_contract_context(
                        fetch_staff_contract_context(target_order["case_no"], assignment_id)
                    )
                except ValueError as error:
                    contract_target_order = None
                    st.warning(f"服務人員契約資料未載入：{error}")

        with c_view_mode_col:
            view_mode = st.radio("畫面排版視角 (INV-UI-FORM-28)", ["🌓 5:5 左右對照維護", "🔍 100% 全寬滿版預覽"], horizontal=True, key=f"v_mode_{curr_cid}")

        contract_edit_key = f"c_editing_{curr_cid}"
        contract_draft_key = f"c_draft_{curr_cid}"
        contract_del_modal_key = f"c_del_modal_{curr_cid}"
        
        is_contract_editing = st.session_state.get(contract_edit_key, False)

        with c_btn_edit:
            st.write("")
            if not is_contract_editing:
                if st.button("✏️ 編輯對照", key=f"btn_c_edit_{curr_cid}", use_container_width=True):
                    st.session_state[contract_edit_key] = True
                    st.session_state[contract_draft_key] = json.loads(json.dumps(curr_contract))
                    st.rerun()
            else:
                if st.button("✖️ 取消編輯", key=f"btn_c_cancel_{curr_cid}", use_container_width=True):
                    st.session_state[contract_edit_key] = False
                    st.session_state.pop(contract_draft_key, None)
                    st.rerun()

        with c_btn_del:
            st.write("")
            if st.button("🗑️ 刪除範本", key=f"btn_c_del_{curr_cid}", type="secondary", use_container_width=True):
                st.session_state[contract_del_modal_key] = True

        # 二次確認刪除 Modal
        if st.session_state.get(contract_del_modal_key, False):
            with st.container(border=True):
                st.error(f"⚠️ **確定要永久刪除契約範本【{curr_contract['name']}】嗎？**")
                st.caption(f"此操作將從 `db/templates/contracts/{curr_cid}.json` 硬碟檔案中徹底移除！")
                
                dm_col1, dm_col2 = st.columns([1, 1])
                with dm_col1:
                    if st.button("💥 確定永久刪除", key=f"btn_c_confirm_del_{curr_cid}", type="primary"):
                        fpath = os.path.join(CONTRACTS_DIR, f"{curr_cid}.json")
                        try:
                            if os.path.exists(fpath):
                                os.remove(fpath)
                        except Exception:
                            pass
                        st.session_state[contract_del_modal_key] = False
                        st.success(f"🗑️ 已成功刪除契約範本【{curr_contract['name']}】！")
                        st.rerun()
                with dm_col2:
                    if st.button("取消", key=f"btn_c_cancel_del_modal_{curr_cid}"):
                        st.session_state[contract_del_modal_key] = False
                        st.rerun()

        st.markdown("---")

        if view_mode == "🔍 100% 全寬滿版預覽":
            st.markdown("#### 👁️ 100% 全螢幕/全寬滿版 A4 沉浸契約預覽視窗")
            st.info("💡 目前已進入全寬滿版預覽視角，文字更清晰大方！切換回【🌓 5:5 左右對照維護】即可邊改邊看！")
            
            param_values = {}
            for p_tag, p_info in curr_contract.get('param_mappings', {}).items():
                db_k = p_info.get('db_key', '')
                if db_k in global_stats:
                    val_raw = global_stats[db_k]
                    param_values[p_tag] = format_db_value(db_k, val_raw)
                elif target_order and db_k in target_order:
                    val_raw = target_order.get(db_k, '')
                    param_values[p_tag] = format_db_value(db_k, val_raw)
                else:
                    param_values[p_tag] = f"<span style='color:#D32F2F; text-decoration:underline;'>___{p_info.get('label')}___</span>"

            if curr_contract.get('template_filename', '').lower().endswith('.xlsx'):
                contract_html = render_excel_contract_mirror(curr_contract, contract_target_order, global_stats)
            else:
                contract_html = f"<div>預設範本</div>"
            
            st.iframe(contract_html, height=1100)
        else:
            col_c_left, col_c_right = st.columns([1, 1])

            with col_c_left:
                if is_contract_editing:
                    st.markdown("#### 🛠️ 編輯模式：{P1}~{PN} 變數代理草稿區")
                    contract_draft = st.session_state.get(contract_draft_key, curr_contract)
                    
                    c_name_val = st.text_input("契約範本顯示名稱", value=contract_draft.get('name', ''), key=f"c_name_in_{curr_cid}")
                    contract_draft['name'] = c_name_val
                    
                    mappings = contract_draft.get('param_mappings', {})
                    updated_mappings = {}

                    for p_tag, p_info in mappings.items():
                        with st.container(border=True):
                            st.markdown(f"**📌 參數標籤 `{{{p_tag}}}`** — `{p_info.get('label', '填空欄位')}`")
                            
                            curr_db_k = p_info.get('db_key', 'client_name')
                            curr_tbl = get_table_for_key(curr_db_k)
                            
                            tbl_list = list(DB_TABLE_FIELDS.keys())
                            c_t_idx = tbl_list.index(curr_tbl) if curr_tbl in tbl_list else 0
                            
                            c_col1, c_col2 = st.columns([1.5, 1.5])
                            with c_col1:
                                sel_tbl = st.selectbox("1️⃣ 選取 DB 資料表", tbl_list, index=c_t_idx, key=f"eppp_tbl_{curr_cid}_{p_tag}")
                            with c_col2:
                                tbl_fmap = DB_TABLE_FIELDS[sel_tbl]
                                f_keys = list(tbl_fmap.keys())
                                c_k_idx = f_keys.index(curr_db_k) if curr_db_k in f_keys else 0
                                sel_fkey = st.selectbox("2️⃣ 綁定目標欄位", f_keys, index=c_k_idx, format_func=lambda x: tbl_fmap[x], key=f"eppp_fkey_{curr_cid}_{p_tag}")
                            
                            updated_mappings[p_tag] = {
                                "label": p_info.get('label', '填空欄位'),
                                "db_table": sel_tbl,
                                "db_key": sel_fkey
                            }

                    st.markdown("---")
                    if st.button("💾 確定更新此契約範本 (寫入 JSON 檔)", key=f"btn_save_c_draft_{curr_cid}", type="primary"):
                        contract_draft['param_mappings'] = updated_mappings
                        save_contract_template(contract_draft)
                        st.session_state[contract_edit_key] = False
                        st.session_state.pop(contract_draft_key, None)
                        st.success(f"🎉 契約【{contract_draft['name']}】變數綁定設定已成功寫入硬碟 JSON 保存！")
                        st.rerun()
                else:
                    st.markdown("#### ⚙️ 左側：{P1}~{PN} 變數代理標籤綁定卡片 (瀏覽模式)")
                    st.info("💡 點擊上方【✏️ 編輯此契約對照】按鈕即可進入靈活編輯草稿模式！")
                    
                    mappings = curr_contract.get('param_mappings', {})
                    for p_tag, p_info in mappings.items():
                        with st.container(border=True):
                            st.markdown(f"**📌 參數標籤 `{{{p_tag}}}`** — `{p_info.get('label', '填空欄位')}`")
                            st.caption(f"目前綁定: `{p_info.get('db_table', '未設定')}` $\rightarrow$ **`{p_info.get('db_key', '未設定')}`**")

            with col_c_right:
                st.markdown("#### 👁️ 右側：1:1 完整定型化契約預覽與套印區")

                param_values = {}
                for p_tag, p_info in curr_contract.get('param_mappings', {}).items():
                    db_k = p_info.get('db_key', '')
                    if db_k in global_stats:
                        val_raw = global_stats[db_k]
                        param_values[p_tag] = format_db_value(db_k, val_raw)
                    elif target_order and db_k in target_order:
                        val_raw = target_order.get(db_k, '')
                        param_values[p_tag] = format_db_value(db_k, val_raw)
                    else:
                        param_values[p_tag] = f"<span style='color:#D32F2F; text-decoration:underline;'>___{p_info.get('label')}___</span>"

                if curr_contract.get('template_filename', '').lower().endswith('.xlsx'):
                    contract_html = render_excel_contract_mirror(curr_contract, contract_target_order, global_stats)
                else:
                    contract_html = f"<div>預設範本</div>"

                st.iframe(contract_html, height=750)

                st.markdown("---")
                pdf_col1, pdf_col2 = st.columns([1, 1])
                with pdf_col1:
                    st.download_button(
                        "📄 一鍵導出為 PDF / 印表機套印",
                        data=contract_html.encode('utf-8'),
                        file_name=f"{curr_contract['name']}_{contract_target_order.get('case_no', 'SAMPLE') if contract_target_order else 'DEMO'}.html",
                        mime="text/html",
                        key=f"dl_c_pdf_{curr_cid}",
                        use_container_width=True
                    )
                with pdf_col2:
                    st.download_button(
                        "📊 匯出實體 .xlsx 填空檔",
                        data=b"",
                        file_name=f"{curr_contract['name']}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"dl_c_xlsx_{curr_cid}",
                        use_container_width=True,
                        disabled=True
                    )
