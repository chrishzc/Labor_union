"""
File: historical_order_workbook.py
Description: 依版本化欄位契約解析歷史訂單 workbook、Excel date system 與雙月嫂來源。
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from hashlib import sha256
from pathlib import Path
import re

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from domains.orders.lifecycle import OrderLifecycleStatus
from shared_kernel.fingerprints import fingerprint_payload


_FIELD_ALIASES = {
    "client_name": {"client_name", "name", "客戶", "客戶姓名", "姓名"},
    "case_no": {"case_no", "案件編號", "查詢序號", "訂單編號"},
    "start_date": {"start_date", "開始日期", "服務開始", "服務開始日", "實際服務開始日"},
    "end_date": {"end_date", "結束日期", "服務結束", "服務結束日", "實際服務結束日"},
    "status": {"status", "狀態", "訂單狀態", "訂單成立狀態"},
}
_STAFF_HEADER = re.compile(r"^(staff_name|服務人員|月嫂|月嫂姓名)$")
_STATUS_MAP = {
    "0": OrderLifecycleStatus.CANCELLED,
    "1": OrderLifecycleStatus.COMPLETED,
    "2": OrderLifecycleStatus.DISCUSSION,
}


@dataclass(frozen=True, slots=True)
class HistoricalCaregiverSource:
    ordinal: int
    name: str | None
    start_date: date | None
    end_date: date | None
    has_individual_interval: bool
    issue_codes: tuple[str, ...]


@dataclass(frozen=True, slots=True)
class HistoricalOrderWorkbookRow:
    source_row: int
    source_identity: str
    source_fingerprint: str
    case_no: str | None
    client_name: str | None
    asserted_status: OrderLifecycleStatus | None
    actual_start_date: date | None
    actual_end_date: date | None
    caregivers: tuple[HistoricalCaregiverSource, ...]
    issue_codes: tuple[str, ...]


@dataclass(frozen=True, slots=True)
class HistoricalOrderWorkbook:
    content_digest: str
    sheet_identity: str
    sheet_name: str
    rows: tuple[HistoricalOrderWorkbookRow, ...]


def load_historical_order_workbook(path: str | Path, sheet: str | None = None) -> HistoricalOrderWorkbook:
    workbook_path = Path(path)
    content_digest = sha256(workbook_path.read_bytes()).hexdigest()
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        selected, header, data_start = _select_sheet(workbook, sheet)
        rows = _parse_rows(selected, header, data_start, workbook.epoch, content_digest)
        sheet_name = str(selected.title)
    finally:
        workbook.close()
    return HistoricalOrderWorkbook(
        content_digest,
        sha256(sheet_name.encode("utf-8")).hexdigest(),
        sheet_name,
        rows,
    )


def parse_historical_status(value) -> OrderLifecycleStatus | None:
    if _blank(value):
        return None
    text = str(value).strip()
    if re.fullmatch(r"[012](?:\.0+)?", text):
        text = text[0]
    return _STATUS_MAP.get(text)


def parse_excel_source_date(value, epoch) -> tuple[date | None, bool]:
    if _blank(value):
        return None, False
    if isinstance(value, datetime):
        return value.date(), False
    if isinstance(value, date):
        return value, False
    if isinstance(value, (int, float)):
        try:
            converted = from_excel(value, epoch=epoch)
            return converted.date() if isinstance(converted, datetime) else converted, False
        except (TypeError, ValueError, OverflowError):
            return None, True
    try:
        return datetime.fromisoformat(str(value).strip().replace("/", "-")).date(), False
    except ValueError:
        return None, True


def _select_sheet(workbook, requested):
    if requested is not None:
        if requested not in workbook.sheetnames:
            raise ValueError("historical_order_sheet_not_found")
        candidate = _sheet_contract(workbook[requested])
        if candidate is None:
            raise ValueError("historical_order_sheet_contract_mismatch")
        return workbook[requested], *candidate
    candidates = [(sheet, _sheet_contract(sheet)) for sheet in workbook.worksheets]
    matches = [(sheet, contract) for sheet, contract in candidates if contract is not None]
    if len(matches) != 1:
        raise ValueError("historical_order_sheet_contract_not_unique")
    selected, contract = matches[0]
    return selected, *contract


def _sheet_contract(sheet):
    first_row = next(
        ((index, tuple(row)) for index, row in enumerate(sheet.iter_rows(values_only=True), start=1)
         if any(not _blank(value) for value in row)),
        None,
    )
    if first_row is None:
        return None
    row_number, first = first_row
    headers = tuple(_header(value) for value in first)
    recognized = {_field_for_header(value) for value in headers}
    if {"client_name", "case_no", "status"}.issubset(recognized):
        return headers, row_number + 1
    if len(first) == 6:
        return ("client_name", "case_no", "start_date", "end_date", "status", "staff_name"), row_number
    return None


def _parse_rows(sheet, headers, data_start, epoch, digest):
    positions = _header_positions(headers)
    rows: list[HistoricalOrderWorkbookRow] = []
    for source_row, values in enumerate(sheet.iter_rows(min_row=data_start, values_only=True), start=data_start):
        if all(_blank(value) for value in values):
            continue
        rows.append(_parse_row(values, positions, source_row, epoch, digest))
    return tuple(rows)


def _parse_row(values, positions, source_row, epoch, digest):
    start, bad_start = parse_excel_source_date(_value(values, positions.get("start_date")), epoch)
    end, bad_end = parse_excel_source_date(_value(values, positions.get("end_date")), epoch)
    issues = _row_issues(values, positions, bad_start, bad_end, start, end)
    caregivers = _caregivers(values, positions, epoch, start, end)
    payload = {
        "case_no": _text(_value(values, positions.get("case_no"))),
        "client_name": _text(_value(values, positions.get("client_name"))),
        "status": str(_value(values, positions.get("status")) or ""),
        "start_date": start.isoformat() if start else None,
        "end_date": end.isoformat() if end else None,
        "caregivers": tuple(
            (
                item.name,
                item.start_date.isoformat() if item.start_date else None,
                item.end_date.isoformat() if item.end_date else None,
            )
            for item in caregivers
        ),
    }
    return HistoricalOrderWorkbookRow(
        source_row,
        f"historical-orders:{digest}:row:{source_row}",
        fingerprint_payload(payload).value,
        _case_no(_value(values, positions.get("case_no"))),
        _text(_value(values, positions.get("client_name"))),
        parse_historical_status(_value(values, positions.get("status"))),
        start,
        end,
        caregivers,
        tuple(sorted(set(issues + tuple(code for item in caregivers for code in item.issue_codes)))),
    )


def _header_positions(headers):
    positions: dict[str, object] = {"staff": [], "staff_start": {}, "staff_end": {}}
    for index, header in enumerate(headers):
        field = _field_for_header(header)
        if field and field not in positions:
            positions[field] = index
            continue
        staff_match = _STAFF_HEADER.fullmatch(header)
        if staff_match:
            positions["staff"].append(index)
            continue
    return positions


def _caregivers(values, positions, epoch, generic_start, generic_end):
    staff_positions = positions["staff"]
    result: list[HistoricalCaregiverSource] = []
    for ordinal, position in enumerate(staff_positions, start=1):
        own_start_position = positions["staff_start"].get(ordinal)
        own_end_position = positions["staff_end"].get(ordinal)
        has_own = own_start_position is not None or own_end_position is not None
        start, bad_start = parse_excel_source_date(_value(values, own_start_position), epoch) if has_own else (generic_start, False)
        end, bad_end = parse_excel_source_date(_value(values, own_end_position), epoch) if has_own else (generic_end, False)
        eligible_generic = len(staff_positions) == 1
        issues = _interval_issues(bad_start, bad_end, start, end, f"caregiver_{ordinal}")
        result.append(HistoricalCaregiverSource(ordinal, _text(_value(values, position)), start, end, has_own or eligible_generic, issues))
    return tuple(result)


def _row_issues(values, positions, bad_start, bad_end, start, end):
    issues = list(_interval_issues(bad_start, bad_end, start, end, "order"))
    if _blank(_value(values, positions.get("case_no"))):
        issues.append("historical_case_no_missing")
    if _blank(_value(values, positions.get("client_name"))):
        issues.append("historical_client_name_missing")
    if parse_historical_status(_value(values, positions.get("status"))) is None:
        issues.append("historical_status_invalid")
    return tuple(issues)


def _interval_issues(bad_start, bad_end, start, end, prefix):
    issues = []
    if bad_start:
        issues.append(f"historical_{prefix}_start_date_invalid")
    if bad_end:
        issues.append(f"historical_{prefix}_end_date_invalid")
    if start is not None and end is not None and start > end:
        issues.append(f"historical_{prefix}_date_range_invalid")
    return tuple(issues)


def _field_for_header(header):
    for field, aliases in _FIELD_ALIASES.items():
        if header in {_header(alias) for alias in aliases}:
            return field
    return None


def _header(value):
    return re.sub(r"\s+", "", str(value or "").strip().casefold())


def _value(values, position):
    return None if position is None or position >= len(values) else values[position]


def _text(value):
    return None if _blank(value) else str(value).strip()


def _case_no(value):
    text = _text(value)
    return text[:-2] if text and text.endswith(".0") and text[:-2].isdigit() else text


def _blank(value):
    return value is None or (isinstance(value, str) and not value.strip())


__all__ = [
    "HistoricalCaregiverSource",
    "HistoricalOrderWorkbook",
    "HistoricalOrderWorkbookRow",
    "load_historical_order_workbook",
    "parse_excel_source_date",
    "parse_historical_status",
]
