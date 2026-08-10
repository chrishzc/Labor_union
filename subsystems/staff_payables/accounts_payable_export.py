"""Canonical monthly Accounts Payable workbook export."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
import hashlib
from io import BytesIO
from typing import Callable, Protocol

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from domains.staff_payables.reconciliation import StaffPayableStatus
from shared_kernel.money import MoneyNTD
from shared_kernel.validation import require_canonical_text, require_positive_integer

_TEXT_MAXIMUM_LENGTH = 191
_HEADERS = (
    "月份-銀行代碼-流水號",
    "銀行名稱",
    "客戶or服務人員姓名",
    "銀行帳號",
    "銀行代號(碼)",
    "金額",
    "身分證字號(匯款到永豐才要填)",
    "案件編號",
    "匯款日期",
)
_OUTGOING_BANKS = {"31": "永豐銀行", "633": "台新銀行"}


@dataclass(frozen=True, slots=True)
class StaffPayableExportFact:
    obligation_identity: str
    case_no: str
    staff_id: int
    recipient_name: str
    bank_code: str
    bank_account: str
    amount: MoneyNTD
    payment_date: date
    status: StaffPayableStatus
    recipient_identity_card: str = ""

    def __post_init__(self) -> None:
        _validate_export_identity(self)
        require_positive_integer(self.staff_id, "staff id")
        _require_positive_money(self.amount)


@dataclass(frozen=True, slots=True)
class ClientRefundExportFact:
    obligation_identity: str
    case_no: str
    recipient_name: str
    bank_code: str
    bank_account: str
    amount: MoneyNTD
    payment_date: date
    payable: bool
    anomaly: bool
    refund_type: str = "customer_refund"

    def __post_init__(self) -> None:
        for value, field_name in (
            (self.obligation_identity, "obligation identity"),
            (self.case_no, "case number"),
            (self.recipient_name, "recipient name"),
            (self.bank_code, "bank code"),
            (self.bank_account, "bank account"),
        ):
            require_canonical_text(value, field_name, _TEXT_MAXIMUM_LENGTH)
        _require_positive_money(self.amount)
        if not isinstance(self.payable, bool) or not isinstance(self.anomaly, bool):
            raise TypeError("refund workflow flags must be bool")
        if self.refund_type not in {"customer_refund", "subsidy_return"}:
            raise ValueError("client refund type is invalid")


@dataclass(frozen=True, slots=True)
class AccountsPayableRow:
    payment_date: date
    payment_type: str
    recipient_name: str
    bank_code: str
    bank_account: str
    amount: MoneyNTD
    obligation_identities: tuple[str, ...]
    case_numbers: tuple[str, ...]
    recipient_identity_card: str = ""


@dataclass(frozen=True, slots=True)
class ArchivedWorkbook:
    absolute_path: str
    sha256: str


@dataclass(frozen=True, slots=True)
class ArchivedWorkbookRecord:
    filename: str
    absolute_path: str
    sha256: str
    size_bytes: int


@dataclass(frozen=True, slots=True)
class AccountsPayableExportReceipt:
    filename: str
    workbook_bytes: bytes
    sha256: str
    archived_path: str
    row_count: int


class StaffPayableExportSource(Protocol):
    def load(self, target_payment_date: date) -> tuple[StaffPayableExportFact, ...]: ...


class ClientRefundExportSource(Protocol):
    def load(self, target_payment_date: date) -> tuple[ClientRefundExportFact, ...]: ...


class WorkbookArchivePort(Protocol):
    def save(self, year: int, filename: str, workbook_bytes: bytes, sha256: str) -> ArchivedWorkbook: ...
    def list(self, year: int) -> tuple[ArchivedWorkbookRecord, ...]: ...


class ReadSnapshot(Protocol):
    def __enter__(self): ...
    def __exit__(self, *_): ...


class AccountsPayableExportWorkflow:
    def __init__(self, staff_source: StaffPayableExportSource, refund_source: ClientRefundExportSource, archive: WorkbookArchivePort, read_snapshot_factory: Callable[[], ReadSnapshot], clock: Callable[[], datetime]) -> None:
        self._staff_source = staff_source
        self._refund_source = refund_source
        self._archive = archive
        self._read_snapshot_factory = read_snapshot_factory
        self._clock = clock

    def export(self, target_payment_date: date) -> AccountsPayableExportReceipt:
        rows = self._load_rows(target_payment_date)
        workbook_bytes = build_accounts_payable_workbook(rows)
        digest = hashlib.sha256(workbook_bytes).hexdigest()
        filename = _workbook_filename(target_payment_date, self._clock(), digest)
        archived = self._archive.save(target_payment_date.year, filename, workbook_bytes, digest)
        if archived.sha256 != digest:
            raise RuntimeError("accounts_payable_archive_failed")
        return AccountsPayableExportReceipt(filename, workbook_bytes, digest, archived.absolute_path, len(rows))

    def query(self, target_payment_date: date) -> tuple[AccountsPayableRow, ...]:
        return self._load_rows(target_payment_date)

    def query_archive(self, year: int) -> tuple[ArchivedWorkbookRecord, ...]:
        return self._archive.list(year)

    def _load_rows(self, target_payment_date: date) -> tuple[AccountsPayableRow, ...]:
        with self._read_snapshot_factory():
            staff = self._staff_source.load(target_payment_date)
            refunds = self._refund_source.load(target_payment_date)
        return aggregate_accounts_payable_rows(staff, refunds)


def aggregate_accounts_payable_rows(staff_facts: tuple[StaffPayableExportFact, ...], refund_facts: tuple[ClientRefundExportFact, ...]) -> tuple[AccountsPayableRow, ...]:
    staff_rows = _aggregate_staff_rows(tuple(item for item in staff_facts if item.status is StaffPayableStatus.PAYABLE))
    refund_rows = tuple(_refund_row(item) for item in refund_facts if item.payable and not item.anomaly)
    return _sort_rows((*staff_rows, *refund_rows))


def _sort_rows(rows):
    return tuple(sorted(rows, key=lambda item: (item.payment_date, item.payment_type, item.recipient_name)))


def _aggregate_staff_rows(facts):
    groups = {}
    for item in facts:
        groups.setdefault((item.staff_id, item.payment_date), []).append(item)
    return tuple(_staff_row(items) for _, items in sorted(groups.items(), key=lambda item: item[0]))


def _staff_row(items):
    _validate_staff_bank_identity(items)
    first = items[0]
    return AccountsPayableRow(
        payment_date=first.payment_date,
        payment_type="staff_payable",
        recipient_name=first.recipient_name,
        bank_code=first.bank_code,
        bank_account=first.bank_account,
        amount=MoneyNTD(sum(item.amount.amount for item in items)),
        obligation_identities=tuple(sorted(item.obligation_identity for item in items)),
        case_numbers=tuple(sorted(item.case_no for item in items)),
        recipient_identity_card=first.recipient_identity_card,
    )


def _validate_staff_bank_identity(items) -> None:
    identities = {(item.recipient_name, item.bank_code, item.bank_account) for item in items}
    if len(identities) != 1:
        raise ValueError("accounts_payable_export_has_anomaly")


def _refund_row(item):
    return AccountsPayableRow(
        payment_date=item.payment_date,
        payment_type=(
            "client_subsidy_return"
            if item.refund_type == "subsidy_return"
            else "client_refund"
        ),
        recipient_name=item.recipient_name,
        bank_code=item.bank_code,
        bank_account=item.bank_account,
        amount=item.amount,
        obligation_identities=(item.obligation_identity,),
        case_numbers=(item.case_no,),
    )


def build_accounts_payable_workbook(rows: tuple[AccountsPayableRow, ...]) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "應付匯款清單"
    worksheet.append(_HEADERS)
    _style_headers(worksheet)
    for row in _transfer_rows(rows):
        worksheet.append(row)
    worksheet.freeze_panes = "A2"
    _set_column_widths(worksheet)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _transfer_rows(rows):
    serials = {bank_code: 0 for bank_code in _OUTGOING_BANKS}
    for row in rows:
        outgoing_bank = _outgoing_bank(row)
        serials[outgoing_bank] += 1
        yield (
            f"{row.payment_date.month}-{outgoing_bank}-{serials[outgoing_bank]}",
            _OUTGOING_BANKS[outgoing_bank],
            row.recipient_name,
            row.bank_account,
            row.bank_code,
            row.amount.amount,
            row.recipient_identity_card if outgoing_bank == "31" else "",
            ",".join(row.case_numbers),
            row.payment_date,
        )


def _outgoing_bank(row):
    return "31" if row.payment_type == "staff_payable" else "633"


def _style_headers(worksheet) -> None:
    for cell in worksheet[1]:
        cell.fill = PatternFill(fill_type="solid", fgColor="FFFF00")
        cell.font = Font(bold=True)


def _set_column_widths(worksheet) -> None:
    for column, width in {
        "A": 24, "B": 14, "C": 22, "D": 22, "E": 16,
        "F": 14, "G": 30, "H": 16, "I": 14,
    }.items():
        worksheet.column_dimensions[column].width = width


def _workbook_filename(target_date: date, generated_at: datetime, digest: str) -> str:
    if generated_at.tzinfo is None:
        raise ValueError("export clock must be timezone-aware")
    timestamp = generated_at.strftime("%Y%m%dT%H%M%S%f")
    return f"accounts-payable-{target_date.isoformat()}-{timestamp}-{digest[:12]}.xlsx"


def _validate_export_identity(item: StaffPayableExportFact) -> None:
    for value, field_name in ((item.obligation_identity, "obligation identity"), (item.case_no, "case number"), (item.recipient_name, "recipient name")):
        require_canonical_text(value, field_name, _TEXT_MAXIMUM_LENGTH)
    if item.status is not StaffPayableStatus.ANOMALY:
        require_canonical_text(item.bank_code, "bank code", _TEXT_MAXIMUM_LENGTH)
        require_canonical_text(item.bank_account, "bank account", _TEXT_MAXIMUM_LENGTH)


def _require_positive_money(value: MoneyNTD) -> None:
    if not isinstance(value, MoneyNTD):
        raise TypeError("export amount must be MoneyNTD")
    require_positive_integer(value.amount, "export amount")


__all__ = [
    "AccountsPayableExportReceipt", "AccountsPayableExportWorkflow", "AccountsPayableRow",
    "ArchivedWorkbook", "ArchivedWorkbookRecord", "ClientRefundExportFact",
    "ClientRefundExportSource", "ReadSnapshot", "StaffPayableExportFact",
    "StaffPayableExportSource", "WorkbookArchivePort", "aggregate_accounts_payable_rows",
    "build_accounts_payable_workbook",
]
