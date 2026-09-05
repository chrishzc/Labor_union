"""
File: weekly_operations_report_export.py
Description: 將營運週報輸出為完全對齊 Excel 模板的三分頁 XLSX。
"""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from subsystems.reporting.weekly_operations_report_query import WeeklyOperationsReport


SHEET_NAMES = ("週報案件受理總表", "補助案件統計表", "每周服務中說明")

_HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
_YELLOW_FILL = PatternFill(fill_type="solid", fgColor="FFFF00")
_LIGHT_GRAY_FILL = PatternFill(fill_type="solid", fgColor="F2F2F2")
_METRIC_FILL = PatternFill(fill_type="solid", fgColor="FEC2FB")
_BORDER_THIN = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

SERVICE_HEADERS = (
    "週數", "序號", "市府案號", "雇主", "休假模式",
    "休數", "服務開始", "服務結束", "特殊休假",
    "每週起始日 ", "每週結束日 ", "服務時數", "每周工作日數 ", "每周工時", "結案",
)


def export_weekly_operations_report(report: WeeklyOperationsReport) -> bytes:
    workbook = Workbook()

    # 1. 週報案件受理總表
    ws_case = workbook.active
    ws_case.title = SHEET_NAMES[0]
    _build_case_sheet(ws_case, report)

    # 2. 補助案件統計表
    ws_subsidy = workbook.create_sheet(SHEET_NAMES[1])
    _build_subsidy_sheet(ws_subsidy, report)

    # 3. 每周服務中說明
    ws_service = workbook.create_sheet(SHEET_NAMES[2])
    _build_service_sheet(ws_service, report)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _build_case_sheet(ws, report: WeeklyOperationsReport) -> None:
    roc_year = report.end_date.year - 1911

    # R1: 報表期間註記
    ws.append(("報表期間", report.period_label))
    ws.cell(row=1, column=1).font = Font(bold=True)
    ws.cell(row=1, column=2).font = Font(bold=True)

    # R2: 表頭第一層
    h1 = [
        "平台序號", "年度月份", None, "週報週數", "申請人",
        "推廣次數", "詢問人次", "平台案件申請數",
        "市民案件審核", None,
        "社福案件審核", None,
        "申請案件狀況", None, None, None,
        "服務天數", "每日服務時數", "預計服務開始日期", "預計服務結束日期", "服務狀況", "區域", "備註",
    ]
    ws.append(h1)

    # R3: 表頭第二層
    h2 = [
        None, None, None, None, None,
        None, None, None,
        "一般市民符合", "一般市民不符合",
        "補助市民符合", "補助市民不符合",
        "訂單成立", "洽談中", "取消", "審核不符合",
        None, None, None, None, None, None, None,
    ]
    ws.append(h2)

    # 合併雙層表頭單元格
    for col_letter in ["A", "D", "E", "F", "G", "H", "Q", "R", "S", "T", "U", "V", "W"]:
        ws.merge_cells(f"{col_letter}2:{col_letter}3")
    ws.merge_cells("B2:C3")
    ws.merge_cells("I2:J2")
    ws.merge_cells("K2:L2")
    ws.merge_cells("M2:P2")

    # 樣式表頭
    for r in (2, 3):
        for c in range(1, 24):
            cell = ws.cell(row=r, column=c)
            cell.font = Font(bold=True)
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = _BORDER_THIN

    # R4: 合計列
    p_cnt = report.summary.promotion_count if report.summary.promotion_count is not None else "未登錄"
    i_cnt = report.summary.inquiry_count if report.summary.inquiry_count is not None else "未登錄"
    g_inelig = report.summary.general_ineligible_count if report.summary.general_ineligible_count is not None else 0
    s_inelig = report.summary.subsidized_ineligible_count if report.summary.subsidized_ineligible_count is not None else 0

    r4 = [
        0, f"{roc_year}年度合計", None, "週數", "雇主",
        p_cnt, i_cnt, report.summary.application_count,
        report.summary.general_eligible_count, g_inelig,
        report.summary.subsidized_eligible_count, s_inelig,
        report.summary.order_established_count, report.summary.negotiating_count,
        report.summary.cancelled_count, report.summary.rejection_unpartitioned_count,
        None, None, None, None, None, None, None,
    ]
    ws.append(r4)
    ws.merge_cells("B4:C4")
    for c in range(1, 24):
        cell = ws.cell(row=4, column=c)
        cell.font = Font(bold=True)
        cell.fill = _YELLOW_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _BORDER_THIN

    # R5: 承上年度列
    r5 = [
        f"承上年度({roc_year - 1})", None, None, None, None,
        None, None, None, None, None, None, None,
        0, None, None, 0,
        None, None, None, None, None, None, None,
    ]
    ws.append(r5)
    ws.merge_cells("A5:E5")
    for c in range(1, 24):
        cell = ws.cell(row=5, column=c)
        cell.font = Font(bold=True)
        cell.fill = _LIGHT_GRAY_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _BORDER_THIN

    # R6+: 明細列 (依週次分組並進行 D、F、G 欄跨列合併)
    groups: list[tuple[str, list[WeeklyCaseRow]]] = []
    current_w: str | None = None
    current_list: list[WeeklyCaseRow] = []
    for r in report.case_rows:
        w = r.week_code or ""
        if w != current_w:
            if current_list:
                groups.append((current_w or "", current_list))
            current_w = w
            current_list = [r]
        else:
            current_list.append(r)
    if current_list:
        groups.append((current_w or "", current_list))

    for w_code, group_rows in groups:
        start_row = ws.max_row + 1
        metrics = report.weekly_metrics.get(w_code)
        promo_val = metrics[0] if metrics is not None else None
        inq_val = metrics[1] if metrics is not None else None

        for idx, row in enumerate(group_rows):
            p_cell_val = promo_val if idx == 0 and promo_val is not None else None
            i_cell_val = inq_val if idx == 0 and inq_val is not None else None
            w_cell_val = row.week_code if idx == 0 else ""

            ws.append([
                row.serial_number,
                row.month_label,
                row.application_date_roc or (row.application_date.isoformat() if row.application_date else ""),
                w_cell_val,
                row.applicant_name,
                p_cell_val, i_cell_val, 1,
                row.general_eligible, row.general_ineligible,
                row.subsidized_eligible, row.subsidized_ineligible,
                row.order_established, row.negotiating,
                row.cancelled, row.review_rejected,
                row.service_days if row.service_days is not None else "",
                row.service_hours_per_day if row.service_hours_per_day is not None else "",
                row.planned_start_date.isoformat() if row.planned_start_date else "",
                row.planned_end_date.isoformat() if row.planned_end_date else "",
                row.service_status,
                row.district or "",
                "",  # 備註欄保持空白供工會自行操作
            ])
            curr_r = ws.max_row
            for c in range(1, 24):
                cell = ws.cell(row=curr_r, column=c)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = _BORDER_THIN
                if c in (6, 7):
                    cell.fill = _METRIC_FILL
                    if idx == 0 and (promo_val is not None or inq_val is not None):
                        cell.font = Font(bold=True)

        end_row = ws.max_row
        if end_row > start_row:
            ws.merge_cells(f"D{start_row}:D{end_row}")
            ws.merge_cells(f"F{start_row}:F{end_row}")
            ws.merge_cells(f"G{start_row}:G{end_row}")
            ws.cell(row=start_row, column=4).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=start_row, column=6).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=start_row, column=7).alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A6"
    _auto_fit_columns(ws, min_col=1, max_col=23)


def _build_subsidy_sheet(ws, report: WeeklyOperationsReport) -> None:
    roc_year = report.end_date.year - 1911

    general_rows = []
    subsidized_rows = []
    for partition in report.subsidy_partitions:
        if partition.citizen_kind == "subsidized":
            subsidized_rows.extend(partition.rows)
        else:
            general_rows.extend(partition.rows)

    total_gen = len(general_rows)
    total_sub = len(subsidized_rows)

    # R1~R4: 看板區
    ws.append([None, None, None, None, None, f"{roc_year - 1}市民總計:", 0, "案", None, None, " "])
    ws.append([f"{roc_year}年度 服務補助案件暨經費 統計明細", None, None, None, None, f"{roc_year}市民總計:", total_gen, "案", "合計", None, None])
    ws.append([None, None, None, None, None, f"{roc_year - 1}社福總計:", 0, "案"])
    ws.append([None, None, None, None, None, f"{roc_year}社福總計:", total_sub, "案", "合計", None, None])

    ws.merge_cells("A2:E2")
    ws.merge_cells("I2:J2")
    ws.merge_cells("I4:J4")

    ws.cell(row=2, column=1).font = Font(bold=True, size=14)
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="center", vertical="center")

    # R5: 表頭
    headers = (
        "序號", "  ", " ", "年度接案", "訂單號碼",
        "起日", "訖日", "補助時數", "補助天數", "備註",
        "補助款金額", "單價", "結案/核銷", "核銷月份",
    )
    ws.append(headers)
    for c in range(1, 15):
        cell = ws.cell(row=5, column=c)
        cell.font = Font(bold=True)
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _BORDER_THIN

    # 一般市民明細 (從 R6 起)
    gen_start = 6
    for idx, r in enumerate(general_rows, start=1):
        ws.append([
            idx,
            r.hc_case_no or r.case_no,
            f"({roc_year})一般市民",
            idx,
            r.case_no,
            r.service_start.isoformat() if r.service_start else "",
            r.service_end.isoformat() if r.service_end else "",
            r.subsidy_hours,
            r.subsidy_days,
            "",  # 備註留空
            r.subsidy_amount_ntd,
            r.unit_price_ntd,
            r.reconciliation_status or "結案",
            r.claim_period_label or "",
        ])
        curr_r = ws.max_row
        for c in range(1, 15):
            cell = ws.cell(row=curr_r, column=c)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _BORDER_THIN

    gen_end = ws.max_row
    if gen_end >= gen_start:
        ws.cell(row=2, column=11, value=f"=SUM(K{gen_start}:K{gen_end})")
    else:
        ws.cell(row=2, column=11, value=0)
    ws.cell(row=2, column=11).font = Font(bold=True)

    # 社福市民區塊
    sub_start = ws.max_row + 2
    ws.append([None, None, f"社福市民總計: {total_sub}案", None, None, "社福市民總計:", total_sub, "案", "合計", None, None, None, "結案/核銷", "核銷月份"])
    ws.merge_cells(f"I{ws.max_row}:J{ws.max_row}")
    sep_row = ws.max_row
    for c in range(1, 15):
        cell = ws.cell(row=sep_row, column=c)
        cell.font = Font(bold=True)
        cell.fill = _YELLOW_FILL
        cell.border = _BORDER_THIN

    sub_rows_start = ws.max_row + 1
    for idx, r in enumerate(subsidized_rows, start=1):
        ws.append([
            idx,
            r.hc_case_no or r.case_no,
            f"({roc_year})社福補助",
            idx,
            r.case_no,
            r.service_start.isoformat() if r.service_start else "",
            r.service_end.isoformat() if r.service_end else "",
            r.subsidy_hours,
            r.subsidy_days,
            "",  # 備註留空
            r.subsidy_amount_ntd,
            r.unit_price_ntd,
            r.reconciliation_status or "結案",
            r.claim_period_label or "",
        ])
        curr_r = ws.max_row
        for c in range(1, 15):
            cell = ws.cell(row=curr_r, column=c)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _BORDER_THIN

    sub_rows_end = ws.max_row
    if sub_rows_end >= sub_rows_start:
        ws.cell(row=4, column=11, value=f"=SUM(K{sub_rows_start}:K{sub_rows_end})")
        ws.cell(row=sep_row, column=11, value=f"=SUM(K{sub_rows_start}:K{sub_rows_end})")
    else:
        ws.cell(row=4, column=11, value=0)
        ws.cell(row=sep_row, column=11, value=0)
    ws.cell(row=4, column=11).font = Font(bold=True)

    ws.freeze_panes = "A6"
    _auto_fit_columns(ws, min_col=1, max_col=14)


def _build_service_sheet(ws, report: WeeklyOperationsReport) -> None:
    # R1: 標題
    ws.append(["服務總表-案件服務中說明(每周)"])
    ws.merge_cells("A1:O1")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="left", vertical="center")

    if not report.service_rows:
        ws.append(SERVICE_HEADERS)
        for c in range(1, 16):
            cell = ws.cell(row=2, column=c)
            cell.font = Font(bold=True)
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _BORDER_THIN
        ws.freeze_panes = "A3"
        _auto_fit_columns(ws, min_col=1, max_col=15)
        return

    # 排序並按 week_code 分組
    groups: dict[str, list] = {}
    for r in report.service_rows:
        w_code = r.week_code or "1-1"
        groups.setdefault(w_code, []).append(r)

    for w_code, rows in groups.items():
        # 每一週區塊開始前插入 15 欄表頭列
        header_row_idx = ws.max_row + 1
        ws.append(SERVICE_HEADERS)
        for c in range(1, 16):
            cell = ws.cell(row=header_row_idx, column=c)
            cell.font = Font(bold=True)
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _BORDER_THIN

        data_start_row = ws.max_row + 1
        for idx, r in enumerate(rows, start=1):
            p_start_str = r.period_start_date.isoformat() if r.period_start_date else ""
            p_end_str = r.period_end_date.isoformat() if r.period_end_date else ""
            svc_start_str = r.service_start_date.isoformat() if r.service_start_date else ""
            svc_end_str = r.service_end_date.isoformat() if r.service_end_date else ""

            ws.append([
                w_code,
                idx,
                r.case_no,
                r.client_name,
                r.rest_mode or "周休二日",
                r.rest_days_count,
                svc_start_str,
                svc_end_str,
                r.special_rest or "",
                p_start_str,
                p_end_str,
                r.service_hours_per_day,
                r.weekly_work_days,
                r.weekly_hours,
                r.is_closed or ("結案" if r.completed else ""),
            ])
            curr_r = ws.max_row
            for c in range(1, 16):
                cell = ws.cell(row=curr_r, column=c)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = _BORDER_THIN

        data_end_row = ws.max_row
        if data_end_row >= data_start_row:
            if data_end_row > data_start_row:
                ws.merge_cells(f"A{data_start_row}:A{data_end_row}")
            ws.cell(row=data_start_row, column=1).alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A3"
    _auto_fit_columns(ws, min_col=1, max_col=15)


def _auto_fit_columns(worksheet, min_col: int, max_col: int) -> None:
    for c in range(min_col, max_col + 1):
        col_letter = get_column_letter(c)
        worksheet.column_dimensions[col_letter].width = 15


__all__ = ["SHEET_NAMES", "SERVICE_HEADERS", "export_weekly_operations_report"]

