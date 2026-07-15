from datetime import date
from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook

from services import subsidy_reconciliation_register as register


class FakeCursor:
    def __init__(self, rows):
        self.rows = rows
        self.executed = []

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc, traceback):
        return False

    def execute(self, sql, params):
        self.executed.append((" ".join(sql.split()), params))

    def fetchall(self):
        return self.rows


class FakeConnection:
    def __init__(self, rows):
        self.cursor_instance = FakeCursor(rows)
        self.closed = False

    def cursor(self):
        return self.cursor_instance

    def close(self):
        self.closed = True


def test_quarterly_register_uses_actual_dates_and_optional_lower_section(monkeypatch):
    connection = FakeConnection([
        {
            "case_no": "115000002", "subsidy_eligibility": "\u4e00\u822c\u5e02\u6c11",
            "actual_start_date": date(2026, 1, 3), "actual_end_date": date(2026, 2, 2),
            "service_days": 20, "service_hours_per_day": 9, "employer_name": "\u738b\u5c0f\u660e",
            "employer_address": "\u53f0\u5317\u5e02\u4e2d\u6b63\u5340", "staff_name": "\u6708\u5ac2\u7532",
            "survey_details": {"\u8eab\u5206\u8b49\u5b57\u865f": "A123456789"},
        },
        {
            "case_no": "115000001", "subsidy_eligibility": "\u88dc\u52a9\u5e02\u6c11",
            "actual_start_date": date(2026, 2, 1), "actual_end_date": date(2026, 3, 1),
            "service_days": 20, "service_hours_per_day": 9, "employer_name": "\u9673\u5c0f\u7f8e",
            "employer_address": "\u65b0\u5317\u5e02\u677f\u6a4b\u5340", "staff_name": "\u6708\u5ac2\u4e59",
            "survey_details": '{"\u8eab\u5206\u8b49\u5b57\u865f": "B223456789"}',
        },
        {
            "case_no": "115000003", "subsidy_eligibility": "\u4e00\u822c\u5e02\u6c11",
            "actual_start_date": date(2026, 4, 1), "actual_end_date": date(2026, 4, 30),
            "service_days": 20, "service_hours_per_day": 8, "employer_name": "\u4e0d\u61c9\u5165\u5217",
            "employer_address": "", "staff_name": "", "survey_details": {},
        },
    ])
    monkeypatch.setattr(register, "get_connection", lambda: connection)

    result = register.build_quarterly_subsidy_register(2026, 1)

    assert [row["\u5e02\u5e9c\u8a02\u55ae\u865f\u78bc"] for row in result["general_citizen_rows"]] == ["115000002"]
    assert [row["\u5e02\u5e9c\u8a02\u55ae\u865f\u78bc"] for row in result["subsidized_citizen_rows"]] == ["115000001"]
    assert result["general_citizen_rows"][0]["\u88dc\u52a9\u5929\u6578"] == Decimal("4.44")
    assert result["subsidized_citizen_rows"][0]["\u7c3d\u9818"] == ""
    assert connection.closed is True
    assert "INSERT" not in connection.cursor_instance.executed[0][0].upper()

    workbook = load_workbook(BytesIO(result["xlsx_bytes"]))
    worksheet = workbook["\u5206\u5b63\u6838\u92b7"]
    assert worksheet.cell(row=1, column=1).value == "\u4e00\u822c\u5e02\u6c11"
    assert worksheet.cell(row=2, column=15).value == "\u7c3d\u9818"
    assert worksheet.cell(row=3, column=7).number_format == "0.00"
    assert worksheet.cell(row=5, column=1).value == "\u88dc\u52a9\u5e02\u6c11"


def test_annual_summary_omits_subsidized_section_and_repairs_legacy_key(monkeypatch):
    legacy_key = "\u8eab\u5206\u8b49\u5b57\u865f".encode("utf-8").decode("latin1")
    monkeypatch.setattr(register, "get_connection", lambda: FakeConnection([
        {
            "case_no": "115000010", "subsidy_eligibility": "\u4e00\u822c\u5e02\u6c11",
            "actual_start_date": "2026-07-01", "actual_end_date": "2026-07-20",
            "service_days": 20, "service_hours_per_day": 8, "employer_name": "\u6797\u592a\u592a",
            "employer_address": "\u6843\u5712\u5e02", "staff_name": "\u6708\u5ac2\u4e19",
            "survey_details": {legacy_key: "C123456789"},
        },
    ]))

    result = register.build_annual_subsidy_summary(2026)
    row = result["general_citizen_rows"][0]
    assert row["\u8eab\u5206\u8b49\u5b57\u865f"] == "C123456789"
    assert result["subsidized_citizen_rows"] == []

    workbook = load_workbook(BytesIO(result["xlsx_bytes"]))
    worksheet = workbook["\u5e74\u5ea6\u7e3d\u8868"]
    values = [cell.value for cell in worksheet["A"]]
    assert "\u88dc\u52a9\u5e02\u6c11" not in values
    assert worksheet.max_column == 10


def test_invalid_quarter_is_rejected_before_database_access(monkeypatch):
    monkeypatch.setattr(
        register,
        "get_connection",
        lambda: (_ for _ in ()).throw(AssertionError("database must not be opened")),
    )

    try:
        register.build_quarterly_subsidy_register(2026, 5)
    except ValueError as exc:
        assert "quarter" in str(exc)
    else:
        raise AssertionError("invalid quarter must be rejected")


def test_register_caps_subsidy_hours_at_case_total_service_hours():
    row = register._to_register_row({
        "case_no": "115000011", "subsidy_eligibility": "一般市民",
        "actual_start_date": date(2026, 1, 1), "actual_end_date": date(2026, 1, 3),
        "service_days": 3, "service_hours_per_day": 9,
        "employer_name": "王小明", "employer_address": "台北市", "staff_name": "月嫂甲",
        "survey_details": {},
    })

    assert row["補助時數"] == Decimal("27")
    assert row["補助款金額"] == Decimal("8100")
