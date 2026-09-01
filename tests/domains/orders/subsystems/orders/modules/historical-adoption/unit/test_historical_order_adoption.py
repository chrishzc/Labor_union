"""
File: tests/domains/orders/test_historical_order_adoption.py
Description: 驗證歷史狀態、nullable日期、六欄工作簿、唯一案號匹配與 replay 契約。
"""

from __future__ import annotations

from datetime import date
from types import SimpleNamespace

from openpyxl import Workbook
from openpyxl.utils.datetime import MAC_EPOCH
import pytest

from domains.orders.historical_adoption import (
    HistoricalOrderCurrentFacts,
    HistoricalOrderOutcome,
    HistoricalOrderResult,
    HistoricalOrderSourceFacts,
    HistoricalOrderSourceStatus,
    build_historical_order_candidate,
)
from domains.orders.actual_start import calculate_service_dates
from domains.orders.lifecycle import OrderLifecycleStatus
from shared_kernel.fingerprints import fingerprint_payload
from infrastructure.mysql.historical_order_adoption_repository import (
    MySqlHistoricalOrderAdoptionRepository,
)
from infrastructure.mysql.historical_assignment_writer import (
    MySqlHistoricalAssignmentWriter,
)
from subsystems.orders.historical_adoption_workflow import (
    HistoricalOrderAdoptionRequest,
    HistoricalOrderAdoptionWorkflow,
    HistoricalPairingResolution,
)
from subsystems.orders.historical_actual_start_rebuild import (
    HistoricalActualStartRebuilder,
)
from subsystems.orders.historical_order_adoption_outbox_consumer import (
    _validate_canonical_event,
)
from subsystems.orders.historical_order_workbook import (
    load_historical_order_workbook,
    parse_historical_status,
)


_BUSINESS_DATE = date(2026, 9, 1)


def test_status_profile_accepts_only_zero_one_two():
    assert parse_historical_status(0) is HistoricalOrderSourceStatus.CANCELLED
    assert parse_historical_status("1.0") is HistoricalOrderSourceStatus.DEPOSIT_PAID
    assert parse_historical_status(2) is HistoricalOrderSourceStatus.DISCUSSION
    assert parse_historical_status(None) is None
    assert parse_historical_status(3) is None
    assert parse_historical_status("訂單完成") is None


def test_numeric_zero_status_has_a_distinct_source_fingerprint_from_blank(tmp_path):
    zero_path = _workbook(
        tmp_path,
        ["客戶姓名", "案件編號", "開始日期", "結束日期", "狀態", "月嫂姓名"],
        ["客戶甲", "CASE-1", None, None, 0, None],
    )
    blank_path = tmp_path / "historical-blank.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["客戶姓名", "案件編號", "開始日期", "結束日期", "狀態", "月嫂姓名"])
    sheet.append(["客戶甲", "CASE-1", None, None, None, None])
    workbook.save(blank_path)

    zero = load_historical_order_workbook(zero_path).rows[0]
    blank = load_historical_order_workbook(blank_path).rows[0]

    assert zero.asserted_status is HistoricalOrderSourceStatus.CANCELLED
    assert blank.asserted_status is None
    assert zero.source_fingerprint != blank.source_fingerprint


def test_six_column_workbook_distinguishes_zero_one_two_and_invalid_statuses(tmp_path):
    path = tmp_path / "historical-statuses.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["客戶姓名", "案件編號", "開始日期", "結束日期", "狀態", "月嫂姓名"])
    for index, status in enumerate((0, 1, 2, None), start=1):
        sheet.append([f"客戶{index}", f"CASE-{index}", None, None, status, None])
    workbook.save(path)

    rows = load_historical_order_workbook(path).rows

    assert tuple(row.asserted_status for row in rows) == (
        HistoricalOrderSourceStatus.CANCELLED,
        HistoricalOrderSourceStatus.DEPOSIT_PAID,
        HistoricalOrderSourceStatus.DISCUSSION,
        None,
    )


def test_valid_status_is_adopted_when_dates_are_null():
    current = _current(OrderLifecycleStatus.DISCUSSION)
    source = HistoricalOrderSourceFacts(HistoricalOrderSourceStatus.DEPOSIT_PAID, None, None)

    candidate = build_historical_order_candidate(current, source, _BUSINESS_DATE)

    assert candidate.outcome is HistoricalOrderOutcome.ADOPTED
    assert candidate.after_status is OrderLifecycleStatus.HISTORICAL_UNSERVED
    assert candidate.result is HistoricalOrderResult.HISTORICAL_UNSERVED
    assert candidate.date_patch == ()
    assert candidate.resulting_version == 4


@pytest.mark.parametrize(
    ("source_status", "expected_status"),
    (
        (HistoricalOrderSourceStatus.CANCELLED, OrderLifecycleStatus.ESTABLISHED),
        (HistoricalOrderSourceStatus.DISCUSSION, OrderLifecycleStatus.ESTABLISHED),
    ),
)
def test_dirty_cancelled_or_discussion_dates_never_become_actual_start(
    source_status,
    expected_status,
):
    current = _current(OrderLifecycleStatus.ESTABLISHED, planned_start=date(2025, 1, 2))
    source = HistoricalOrderSourceFacts(
        source_status,
        date(2025, 1, 3),
        date(2025, 1, 31),
    )

    candidate = build_historical_order_candidate(current, source, _BUSINESS_DATE)

    assert candidate.outcome is (
        HistoricalOrderOutcome.UNMATCHED_CASE
        if source_status is HistoricalOrderSourceStatus.CANCELLED
        else HistoricalOrderOutcome.ADOPTED
    )
    assert candidate.after_status is expected_status
    assert candidate.date_patch == ()
    if source_status is HistoricalOrderSourceStatus.DISCUSSION:
        assert candidate.result is HistoricalOrderResult.NOT_ADOPTED


def test_discussion_requires_blank_start_and_one_uniquely_resolved_caregiver(tmp_path):
    blank_start = _workbook(
        tmp_path,
        ["客戶姓名", "案件編號", "開始日期", "結束日期", "狀態", "月嫂姓名"],
        ["客戶甲", "CASE-1", None, None, 2, "月嫂甲"],
    )
    row = load_historical_order_workbook(blank_start).rows[0]
    repository = _Repository(row)

    preview = HistoricalOrderAdoptionWorkflow(
        repository,
        _UnitOfWork,
        _SchedulingHistoricalAssignment(),
    ).preview(row)

    assert preview.outcome is HistoricalOrderOutcome.ADOPTED
    assert preview.result is HistoricalOrderResult.MATCHING_PENDING_DEPOSIT
    assert preview.after_status == OrderLifecycleStatus.DISCUSSION.value
    assert preview.pairings[0].staff_id == 11

    nonblank_start = _workbook(
        tmp_path,
        ["客戶姓名", "案件編號", "開始日期", "結束日期", "狀態", "月嫂姓名"],
        ["客戶甲", "CASE-1", date(2026, 8, 1), date(2026, 8, 10), 2, "月嫂甲"],
    )
    nonblank_row = load_historical_order_workbook(nonblank_start).rows[0]
    nonblank_repository = _Repository(nonblank_row)

    skipped = HistoricalOrderAdoptionWorkflow(
        nonblank_repository,
        _UnitOfWork,
        _SchedulingHistoricalAssignment(),
    ).preview(nonblank_row)

    assert skipped.outcome is HistoricalOrderOutcome.UNMATCHED_CASE
    assert skipped.result is HistoricalOrderResult.NOT_ADOPTED
    assert nonblank_repository.staff_resolution_count == 0


def test_matching_pending_deposit_apply_writes_formal_matching_root_before_receipt(tmp_path):
    path = _workbook(
        tmp_path,
        ["客戶姓名", "案件編號", "開始日期", "結束日期", "狀態", "月嫂姓名"],
        ["客戶甲", "CASE-1", None, None, 2, "月嫂甲"],
    )
    row = load_historical_order_workbook(path).rows[0]
    ordering: list[str] = []

    class PersistingRepository(_Repository):
        def persist(self, request, preview, assignment_ids):
            del request, assignment_ids
            ordering.append("receipt")
            self.persist_count += 1
            return SimpleNamespace(
                outcome=preview.outcome,
                case_no=preview.case_no,
                resulting_version=preview.resulting_version,
                assignment_count=0,
                review_identity=None,
                replayed=False,
                preview_fingerprint=preview.fingerprint,
            )

    class MatchingPort:
        def __init__(self):
            self.command = None

        def ensure_pending_deposit_match(self, command):
            ordering.append("matching")
            self.command = command
            return SimpleNamespace(plan_id=81, plan_version=1, created=True)

    repository = PersistingRepository(row)
    matching = MatchingPort()
    workflow = HistoricalOrderAdoptionWorkflow(
        repository,
        _UnitOfWork,
        _SchedulingHistoricalAssignment(),
        matching_pending_deposit=matching,
    )
    preview = workflow.preview(row)

    workflow.apply(
        HistoricalOrderAdoptionRequest(
            row,
            preview.fingerprint,
            "historical-order:matching-pending",
            "test-operator",
            "adopt historical pending-deposit match",
            "historical-order:matching-pending:correlation",
        )
    )

    assert ordering == ["matching", "receipt"]
    assert matching.command.case_no == "CASE-1"
    assert matching.command.staff_id == 11
    assert matching.command.source_identity == row.source_identity


def test_historical_start_matching_hcm_plan_does_not_create_actual_start():
    current = _current(OrderLifecycleStatus.DISCUSSION, planned_start=date(2025, 1, 3))
    source = HistoricalOrderSourceFacts(
        HistoricalOrderSourceStatus.DEPOSIT_PAID,
        date(2025, 1, 3),
        date(2025, 1, 31),
    )

    candidate = build_historical_order_candidate(current, source, _BUSINESS_DATE)

    assert candidate.outcome is HistoricalOrderOutcome.ADOPTED
    assert candidate.date_patch == ()
    assert candidate.issue_codes == ()
    assert candidate.after_status is OrderLifecycleStatus.HISTORICAL_UNSERVED


def test_historical_start_different_from_hcm_plan_becomes_actual_start():
    current = _current(OrderLifecycleStatus.COMPLETED, planned_start=date(2025, 1, 2))
    source = HistoricalOrderSourceFacts(
        HistoricalOrderSourceStatus.DEPOSIT_PAID,
        date(2025, 1, 3),
        date(2025, 1, 31),
    )

    candidate = build_historical_order_candidate(current, source, _BUSINESS_DATE)

    assert candidate.date_patch == (
        ("actual_start_date", date(2025, 1, 3)),
        ("actual_end_date", date(2025, 1, 31)),
    )
    assert candidate.issue_codes == ()
    assert candidate.after_status is OrderLifecycleStatus.HISTORICAL_SERVICE_COMPLETED


def test_historical_actual_start_rebuild_uses_order_rest_days_and_holidays():
    assert calculate_service_dates(
        date(2026, 8, 8),
        3,
        "週休1日",
        (date(2026, 8, 10),),
    ) == (
        date(2026, 8, 8),
        date(2026, 8, 11),
        date(2026, 8, 12),
    )


def test_matching_hcm_plan_clears_previously_inferred_actual_start():
    current = _current(
        OrderLifecycleStatus.COMPLETED,
        planned_start=date(2025, 1, 2),
        actual_start=date(2025, 1, 2),
    )

    candidate = build_historical_order_candidate(
        current,
        HistoricalOrderSourceFacts(
            HistoricalOrderSourceStatus.DEPOSIT_PAID,
            date(2025, 1, 2),
            date(2025, 1, 31),
        ),
        _BUSINESS_DATE,
    )

    assert candidate.date_patch == (("actual_start_date", None),)


def test_valid_historical_status_overwrites_current_value_without_false_warning():
    candidate = build_historical_order_candidate(
        _current(OrderLifecycleStatus.CANCELLED),
        HistoricalOrderSourceFacts(HistoricalOrderSourceStatus.DEPOSIT_PAID, None, None),
        _BUSINESS_DATE,
    )

    assert candidate.outcome is HistoricalOrderOutcome.ADOPTED
    assert candidate.after_status is OrderLifecycleStatus.HISTORICAL_UNSERVED
    assert candidate.resulting_version == 4
    assert candidate.issue_codes == ()


def test_same_status_and_dates_do_not_repeat_order_lifecycle_version():
    current = _current(OrderLifecycleStatus.HISTORICAL_UNSERVED)
    candidate = build_historical_order_candidate(
        current,
        HistoricalOrderSourceFacts(HistoricalOrderSourceStatus.DEPOSIT_PAID, None, None),
        _BUSINESS_DATE,
    )

    assert candidate.outcome is HistoricalOrderOutcome.ADOPTED
    assert candidate.mutates_order is False
    assert candidate.resulting_version == current.lifecycle_version


def test_repository_skips_lifecycle_write_for_unchanged_adoption_preview():
    class NoSqlConnection:
        def cursor(self):
            raise AssertionError("unchanged adoption must not execute lifecycle SQL")

    preview = type(
        "Preview",
        (),
        {
            "outcome": HistoricalOrderOutcome.ADOPTED,
            "expected_version": 3,
            "resulting_version": 3,
        },
    )()

    assert (
        MySqlHistoricalOrderAdoptionRepository(NoSqlConnection())._apply_order(
            object(), preview
        )
        is None
    )


def test_repository_writes_historical_actual_dates_without_precision():
    class Cursor:
        rowcount = 1
        lastrowid = 81

        def __init__(self):
            self.calls = []

        def execute(self, statement, parameters):
            self.calls.append((statement, parameters))

        def close(self):
            return None

    class Connection:
        def __init__(self, cursor):
            self._cursor = cursor

        def cursor(self):
            return self._cursor

    cursor = Cursor()
    preview = SimpleNamespace(
        outcome=HistoricalOrderOutcome.ADOPTED,
        expected_version=3,
        resulting_version=4,
        after_status=OrderLifecycleStatus.COMPLETED,
        before_status=OrderLifecycleStatus.COMPLETED.value,
        case_no="CASE-1",
        date_patch=(
            ("actual_start_date", date(2026, 5, 6)),
            ("actual_end_date", date(2026, 5, 8)),
        ),
        issue_codes=(),
    )
    request = SimpleNamespace(
        actor="operator",
        idempotency_key="historical-row:key",
        row=SimpleNamespace(
            source_identity="historical-orders:digest:row:2",
            source_fingerprint="f" * 64,
        ),
    )

    event_id = MySqlHistoricalOrderAdoptionRepository(
        Connection(cursor)
    )._apply_order(request, preview)

    update_sql, update_parameters = cursor.calls[0]
    assert event_id == 81
    assert "actual_end_date" in update_sql
    assert "actual_start_date" in update_sql
    assert update_parameters == (
        OrderLifecycleStatus.COMPLETED,
        4,
        True,
        date(2026, 5, 6),
        True,
        date(2026, 5, 8),
        "CASE-1",
        3,
    )


def test_adopted_accounting_review_outbox_is_acknowledgeable():
    review_identity = "historical-order-review:accounting"

    _validate_canonical_event(
        {
            "receipt_id": 81,
            "intent_type": "historical_order_review_required",
            "bounded_snapshot": {"review_identity": review_identity},
        },
        {
            "id": 81,
            "outcome": "adopted",
            "review_identity": review_identity,
            "result_snapshot": {
                "service_calendar_status": "accounting_review_required"
            },
        },
        {
            "review_identity": review_identity,
            "source_event_identity": "historical-orders:row:019",
            "masked_case_identity": "***0019",
            "issue_codes": [
                "historical_accounting_service_calendar_unconfirmed"
            ],
            "evidence_snapshot": {},
        },
        review_identity,
    )


def test_columns_after_canonical_six_are_ignored(tmp_path):
    path = tmp_path / "historical.xlsx"
    workbook = Workbook()
    workbook.epoch = MAC_EPOCH
    sheet = workbook.active
    sheet.title = "任意名稱"
    sheet.append(["客戶姓名", "案件編號", "開始日期", "結束日期", "狀態", "月嫂姓名", "月嫂姓名2"])
    sheet.append(["客戶甲", "CASE-1", 1, 2, 1, "月嫂甲", "月嫂乙"])
    workbook.save(path)

    parsed = load_historical_order_workbook(path)
    row = parsed.rows[0]
    workflow = HistoricalOrderAdoptionWorkflow(_Repository(row), _UnitOfWork, _SchedulingHistoricalAssignment())
    preview = workflow.preview(row)

    assert row.actual_start_date == date(1904, 1, 2)
    assert row.actual_end_date == date(1904, 1, 3)
    assert tuple(item.name for item in row.caregivers) == ("月嫂甲",)
    assert tuple(item.resolution for item in preview.pairings) == (
        HistoricalPairingResolution.ASSIGNMENT_CANDIDATE,
    )
    assert preview.issue_codes == ()


def test_deposit_paid_row_does_not_build_completed_assignment_candidate(tmp_path):
    path = _workbook(tmp_path, ["客戶姓名", "案件編號", "開始日期", "結束日期", "狀態", "月嫂姓名"], ["客戶甲", "CASE-1", 45937, 45978, 1, "月嫂甲"])
    row = load_historical_order_workbook(path).rows[0]

    preview = HistoricalOrderAdoptionWorkflow(_Repository(row), _UnitOfWork, _SchedulingHistoricalAssignment()).preview(row)

    assert preview.outcome is HistoricalOrderOutcome.ADOPTED
    assert preview.pairings[0].resolution is HistoricalPairingResolution.ASSIGNMENT_CANDIDATE


def test_unique_case_number_adopts_client_name_suffix_with_review_evidence(tmp_path):
    path = _workbook(
        tmp_path,
        ["客戶姓名", "案件編號", "狀態"],
        ["客戶甲-2", "CASE-1", 1],
    )
    row = load_historical_order_workbook(path).rows[0]

    class CaseNumberRepository(_Repository):
        def load_order(self, case_no, client_name, *, for_update):
            del client_name, for_update
            return _current(OrderLifecycleStatus.DISCUSSION) if case_no == "CASE-1" else None

    preview = HistoricalOrderAdoptionWorkflow(
        CaseNumberRepository(row), _UnitOfWork, _SchedulingHistoricalAssignment()
    ).preview(row)

    assert preview.outcome is HistoricalOrderOutcome.UNMATCHED_CASE
    assert preview.issue_codes == ()


def test_repository_loads_a_unique_case_without_using_source_client_name():
    class Cursor:
        def __init__(self):
            self.statement = ""
            self.parameters = ()

        def execute(self, statement, parameters):
            self.statement = statement
            self.parameters = parameters

        def fetchall(self):
            return [{
                "case_no": "CASE-1",
                "name": "客戶甲",
                "status": OrderLifecycleStatus.DISCUSSION.value,
                "lifecycle_version": 3,
                "start_date": None,
                "actual_start_date": None,
                "actual_end_date": None,
            }]

        def close(self):
            return None

    class Connection:
        def __init__(self):
            self.cursor_value = Cursor()

        def cursor(self):
            return self.cursor_value

    connection = Connection()

    current = MySqlHistoricalOrderAdoptionRepository(connection).load_order(
        "CASE-1", "客戶甲-2", for_update=True
    )

    assert current is not None
    assert current.client_name == "客戶甲"
    assert connection.cursor_value.parameters == ("CASE-1",)
    assert "c.name=%s" not in connection.cursor_value.statement
    assert connection.cursor_value.statement.endswith("FOR UPDATE")


def test_existing_assignment_never_builds_duplicate_assignment_candidate(tmp_path):
    path = _workbook(
        tmp_path,
        ["客戶姓名", "案件編號", "開始日期", "結束日期", "狀態", "月嫂姓名"],
        ["客戶甲", "CASE-1", 45937, 45978, 1, "月嫂甲"],
    )
    row = load_historical_order_workbook(path).rows[0]

    class RepositoryWithExistingAssignment(_Repository):
        def active_assignments(self, case_no, *, for_update):
            del case_no, for_update
            return ({"id": 91, "staff_id": 11, "status": "completed"},)

    preview = HistoricalOrderAdoptionWorkflow(
        RepositoryWithExistingAssignment(row), _UnitOfWork, _SchedulingHistoricalAssignment()
    ).preview(row)

    assert preview.pairings[0].resolution is HistoricalPairingResolution.ASSIGNMENT_CANDIDATE
    assert preview.issue_codes == ()


def test_unmatched_case_does_not_resolve_staff_or_create_review(tmp_path):
    path = _workbook(tmp_path, ["客戶姓名", "案件編號", "狀態", "月嫂姓名"], ["不存在", "MISSING", None, "月嫂甲"])
    row = load_historical_order_workbook(path).rows[0]
    repository = _Repository(row, matched=False)

    preview = HistoricalOrderAdoptionWorkflow(repository, _UnitOfWork, _SchedulingHistoricalAssignment()).preview(row)

    assert preview.outcome is HistoricalOrderOutcome.UNMATCHED_CASE
    assert preview.pairings == ()
    assert repository.staff_resolution_count == 0


def test_unmatched_case_apply_is_a_zero_write_skip(tmp_path):
    path = _workbook(
        tmp_path,
        ["客戶姓名", "案件編號", "狀態", "月嫂姓名"],
        ["不存在", "MISSING", 1, "月嫂甲"],
    )
    row = load_historical_order_workbook(path).rows[0]
    repository = _Repository(row, matched=False)
    workflow = HistoricalOrderAdoptionWorkflow(repository, _UnitOfWork, _SchedulingHistoricalAssignment())
    preview = workflow.preview(row)
    request = HistoricalOrderAdoptionRequest(
        row,
        preview.fingerprint,
        "historical-order:unmatched",
        "test-operator",
        "verify unmatched skip",
        "historical-order:unmatched:correlation",
    )

    receipt = workflow.apply(request)

    assert receipt.outcome is HistoricalOrderOutcome.UNMATCHED_CASE
    assert receipt.replayed is False
    assert repository.persist_count == 0


def test_completed_historical_actual_start_skips_formal_rebuild(tmp_path):
    path = _workbook(
        tmp_path,
        ["客戶姓名", "案件編號", "開始日期", "結束日期", "狀態", "月嫂姓名"],
        ["客戶甲", "CASE-1", date(2026, 8, 7), date(2026, 9, 7), 1, "月嫂甲"],
    )
    row = load_historical_order_workbook(path).rows[0]

    class PersistingRepository(_Repository):
        def load_order(self, case_no, client_name, *, for_update):
            del case_no, client_name, for_update
            return _current(
                OrderLifecycleStatus.DISCUSSION,
                planned_start=date(2026, 8, 6),
            )

        def persist(self, request, preview, assignment_ids):
            self.persist_count += 1
            return SimpleNamespace(
                outcome=preview.outcome,
                case_no=preview.case_no,
                resulting_version=preview.resulting_version,
                assignment_count=len(assignment_ids),
                review_identity=None,
                replayed=False,
                preview_fingerprint=preview.fingerprint,
            )

    class Rebuilder:
        def __init__(self):
            self.calls = []
            self.preview_calls = []

        def preview(self, **values):
            self.preview_calls.append(values)
            return None

        def apply_in_current_unit_of_work(self, **values):
            self.calls.append(values)

    repository = PersistingRepository(row)
    rebuilder = Rebuilder()
    workflow = HistoricalOrderAdoptionWorkflow(
        repository,
        _UnitOfWork,
        _SchedulingHistoricalAssignment(),
        rebuilder,
    )
    preview = workflow.preview(row)
    assert preview.pairings[0].resolution is HistoricalPairingResolution.ASSIGNMENT_CANDIDATE
    assert preview.date_patch == (
        ("actual_start_date", date(2026, 8, 7)),
        ("actual_end_date", date(2026, 9, 7)),
    )
    assert preview.issue_codes == ()

    workflow.apply(
        HistoricalOrderAdoptionRequest(
            row,
            preview.fingerprint,
            "historical-order:actual-start-rebuild",
            "test-operator",
            "verify formal rebuild delegation",
            "historical-order:actual-start-rebuild:correlation",
        )
    )

    assert repository.persist_count == 1
    assert rebuilder.preview_calls == []
    assert rebuilder.calls == []


def test_historical_rebuilder_passes_recalculated_dates_to_canonical_actual_start():
    class Planner:
        def calculate(self, case_no, actual_start_date, *, for_update):
            assert (case_no, actual_start_date, for_update) == (
                "CASE-1",
                date(2026, 8, 8),
                True,
            )
            return (date(2026, 8, 8), date(2026, 8, 11))

    class ActualStart:
        def __init__(self):
            self.applied = []

        def replay_from_immutable_source(self, _idempotency_key):
            return None

        def preview(self, case_no, actual_start_date, *, recalculated_service_dates):
            assert (case_no, actual_start_date, recalculated_service_dates) == (
                "CASE-1",
                date(2026, 8, 8),
                (date(2026, 8, 8), date(2026, 8, 11)),
            )
            return SimpleNamespace(
                order_version=4,
                scheduling_version=5,
                client_finance_version=6,
                payroll_version=7,
                fingerprint=fingerprint_payload({"preview": "historical"}),
            )

        def apply_in_current_unit_of_work(self, request, *, recalculated_service_dates):
            self.applied.append((request, recalculated_service_dates))

    actual_start = ActualStart()
    HistoricalActualStartRebuilder(actual_start, Planner()).apply_in_current_unit_of_work(
        case_no="CASE-1",
        actual_start_date=date(2026, 8, 8),
        source_identity="historical-orders:digest:row:5",
        actor="test-operator",
        correlation_id="historical-rebuild-correlation",
    )

    request, service_dates = actual_start.applied[0]
    assert request.new_actual_start_date == date(2026, 8, 8)
    assert request.expected_order_version.value == 4
    assert request.idempotency_key.value.startswith("historical-actual-start:")
    assert service_dates == (date(2026, 8, 8), date(2026, 8, 11))


def test_same_source_and_fingerprint_replays_across_operator_metadata(tmp_path):
    path = _workbook(
        tmp_path,
        ["客戶姓名", "案件編號", "狀態"],
        ["客戶甲", "CASE-1", 1],
    )
    row = load_historical_order_workbook(path).rows[0]
    repository = _Repository(row)
    workflow = HistoricalOrderAdoptionWorkflow(repository, _UnitOfWork, _SchedulingHistoricalAssignment())
    preview = workflow.preview(row)
    repository.receipt = {
        "command_fingerprint": fingerprint_payload({
            "source_identity": row.source_identity,
            "source_fingerprint": row.source_fingerprint,
        }).value,
        "outcome": "adopted",
        "case_no": "CASE-1",
        "resulting_version": 4,
        "assignment_count": 0,
        "review_identity": None,
        "preview_fingerprint": preview.fingerprint.value,
    }
    request = HistoricalOrderAdoptionRequest(
        row,
        preview.fingerprint,
        "historical-order:replay",
        "different-operator",
        "different-reason",
        "historical-order:replay:correlation",
    )

    receipt = workflow.apply(request)

    assert receipt.replayed is True
    assert repository.persist_count == 0


def test_replayed_historical_actual_start_does_not_run_precision(tmp_path):
    path = _workbook(
        tmp_path,
        ["客戶姓名", "案件編號", "開始日期", "結束日期", "狀態", "月嫂姓名"],
        ["客戶甲", "CASE-1", date(2026, 8, 8), date(2026, 9, 7), 1, "月嫂甲"],
    )
    row = load_historical_order_workbook(path).rows[0]

    class ExistingReceiptRepository(_Repository):
        def load_order(self, case_no, client_name, *, for_update):
            del case_no, client_name, for_update
            return _current(
                OrderLifecycleStatus.COMPLETED,
                planned_start=date(2026, 8, 7),
                actual_start=date(2026, 8, 8),
            )

    class Rebuilder:
        def __init__(self):
            self.calls = []

        def preview(self, **_values):
            return None

        def apply_in_current_unit_of_work(self, **values):
            self.calls.append(values)

    repository = ExistingReceiptRepository(row)
    rebuilder = Rebuilder()
    workflow = HistoricalOrderAdoptionWorkflow(
        repository,
        _UnitOfWork,
        _SchedulingHistoricalAssignment(),
        rebuilder,
    )
    preview = workflow.preview(row)
    assert preview.pairings[0].resolution is HistoricalPairingResolution.ASSIGNMENT_CANDIDATE
    assert preview.date_patch == (("actual_end_date", date(2026, 9, 7)),)
    repository.receipt = {
        "command_fingerprint": fingerprint_payload({
            "source_identity": row.source_identity,
            "source_fingerprint": row.source_fingerprint,
        }).value,
        "outcome": "adopted",
        "case_no": "CASE-1",
        "resulting_version": 4,
        "assignment_count": 0,
        "review_identity": None,
        "preview_fingerprint": preview.fingerprint.value,
    }
    receipt = workflow.apply(
        HistoricalOrderAdoptionRequest(
            row,
            preview.fingerprint,
            "historical-order:replay-predelegation",
            "test-operator",
            "repair predelegation historical receipt",
            "historical-order:replay-predelegation:correlation",
        )
    )

    assert receipt.replayed is True
    assert repository.persist_count == 0
    assert rebuilder.calls == []


def test_replayed_historical_actual_start_does_not_repeat_completed_rebuild(tmp_path):
    path = _workbook(
        tmp_path,
        ["客戶姓名", "案件編號", "開始日期", "結束日期", "狀態"],
        ["客戶甲", "CASE-1", date(2026, 8, 8), None, 1],
    )
    row = load_historical_order_workbook(path).rows[0]

    class ExistingReceiptRepository(_Repository):
        def load_order(self, case_no, client_name, *, for_update):
            del case_no, client_name, for_update
            return _current(
                OrderLifecycleStatus.COMPLETED,
                planned_start=date(2026, 8, 7),
                actual_start=date(2026, 8, 8),
                actual_end=date(2026, 9, 7),
            )

    class Rebuilder:
        def __init__(self):
            self.calls = []

        def preview(self, **_values):
            return None

        def apply_in_current_unit_of_work(self, **values):
            self.calls.append(values)

    repository = ExistingReceiptRepository(row)
    rebuilder = Rebuilder()
    workflow = HistoricalOrderAdoptionWorkflow(
        repository,
        _UnitOfWork,
        _SchedulingHistoricalAssignment(),
        rebuilder,
    )
    preview = workflow.preview(row)
    repository.receipt = {
        "command_fingerprint": fingerprint_payload({
            "source_identity": row.source_identity,
            "source_fingerprint": row.source_fingerprint,
        }).value,
        "outcome": "adopted",
        "case_no": "CASE-1",
        "resulting_version": 4,
        "assignment_count": 0,
        "review_identity": None,
        "preview_fingerprint": preview.fingerprint.value,
    }

    receipt = workflow.apply(
        HistoricalOrderAdoptionRequest(
            row,
            preview.fingerprint,
            "historical-order:replay-complete",
            "test-operator",
            "do not repeat completed historical rebuild",
            "historical-order:replay-complete:correlation",
        )
    )

    assert receipt.replayed is True
    assert repository.persist_count == 0
    assert rebuilder.calls == []


def test_multiple_matching_sheets_fail_closed(tmp_path):
    path = tmp_path / "ambiguous.xlsx"
    workbook = Workbook()
    for index, name in enumerate(("A", "B")):
        sheet = workbook.active if index == 0 else workbook.create_sheet()
        sheet.title = name
        sheet.append(["客戶姓名", "案件編號", "狀態"])
        sheet.append(["客戶甲", f"CASE-{index}", 1])
    workbook.save(path)

    with pytest.raises(ValueError, match="historical_order_sheet_contract_not_unique"):
        load_historical_order_workbook(path)


def _current(status, planned_start=None, actual_start=None, actual_end=None):
    return HistoricalOrderCurrentFacts(
        "CASE-1",
        "客戶甲",
        status,
        3,
        planned_start,
        actual_start,
        actual_end,
    )


def _workbook(tmp_path, headers, values):
    path = tmp_path / "historical.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "來源"
    sheet.append(headers)
    sheet.append(values)
    workbook.save(path)
    return path


class _Repository:
    def __init__(self, row, *, matched=True):
        self._matched = matched
        self._row = row
        self.staff_resolution_count = 0
        self.persist_count = 0
        self.receipt = None

    def load_order(self, case_no, client_name, *, for_update):
        if not self._matched or (case_no, client_name) != ("CASE-1", "客戶甲"):
            return None
        return _current(OrderLifecycleStatus.DISCUSSION)

    def resolve_staff(self, name, *, for_update):
        self.staff_resolution_count += 1
        return {"月嫂甲": (11,), "月嫂乙": (12,)}.get(name, ())

    def active_assignments(self, case_no, *, for_update):
        return ()

    def find_receipt(self, key, source_identity):
        return self.receipt

    def persist(self, request, preview, assignment_ids):
        self.persist_count += 1
        raise AssertionError("unit preview must not persist")


class _SchedulingHistoricalAssignment:
    def append_completed_assignments(self, case_no, assignments):
        return ()


def test_historical_assignment_writer_uses_borrowed_connection_without_commit():
    class Cursor:
        lastrowid = 91

        def __init__(self):
            self.calls = []

        def __enter__(self):
            return self

        def __exit__(self, *_):
            return False

        def execute(self, statement, parameters):
            self.calls.append((statement, parameters))

        def fetchone(self):
            return {"last_sequence": 3}

    class Connection:
        def __init__(self, cursor):
            self.cursor_value = cursor
            self.commits = 0

        def cursor(self):
            return self.cursor_value

        def commit(self):
            self.commits += 1

    cursor = Cursor()
    connection = Connection(cursor)
    writer = MySqlHistoricalAssignmentWriter(connection)

    assert writer.append_completed_assignments(
        "CASE-1", ((11, date(2026, 1, 2), date(2026, 1, 4)),)
    ) == (91,)
    assert connection.commits == 0
    assert cursor.calls == [
        (
            "SELECT COALESCE(MAX(assignment_sequence),0) AS last_sequence "
            "FROM case_staff_assignments WHERE case_no=%s FOR UPDATE",
            ("CASE-1",),
        ),
        (
            "INSERT INTO case_staff_assignments "
            "(case_no,staff_id,assignment_sequence,assigned_start_date,assigned_end_date,"
            "original_assigned_start_date,original_assigned_end_date,status) "
            "VALUES (%s,%s,%s,%s,%s,%s,%s,'completed')",
            ("CASE-1", 11, 4, date(2026, 1, 2), date(2026, 1, 4), date(2026, 1, 2), date(2026, 1, 4)),
        ),
    ]


class _UnitOfWork:
    def __enter__(self):
        return self

    def __exit__(self, exception_type, exception, traceback):
        return False

    def commit(self):
        return None
