"""
File: test_contract_renderer.py
Description: 驗證契約 renderer port、相容 XLSX 填值與公式型文字的 literal 安全契約。
"""

from io import BytesIO
import json
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from subsystems.contract_signing.contract_renderer import (
    ContractRenderer,
    ContractRendererError,
    RenderedContract,
    render_contract_template,
)


def test_renderer_fills_only_declared_snapshot_values(tmp_path):
    template = tmp_path / "template.xlsx"
    workbook = Workbook()
    workbook.active["A1"] = "unchanged"
    workbook.save(template)
    mapping = tmp_path / "mapping.json"
    mapping.write_text(json.dumps({"param_mappings": {"B2": {"db_key": "case_no"}, "C3": {"db_key": "pending"}}}), encoding="utf-8")

    rendered = render_contract_template(template_path=template, mapping_path=mapping, facts={"case_no": "CASE-1"})
    worksheet = load_workbook(BytesIO(rendered)).active

    assert worksheet["A1"].value == "unchanged"
    assert worksheet["B2"].value == "CASE-1"
    assert worksheet["C3"].value is None


def test_renderer_writes_formula_like_facts_as_literal_text(tmp_path):
    template = tmp_path / "template.xlsx"
    workbook = Workbook()
    workbook.save(template)
    mapping = tmp_path / "mapping.json"
    mapping.write_text(
        json.dumps(
            {
                "param_mappings": {
                    "A1": {"db_key": "formula"},
                    "A2": {"db_key": "plus"},
                    "A3": {"db_key": "minus"},
                    "A4": {"db_key": "at"},
                }
            }
        ),
        encoding="utf-8",
    )
    facts = {
        "formula": "=WEBSERVICE(\"https://example.invalid\")",
        "plus": "+1+1",
        "minus": "-1+1",
        "at": "  @SUM(1,1)",
    }

    rendered = render_contract_template(
        template_path=template,
        mapping_path=mapping,
        facts=facts,
    )
    worksheet = load_workbook(BytesIO(rendered), data_only=False).active

    for row, expected in enumerate(facts.values(), start=1):
        cell = worksheet.cell(row=row, column=1)
        assert cell.value == expected
        assert cell.data_type == "s"


def test_renderer_fails_closed_for_unresolved_mapping_descriptor(tmp_path):
    template = tmp_path / "template.xlsx"
    Workbook().save(template)
    mapping = tmp_path / "mapping.json"
    mapping.write_text(
        json.dumps(
            {
                "param_mappings": {
                    "A1": {
                        "db_key": "",
                        "status": "pending",
                    }
                }
            }
        ),
        encoding="utf-8",
    )

    with pytest.raises(ContractRendererError) as captured:
        render_contract_template(
            template_path=template,
            mapping_path=mapping,
            facts={"case_no": "CASE-1"},
        )

    assert captured.value.code == "contract_pdf_required_mapping_unresolved"


def test_renderer_fails_closed_for_missing_fact_in_approved_mapping(tmp_path):
    template = tmp_path / "template.xlsx"
    Workbook().save(template)
    mapping = tmp_path / "mapping.json"
    mapping.write_text(
        json.dumps(
            {
                "id": "contract_client_copy",
                "param_mappings": {
                    "A1": {
                        "db_key": "typed_owner_fact",
                        "requiredness": "required",
                    }
                },
            }
        ),
        encoding="utf-8",
    )

    with pytest.raises(ContractRendererError) as captured:
        render_contract_template(
            template_path=template,
            mapping_path=mapping,
            facts={"case_no": "CASE-1"},
        )

    assert captured.value.code == "contract_pdf_required_mapping_missing"


def test_renderer_fails_closed_when_approved_mapping_lacks_requiredness(tmp_path):
    template = tmp_path / "template.xlsx"
    Workbook().save(template)
    mapping = tmp_path / "mapping.json"
    mapping.write_text(
        json.dumps(
            {
                "id": "contract_client_copy",
                "param_mappings": {"A1": {"db_key": "typed_owner_fact"}},
            }
        ),
        encoding="utf-8",
    )

    with pytest.raises(ContractRendererError) as captured:
        render_contract_template(
            template_path=template,
            mapping_path=mapping,
            facts={"typed_owner_fact": "value"},
        )

    assert captured.value.code == "contract_pdf_required_mapping_unresolved"


def test_rendered_contract_validates_its_pdf_contract():
    rendered = RenderedContract.from_pdf_bytes(
        content=b"%PDF-1.7\nbody\n%%EOF\n",
        filename="approved-contract.pdf",
        renderer_identity="libreoffice-headless",
    )

    assert rendered.mime_type == "application/pdf"
    assert rendered.filename == "approved-contract.pdf"
    assert len(rendered.sha256) == 64
    assert isinstance(object(), ContractRenderer) is False
