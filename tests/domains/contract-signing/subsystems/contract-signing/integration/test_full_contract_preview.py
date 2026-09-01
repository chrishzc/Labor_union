"""Contract tests for the exact-target Full Contract Query/Preview boundary."""

from __future__ import annotations

import json
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace
from datetime import UTC, datetime

import pytest
from openpyxl import Workbook, load_workbook

from subsystems.contract_signing.contract_renderer import render_contract_template
from subsystems.contract_signing.full_contract_preview import (
    ContractPreviewScope,
    FullContractOwnerProjection,
    FullContractPreviewApplication,
    FullContractPreviewError,
    _mapping_blockers,
)
from shared_kernel.clock import FixedBusinessClock
from infrastructure.mysql.contract_full_preview_repository import (
    _canonical_service_mode,
    _due_date_from_due_month,
    _load_approved_subsidy_claim,
    _project_subsidy_coverage,
    _special_holidays_text,
)


class _Repository:
    def __init__(self, projection):
        self.projection = projection

    def load_client_projection(self, case_no):
        return self.projection if self.projection and case_no == self.projection.case_no else None

    def load_staff_projection(self, case_no, assignment_id):
        return (
            self.projection
            if self.projection
            and case_no == self.projection.case_no
            and assignment_id == self.projection.assignment_id
            else None
        )


def _projection(scope=ContractPreviewScope.CLIENT, assignment_id=None):
    return FullContractOwnerProjection(
        case_no="CASE-1",
        scope=scope,
        assignment_id=assignment_id,
        facts={"case_no": "CASE-1", "typed_owner_fact": "value"},
        owner_fingerprints={"orders": "a" * 64},
    )


def _approved_mapping(tmp_path: Path):
    mapping = tmp_path / "mapping.json"
    mapping.write_text(
        json.dumps(
            {
                "id": "contract_client_copy",
                "param_mappings": {
                    "A1": {
                        "db_key": "typed_owner_fact",
                        "requiredness": "required",
                    }
                },
            }
        ),
        encoding="utf-8",
    )
    template = tmp_path / "contract.xlsx"
    Workbook().save(template)
    return mapping, template


def test_client_preview_uses_exact_target_and_returns_typed_cell_values(monkeypatch, tmp_path):
    mapping, template = _approved_mapping(tmp_path)
    monkeypatch.setattr(
        "subsystems.contract_signing.full_contract_preview.load_approved_template",
        lambda key: SimpleNamespace(
            template_key=key,
            mapping_sha256="b" * 64,
            template_sha256="c" * 64,
            template_filename=template.name,
        ),
    )
    monkeypatch.setattr(
        "subsystems.contract_signing.full_contract_preview.approved_template_mapping_path",
        lambda key: mapping,
    )
    result = FullContractPreviewApplication(
        _Repository(_projection()),
        FixedBusinessClock(datetime(2026, 9, 2, 9, 0, 0, tzinfo=UTC)),
    ).preview_client("CASE-1")

    assert result.scope is ContractPreviewScope.CLIENT
    assert result.assignment_id is None
    assert result.blockers == ()
    assert result.ready_to_print is True
    assert result.field_values == {"A1": "value"}


def test_preview_rejects_null_required_owner_fact(monkeypatch, tmp_path):
    mapping, template = _approved_mapping(tmp_path)
    mapping.write_text(
        json.dumps(
            {
                "id": "contract_client_copy",
                "param_mappings": {
                    "A1": {"db_key": "typed_owner_fact", "requiredness": "required"}
                },
            }
        ),
        encoding="utf-8",
    )
    import subsystems.contract_signing.full_contract_preview as module

    monkeypatch.setattr(module, "load_approved_template", lambda key: SimpleNamespace(
        template_key=key,
        mapping_sha256="b" * 64,
        template_sha256="c" * 64,
        template_filename=template.name,
    ))
    monkeypatch.setattr(module, "approved_template_mapping_path", lambda key: mapping)
    projection = FullContractOwnerProjection(
        case_no="CASE-1", scope=ContractPreviewScope.CLIENT, assignment_id=None,
        facts={"case_no": "CASE-1", "typed_owner_fact": None},
        owner_fingerprints={"orders": "a" * 64},
    )
    result = FullContractPreviewApplication(_Repository(projection)).preview_client("CASE-1")
    assert result.ready_to_print is False
    assert result.blockers == ("contract_pdf_required_mapping_missing",)


@pytest.mark.parametrize("value", ["週休1日", "週休2日", "連續服務"])
def test_client_service_type_is_an_exact_canonical_rest_mode(value):
    assert _canonical_service_mode(value) == value


@pytest.mark.parametrize("value", [None, "care", "居家", "週休一日"])
def test_legacy_or_ambiguous_service_type_is_not_reinterpreted(value):
    assert _canonical_service_mode(value) is None


@pytest.mark.parametrize(
    ("value", "expected"),
    [
        ('["2026-09-07", "2026-09-14"]', "2026-09-07、2026-09-14"),
        (["2026-09-07"], "2026-09-07"),
        (None, None),
        ("not-json", None),
        ('{"date":"2026-09-07"}', None),
    ],
)
def test_orders_custom_rest_dates_are_projected_as_typed_text(value, expected):
    assert _special_holidays_text(value) == expected


@pytest.mark.parametrize(
    ("value", "expected"),
    [
        ("2026/09/15", "2026-09-15"),
        ("2026/02/30", None),
        ("2026/09", None),
        ("2026-09-15", None),
        (None, None),
    ],
)
def test_legacy_due_month_only_projects_explicit_full_dates(value, expected):
    result = _due_date_from_due_month(value)
    assert (result.isoformat() if result else None) == expected


def test_conditional_unresolved_mapping_is_skipped_when_owner_says_not_applicable(tmp_path):
    mapping = tmp_path / "mapping.json"
    mapping.write_text(
        json.dumps(
            {
                "id": "contract_client_copy",
                "param_mappings": {
                    "C37": {
                        "db_key": "deposit_date",
                        "requiredness": "conditional",
                        "status": "unresolved",
                        "applicability": "floor_fee_positive",
                    }
                },
            }
        ),
        encoding="utf-8",
    )
    assert _mapping_blockers("contract_client_copy", mapping, {"floor_fee": 0}) == ()
    assert _mapping_blockers("contract_client_copy", mapping, {"floor_fee": 100}) == (
        "contract_pdf_required_mapping_unresolved",
    )


def test_subsidy_unresolved_mapping_blocks_only_for_typed_eligible_identity(tmp_path):
    mapping = tmp_path / "mapping.json"
    mapping.write_text(
        json.dumps(
            {
                "id": "contract_client_copy",
                "param_mappings": {
                    "B28": {
                        "db_key": "subsidy_hours",
                        "requiredness": "conditional",
                        "status": "unresolved",
                        "applicability": "subsidy_eligible",
                    }
                },
            }
        ),
        encoding="utf-8",
    )
    assert _mapping_blockers("contract_client_copy", mapping, {"identity_status": "待確認"}) == ()
    assert _mapping_blockers("contract_client_copy", mapping, {"identity_status": "補助市民"}) == (
        "contract_pdf_required_mapping_unresolved",
    )


def test_staff_legacy_funding_split_cells_stay_blank_and_whole_obligation_populates_totals(
    tmp_path,
):
    canonical_path = Path(__file__).resolve().parents[6] / "db/templates/contracts/contract_staff_service.json"
    canonical = json.loads(canonical_path.read_text(encoding="utf-8"))
    mapping = tmp_path / "mapping.json"
    split_cells = {}
    for cell in ("B13", "C13", "B15", "C15"):
        descriptor = canonical["param_mappings"][cell]
        assert descriptor["status"] == "not_applicable"
        assert descriptor["db_key"] == ""
        split_cells[cell] = descriptor
    split_cells.update(
        {
            "F10": {"db_key": "staff_payable_total", "requiredness": "required"},
            "B19": {"db_key": "staff_payable_total", "requiredness": "required"},
        }
    )
    mapping.write_text(
        json.dumps({"id": "contract_staff_service", "param_mappings": split_cells}),
        encoding="utf-8",
    )
    template = tmp_path / "contract.xlsx"
    Workbook().save(template)
    facts = {"staff_payable_total": 42000}

    assert _mapping_blockers("contract_staff_service", mapping, facts) == ()
    rendered = render_contract_template(
        template_path=template,
        mapping_path=mapping,
        facts=facts,
    )
    worksheet = load_workbook(BytesIO(rendered), data_only=False).active
    for cell in ("B13", "C13", "B15", "C15"):
        assert worksheet[cell].value is None
    assert worksheet["F10"].value == 42000
    assert worksheet["B19"].value == 42000


def test_real_staff_template_clears_legacy_funding_placeholders():
    root = Path(__file__).resolve().parents[6]
    template = root / "db/templates/contracts/服務人員契約.xlsx"
    mapping = root / "db/templates/contracts/contract_staff_service.json"
    descriptors = json.loads(mapping.read_text(encoding="utf-8"))["param_mappings"]
    facts = {
        descriptor["db_key"]: "測試值"
        for descriptor in descriptors.values()
        if descriptor.get("db_key") and descriptor.get("requiredness") == "required"
    }
    facts.update(
        {
            "case_no": "CASE-1", "staff_name": "服務人員", "client_name": "客戶",
            "assigned_start_date": "2026-09-01", "assigned_end_date": "2026-09-10",
            "service_days": 10, "assignment_service_days": 10, "service_time": "09:00-17:00",
            "service_type": "週休1日", "service_unit_price": 300,
            "staff_payable_total": 24000, "payroll_payment_date": "2026-09-15",
            "client_city": "新竹市", "client_address": "測試地址", "staff_phone": "0900000000",
            "contract_signed_date": "2026-09-01", "__today__": "2026-09-01",
        }
    )

    rendered = render_contract_template(
        template_path=template,
        mapping_path=mapping,
        facts=facts,
    )
    worksheet = load_workbook(BytesIO(rendered), data_only=False).active
    assert [worksheet[cell].value for cell in ("B13", "C13", "B15", "C15")] == [None, None, None, None]


def test_client_contract_payment_destination_and_floor_fee_due_date_use_client_finance_owner():
    root = Path(__file__).resolve().parents[6]
    mapping = json.loads((root / "db/templates/contracts/contract_client_copy.json").read_text(encoding="utf-8"))
    assert mapping["param_mappings"]["D36"] == {
        "label": "工會／代收付帳戶 (D36)",
        "db_table": "client_payment_destination_configuration_current (Client Finance typed current configuration)",
        "db_key": "client_payment_destination_account",
        "requiredness": "required",
        "status": "approved",
    }
    floor_fee_date = mapping["param_mappings"]["C37"]
    assert floor_fee_date["db_key"] == "deposit_due_date"
    assert floor_fee_date["status"] == "approved"


@pytest.mark.parametrize(
    ("filename", "print_area"),
    [("服務人員契約.xlsx", "'工作表1'!$A$1:$H$97"), ("contract_client_copy.xlsx", "'客戶契約'!$A$1:$G$185")],
)
def test_contract_templates_print_one_page_wide_without_horizontal_fragment_pages(filename, print_area):
    root = Path(__file__).resolve().parents[6]
    worksheet = load_workbook(root / "db/templates/contracts" / filename).active
    assert str(worksheet.print_area) == print_area
    assert worksheet.page_setup.fitToWidth == 1
    assert worksheet.page_setup.fitToHeight == 0
    assert worksheet.page_setup.scale is None
    assert worksheet.sheet_properties.pageSetUpPr.fitToPage is True


def test_client_finance_coverage_projection_uses_exact_planned_hours():
    facts = {
        "identity_status": "補助市民",
        "total_hours": 80,
        "floor_fee": 0,
    }
    owners = {"client_finance": "a" * 64}
    _project_subsidy_coverage(facts, owners)
    assert facts["subsidy_hours"] == 80
    assert owners["client_finance"] != "a" * 64


def test_client_finance_coverage_does_not_invent_subsidy_for_noneligible_identity():
    facts = {"identity_status": "非市民", "total_hours": 80, "floor_fee": 0}
    owners = {"client_finance": "a" * 64}
    _project_subsidy_coverage(facts, owners)
    assert "subsidy_hours" not in facts


class _Cursor:
    def __init__(self, rows):
        self.rows = rows

    def execute(self, *_args):
        return None

    def fetchall(self):
        return self.rows


def test_government_claim_item_projection_requires_one_exact_approved_item():
    cursor = _Cursor(
        [
            {
                "id": 7,
                "batch_id": 9,
                "assignment_id": 11,
                "staff_id": 13,
                "claimed_hours": 40,
                "unit_price": 300,
                "requested_amount": 12000,
                "approved_amount": 11800,
                "aggregate_version": 2,
            }
        ]
    )
    result = _load_approved_subsidy_claim(cursor, "CASE-1", 11)
    assert result["claimed_hours"] == 40
    assert result["approved_amount"] == 11800
    assert _load_approved_subsidy_claim(_Cursor(cursor.rows * 2), "CASE-1", 11) is None


def test_case_import_named_projection_is_the_only_multi_birth_source():
    from infrastructure.mysql.contract_full_preview_repository import _common_facts

    facts = _common_facts(
        {
            "case_no": "CASE-1",
            "survey_details": '{"特殊計費:胎數":"雙胞胎"}',
        }
    )
    assert facts["multi_birth_count"] == "雙胞胎"
    assert "survey_details" not in facts


def test_staff_preview_requires_exact_assignment_and_uses_no_client_fallback():
    application = FullContractPreviewApplication(
        _Repository(_projection(ContractPreviewScope.STAFF, 7)),
    )

    with pytest.raises(FullContractPreviewError) as captured:
        application.preview_staff("CASE-1", 8)

    assert captured.value.code == "contract_preview_target_not_found"


def test_preview_exposes_mapping_blocker(tmp_path):
    mapping, template = _approved_mapping(tmp_path)
    mapping.write_text(
        json.dumps(
            {
                "id": "contract_client_copy",
                "param_mappings": {
                    "A1": {"db_key": "typed_owner_fact"}
                },
            }
        ),
        encoding="utf-8",
    )
    import subsystems.contract_signing.full_contract_preview as module

    original_loader = module.load_approved_template
    original_mapping = module.approved_template_mapping_path
    module.load_approved_template = lambda key: SimpleNamespace(
        template_key=key,
        mapping_sha256="b" * 64,
        template_sha256="c" * 64,
        template_filename=template.name,
    )
    module.approved_template_mapping_path = lambda key: mapping
    try:
        result = FullContractPreviewApplication(_Repository(_projection())).preview_client("CASE-1")
    finally:
        module.load_approved_template = original_loader
        module.approved_template_mapping_path = original_mapping
    assert result.ready_to_print is False
    assert result.blockers == ("contract_pdf_required_mapping_unresolved",)
