"""
File: test_wp80_historical_order_adoption.py
Description: 驗證歷史狀態、nullable日期、雙月嫂 evidence、精確案件匹配與 replay 契約。
"""

from __future__ import annotations

from datetime import date

from openpyxl import Workbook
from openpyxl.utils.datetime import MAC_EPOCH
import pytest

from domains.orders.historical_adoption import (
    HistoricalOrderCurrentFacts,
    HistoricalOrderOutcome,
    HistoricalOrderSourceFacts,
    build_historical_order_candidate,
)
from domains.orders.lifecycle import OrderLifecycleStatus
from shared_kernel.fingerprints import fingerprint_payload
from subsystems.orders.historical_adoption_workflow import (
    HistoricalOrderAdoptionRequest,
    HistoricalOrderAdoptionWorkflow,
    HistoricalPairingResolution,
)
from subsystems.orders.historical_order_workbook import (
    load_historical_order_workbook,
    parse_historical_status,
)


def test_status_profile_accepts_only_zero_one_two():
    assert parse_historical_status(0) is OrderLifecycleStatus.CANCELLED
    assert parse_historical_status("1.0") is OrderLifecycleStatus.COMPLETED
    assert parse_historical_status(2) is OrderLifecycleStatus.DISCUSSION
    assert parse_historical_status(None) is None
    assert parse_historical_status(3) is None
    assert parse_historical_status("訂單完成") is None


def test_valid_status_is_adopted_when_dates_are_null():
    current = _current(OrderLifecycleStatus.DISCUSSION)
    source = HistoricalOrderSourceFacts(OrderLifecycleStatus.COMPLETED, None, None)

    candidate = build_historical_order_candidate(current, source)

    assert candidate.outcome is HistoricalOrderOutcome.ADOPTED
    assert candidate.after_status is OrderLifecycleStatus.COMPLETED
    assert candidate.date_patch == ()
    assert candidate.resulting_version == 4


def test_valid_historical_date_overwrites_current_value_without_false_warning():
    current = _current(OrderLifecycleStatus.DISCUSSION, actual_start=date(2025, 1, 2))
    source = HistoricalOrderSourceFacts(
        OrderLifecycleStatus.COMPLETED,
        date(2025, 1, 3),
        None,
    )

    candidate = build_historical_order_candidate(current, source)

    assert candidate.outcome is HistoricalOrderOutcome.ADOPTED
    assert candidate.date_patch == (("actual_start_date", date(2025, 1, 3)),)
    assert candidate.issue_codes == ()


def test_valid_historical_status_overwrites_current_value_without_false_warning():
    candidate = build_historical_order_candidate(
        _current(OrderLifecycleStatus.CANCELLED),
        HistoricalOrderSourceFacts(OrderLifecycleStatus.COMPLETED, None, None),
    )

    assert candidate.outcome is HistoricalOrderOutcome.ADOPTED
    assert candidate.after_status is OrderLifecycleStatus.COMPLETED
    assert candidate.resulting_version == 4
    assert candidate.issue_codes == ()


def test_two_caregivers_without_individual_intervals_require_assignment_review(tmp_path):
    path = tmp_path / "historical.xlsx"
    workbook = Workbook()
    workbook.epoch = MAC_EPOCH
    sheet = workbook.active
    sheet.title = "任意名稱"
    sheet.append(["客戶姓名", "案件編號", "開始日期", "結束日期", "狀態", "月嫂姓名", "月嫂姓名2"])
    sheet.append(["客戶甲", "CASE-1", 1, 2, 1, "月嫂甲", "月嫂乙"])
    workbook.save(path)

    parsed = load_historical_order_workbook(path)
    row = parsed.rows[0]
    workflow = HistoricalOrderAdoptionWorkflow(_Repository(row), _UnitOfWork)
    preview = workflow.preview(row)

    assert row.actual_start_date == date(1904, 1, 2)
    assert row.actual_end_date == date(1904, 1, 3)
    assert tuple(item.name for item in row.caregivers) == ("月嫂甲", "月嫂乙")
    assert {item.resolution for item in preview.pairings} == {
        HistoricalPairingResolution.ASSIGNMENT_CONFLICT
    }
    assert preview.issue_codes == ("historical_assignment_evidence_insufficient",)


def test_single_unique_caregiver_with_interval_builds_completed_assignment_candidate(tmp_path):
    path = _workbook(tmp_path, ["客戶姓名", "案件編號", "開始日期", "結束日期", "狀態", "月嫂姓名"], ["客戶甲", "CASE-1", 45937, 45978, 1, "月嫂甲"])
    row = load_historical_order_workbook(path).rows[0]

    preview = HistoricalOrderAdoptionWorkflow(_Repository(row), _UnitOfWork).preview(row)

    assert preview.outcome is HistoricalOrderOutcome.ADOPTED
    assert preview.pairings[0].resolution is HistoricalPairingResolution.ASSIGNMENT_CANDIDATE


def test_unmatched_case_does_not_resolve_staff_or_create_review(tmp_path):
    path = _workbook(tmp_path, ["客戶姓名", "案件編號", "狀態", "月嫂姓名"], ["不存在", "MISSING", None, "月嫂甲"])
    row = load_historical_order_workbook(path).rows[0]
    repository = _Repository(row, matched=False)

    preview = HistoricalOrderAdoptionWorkflow(repository, _UnitOfWork).preview(row)

    assert preview.outcome is HistoricalOrderOutcome.UNMATCHED_CASE
    assert preview.pairings == ()
    assert repository.staff_resolution_count == 0


def test_unmatched_case_apply_is_a_zero_write_skip(tmp_path):
    path = _workbook(
        tmp_path,
        ["客戶姓名", "案件編號", "狀態", "月嫂姓名"],
        ["不存在", "MISSING", 1, "月嫂甲"],
    )
    row = load_historical_order_workbook(path).rows[0]
    repository = _Repository(row, matched=False)
    workflow = HistoricalOrderAdoptionWorkflow(repository, _UnitOfWork)
    preview = workflow.preview(row)
    request = HistoricalOrderAdoptionRequest(
        row,
        preview.fingerprint,
        "wp80:unmatched",
        "test-operator",
        "verify unmatched skip",
        "wp80:unmatched:correlation",
    )

    receipt = workflow.apply(request)

    assert receipt.outcome is HistoricalOrderOutcome.UNMATCHED_CASE
    assert receipt.replayed is False
    assert repository.persist_count == 0


def test_same_source_and_fingerprint_replays_across_operator_metadata(tmp_path):
    path = _workbook(
        tmp_path,
        ["客戶姓名", "案件編號", "狀態"],
        ["客戶甲", "CASE-1", 1],
    )
    row = load_historical_order_workbook(path).rows[0]
    repository = _Repository(row)
    workflow = HistoricalOrderAdoptionWorkflow(repository, _UnitOfWork)
    preview = workflow.preview(row)
    repository.receipt = {
        "command_fingerprint": fingerprint_payload({
            "source_identity": row.source_identity,
            "source_fingerprint": row.source_fingerprint,
        }).value,
        "outcome": "adopted",
        "case_no": "CASE-1",
        "resulting_version": 4,
        "assignment_count": 0,
        "review_identity": None,
        "preview_fingerprint": preview.fingerprint.value,
    }
    request = HistoricalOrderAdoptionRequest(
        row,
        preview.fingerprint,
        "wp80:different-key",
        "different-operator",
        "different-reason",
        "wp80:replay:correlation",
    )

    receipt = workflow.apply(request)

    assert receipt.replayed is True
    assert repository.persist_count == 0


def test_multiple_matching_sheets_fail_closed(tmp_path):
    path = tmp_path / "ambiguous.xlsx"
    workbook = Workbook()
    for index, name in enumerate(("A", "B")):
        sheet = workbook.active if index == 0 else workbook.create_sheet()
        sheet.title = name
        sheet.append(["客戶姓名", "案件編號", "狀態"])
        sheet.append(["客戶甲", f"CASE-{index}", 1])
    workbook.save(path)

    with pytest.raises(ValueError, match="historical_order_sheet_contract_not_unique"):
        load_historical_order_workbook(path)


def _current(status, actual_start=None):
    return HistoricalOrderCurrentFacts("CASE-1", "客戶甲", status, 3, actual_start, None)


def _workbook(tmp_path, headers, values):
    path = tmp_path / "historical.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "來源"
    sheet.append(headers)
    sheet.append(values)
    workbook.save(path)
    return path


class _Repository:
    def __init__(self, row, *, matched=True):
        self._matched = matched
        self._row = row
        self.staff_resolution_count = 0
        self.persist_count = 0
        self.receipt = None

    def load_order(self, case_no, client_name, *, for_update):
        if not self._matched or (case_no, client_name) != ("CASE-1", "客戶甲"):
            return None
        return _current(OrderLifecycleStatus.DISCUSSION)

    def resolve_staff(self, name, *, for_update):
        self.staff_resolution_count += 1
        return {"月嫂甲": (11,), "月嫂乙": (12,)}.get(name, ())

    def active_assignments(self, case_no, *, for_update):
        return ()

    def find_receipt(self, key, source_identity):
        return self.receipt

    def persist(self, request, preview):
        self.persist_count += 1
        raise AssertionError("unit preview must not persist")


class _UnitOfWork:
    def __enter__(self):
        return self

    def __exit__(self, exception_type, exception, traceback):
        return False

    def commit(self):
        return None
