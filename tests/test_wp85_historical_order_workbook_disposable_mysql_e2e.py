"""
File: test_wp85_historical_order_workbook_disposable_mysql_e2e.py
Description: 在 disposable MySQL 驗證歷史訂單配對、rollback 與未知警示重試停損。
"""

from __future__ import annotations

from datetime import date
import os
from pathlib import Path
import time
from uuid import uuid4

from openpyxl import Workbook
import pytest

from domains.orders.lifecycle import OrderLifecycleStatus
from infrastructure.mysql.historical_order_adoption_repository import MySqlHistoricalOrderAdoptionRepository
from infrastructure.mysql.historical_order_workbook_import_repository import HistoricalOrderWorkbookImportRepository
from infrastructure.mysql.mysql_adapter import get_connection
from infrastructure.mysql.unit_of_work import MySqlUnitOfWork
from subsystems.anomalies.historical_order_adoption_outbox_consumer import (
    consume_historical_order_adoption_review_events,
)
from subsystems.orders.historical_adoption_workflow import HistoricalOrderAdoptionRequest, HistoricalOrderAdoptionWorkflow
from subsystems.orders.historical_order_workbook import (
    HistoricalCaregiverSource,
    HistoricalOrderWorkbookRow,
    load_historical_order_workbook,
)
from subsystems.orders.historical_order_workbook_import import (
    HistoricalOrderWorkbookConflict,
    HistoricalOrderWorkbookImportService,
)


DATABASE = os.getenv("LABOR_UNION_TEST_MYSQL_DATABASE")
pytestmark = pytest.mark.skipif(
    not DATABASE or os.getenv("DB_DATABASE") != DATABASE,
    reason="requires an explicitly configured disposable lu_test_* MySQL database",
)


def test_workbook_apply_uses_only_the_canonical_six_columns(tmp_path):
    token = uuid4().hex
    connection = get_connection()
    try:
        first_case, first_staff = _seed_case(connection, token, "single")
        second_case, second_staff = _seed_case(connection, token, "dual")
        second_staff_two = _seed_staff(connection, token, "dual-two")
        workbook_path = _write_workbook(tmp_path, first_case, first_staff, second_case, second_staff, second_staff_two)
        service = _service(connection)

        preview = service.preview(str(workbook_path))
        receipt = service.apply(str(workbook_path), f"wp85:{token}", preview.preview_fingerprint, "wp85-test", token)
        replay = service.apply(str(workbook_path), f"wp85:{token}", preview.preview_fingerprint, "wp85-test", token)

        assert receipt.adopted_count == 2
        assert receipt.assignments_created == 2
        assert replay.replayed_workbook is True
        _assert_assignment_and_evidence(connection, first_case, second_case)
    finally:
        connection.close()


def test_six_column_workbook_applies_zero_one_two_as_distinct_order_statuses(tmp_path):
    token = uuid4().hex
    connection = get_connection()
    try:
        cancelled_case, _ = _seed_case(connection, token, "zero")
        completed_case, _ = _seed_case(connection, token, "one")
        discussion_case, _ = _seed_case(connection, token, "two")
        with connection.cursor() as cursor:
            cursor.execute(
                "UPDATE orders SET status='訂單完成',lifecycle_version=1 WHERE case_no=%s",
                (discussion_case,),
            )
        connection.commit()
        workbook_path = _write_status_workbook(
            tmp_path / "statuses-012.xlsx",
            ((cancelled_case, 0), (completed_case, 1), (discussion_case, 2)),
        )
        service = _service(connection)

        preview = service.preview(str(workbook_path))
        receipt = service.apply(
            str(workbook_path), f"wp85-status-012:{token}", preview.preview_fingerprint,
            "wp85-test", token,
        )
        replay = service.apply(
            str(workbook_path), f"wp85-status-012:{token}", preview.preview_fingerprint,
            "wp85-test", token,
        )

        expected_counts = {
            "cancelled_0": 1,
            "completed_1": 1,
            "discussion_2": 1,
            "invalid_or_blank": 0,
        }
        assert preview.status_counts.as_dict() == expected_counts
        assert receipt.status_counts.as_dict() == expected_counts
        assert replay.status_counts.as_dict() == expected_counts
        assert replay.replayed_workbook is True
        assert _order_status(connection, cancelled_case) == "訂單取消"
        assert _order_status(connection, completed_case) == "訂單完成"
        assert _order_status(connection, discussion_case) == "洽談中"
        for case_no in (cancelled_case, completed_case, discussion_case):
            assert _count(connection, "order_lifecycle_state_events", case_no) == 1
            assert _count(connection, "historical_order_adoption_receipts", case_no) == 1
    finally:
        connection.close()


def test_workbook_conflict_rejects_changed_source_before_row_apply(tmp_path):
    token = uuid4().hex
    connection = get_connection()
    try:
        case_no, staff_name = _seed_case(connection, token, "conflict")
        first_path = _write_single_workbook(tmp_path / "first.xlsx", case_no, staff_name, status=1)
        changed_path = _write_single_workbook(tmp_path / "changed.xlsx", case_no, staff_name, status=2)
        service = _service(connection)
        first_preview = service.preview(str(first_path))
        service.apply(str(first_path), f"wp85-conflict:{token}", first_preview.preview_fingerprint, "wp85-test", token)

        changed_preview = service.preview(str(changed_path))
        with pytest.raises(HistoricalOrderWorkbookConflict):
            service.apply(str(changed_path), f"wp85-conflict:{token}", changed_preview.preview_fingerprint, "wp85-test", token)

        assert _order_status(connection, case_no) == "訂單完成"
    finally:
        connection.close()


def test_valid_historical_values_replace_current_values_without_false_conflict(tmp_path):
    token = uuid4().hex
    connection = get_connection()
    try:
        case_no, staff_name = _seed_case(connection, token, "replace")
        with connection.cursor() as cursor:
            cursor.execute(
                "UPDATE orders SET status='訂單取消',lifecycle_version=3,"
                "actual_start_date='2024-12-01',actual_end_date='2024-12-31' "
                "WHERE case_no=%s",
                (case_no,),
            )
        connection.commit()
        workbook = _write_single_workbook(
            tmp_path / "replace.xlsx", case_no, staff_name, status=1
        )
        service = _service(connection)
        preview = service.preview(str(workbook))

        receipt = service.apply(
            str(workbook), f"wp85-replace:{token}", preview.preview_fingerprint,
            "wp85-test", token,
        )

        assert receipt.adopted_count == 1
        assert preview.review_required_count == preview.current_conflict_count == 0
        assert receipt.review_required_count == receipt.current_conflict_count == 0
        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT status,lifecycle_version,actual_start_date,actual_end_date "
                "FROM orders WHERE case_no=%s",
                (case_no,),
            )
            assert cursor.fetchone() == {
                "status": "訂單完成",
                "lifecycle_version": 4,
                "actual_start_date": date(2025, 1, 2),
                "actual_end_date": date(2025, 1, 31),
            }
    finally:
        connection.close()


def test_row_apply_rolls_back_when_persist_fails(tmp_path, monkeypatch):
    token = uuid4().hex
    connection = get_connection()
    try:
        case_no, staff_name = _seed_case(connection, token, "rollback")
        row = load_historical_order_workbook(str(_write_single_workbook(tmp_path / "rollback.xlsx", case_no, staff_name, status=1))).rows[0]
        repository = MySqlHistoricalOrderAdoptionRepository(connection)
        workflow = HistoricalOrderAdoptionWorkflow(repository, lambda: MySqlUnitOfWork(connection))
        preview = workflow.preview(row)
        monkeypatch.setattr(repository, "_append_outbox", _raise_after_domain_writes)

        with pytest.raises(RuntimeError, match="wp85_forced_outbox_failure"):
            workflow.apply(HistoricalOrderAdoptionRequest(row, preview.fingerprint, f"wp85-rollback:{token}", "wp85-test", "rollback", token))

        assert _order_status(connection, case_no) == "洽談中"
        assert _count(connection, "order_lifecycle_state_events", case_no) == 0
        assert _count(connection, "case_staff_assignments", case_no) == 0
    finally:
        connection.close()


def test_matched_dirty_row_projects_masked_historical_order_anomaly():
    token = uuid4().hex
    connection = get_connection()
    try:
        case_no, _ = _seed_case(connection, token, "review")
        row = HistoricalOrderWorkbookRow(
            2,
            f"historical-orders:{token}:review",
            "a" * 64,
            case_no,
            _client_name(case_no),
            OrderLifecycleStatus.COMPLETED,
            None,
            None,
            (HistoricalCaregiverSource(1, f"missing-{token[:8]}", None, None, False, ("staff_missing",)),),
            ("staff_missing",),
        )
        workflow = HistoricalOrderAdoptionWorkflow(
            MySqlHistoricalOrderAdoptionRepository(connection), lambda: MySqlUnitOfWork(connection)
        )
        preview = workflow.preview(row)
        receipt = workflow.apply(
            HistoricalOrderAdoptionRequest(
                row, preview.fingerprint, f"wp85-review:{token}", "wp85-test", "review", token
            )
        )
        result = consume_historical_order_adoption_review_events(connection)

        assert result.failed_count == 0
        assert receipt.review_identity is not None
        _assert_review_alert(connection, receipt.review_identity)
        _assert_warning_task(connection, receipt.review_identity)
    finally:
        connection.close()


def test_historical_unknown_issue_obeys_one_second_three_attempt_dead_letter_policy():
    token = uuid4().hex
    raw_issue = "future_order_state:完整客戶名不得寫入錯誤"
    connection = get_connection()
    try:
        consume_historical_order_adoption_review_events(connection)
        case_no, _ = _seed_case(connection, token, "unknown")
        row = HistoricalOrderWorkbookRow(
            3,
            f"historical-orders:{token}:unknown",
            "b" * 64,
            case_no,
            _client_name(case_no),
            OrderLifecycleStatus.COMPLETED,
            None,
            None,
            (),
            (raw_issue,),
        )
        workflow = HistoricalOrderAdoptionWorkflow(
            MySqlHistoricalOrderAdoptionRepository(connection),
            lambda: MySqlUnitOfWork(connection),
        )
        preview = workflow.preview(row)
        receipt = workflow.apply(
            HistoricalOrderAdoptionRequest(
                row, preview.fingerprint, f"wp85-unknown:{token}",
                "wp85-test", "unknown", token,
            )
        )
        assert receipt.review_identity is not None

        for attempt in range(3):
            result = consume_historical_order_adoption_review_events(
                connection, maximum_events=1
            )
            assert result.failed_count == 1
            immediate = consume_historical_order_adoption_review_events(
                connection, maximum_events=1
            )
            assert immediate.failed_count == 0
            if attempt < 2:
                time.sleep(1.05)

        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT COUNT(*) AS count FROM import_warning_occurrences "
                "WHERE source_receipt_identity=%s",
                (receipt.review_identity,),
            )
            assert cursor.fetchone() == {"count": 0}
            cursor.execute(
                "SELECT COUNT(*) AS count FROM anomaly_current_alerts "
                "WHERE definition_code='HISTORICAL-ORDER-001' AND source_identity=%s",
                (receipt.review_identity,),
            )
            assert cursor.fetchone() == {"count": 0}
            cursor.execute(
                "SELECT outbox.published_at,outbox.attempts,outbox.last_error "
                "FROM historical_order_adoption_outbox outbox "
                "JOIN historical_order_adoption_receipts receipt "
                "ON receipt.id=outbox.receipt_id "
                "WHERE receipt.review_identity=%s "
                "AND outbox.intent_type='historical_order_review_required'",
                (receipt.review_identity,),
            )
            outbox = cursor.fetchone()
            failure = __import__("json").loads(outbox["last_error"])
            assert outbox["published_at"] is None
            assert outbox["attempts"] == 3
            assert failure["terminal"] == 1
            assert failure["error_code"].startswith(
                "import_warning_projection_unknown_issue:historical_order:"
            )
            assert raw_issue not in outbox["last_error"]
    finally:
        connection.close()


def test_controlled_deidentified_historical_workbook_rebuilds_and_replays():
    workbook_path = Path(__file__).parents[1] / "document" / "資料庫、資料處理" / "假資料_歷史訂單.xlsx"
    parsed = load_historical_order_workbook(str(workbook_path))
    assert len(parsed.rows) == 1
    source_row = parsed.rows[0]
    assert source_row.case_no is not None
    assert source_row.client_name is not None
    connection = get_connection()
    try:
        _seed_controlled_source_roots(connection, source_row)
        service = _service(connection)
        preview = service.preview(str(workbook_path))
        receipt = service.apply(
            str(workbook_path), "wp85-controlled-deidentified-source",
            preview.preview_fingerprint, "wp85-test", "controlled-deidentified-source",
        )
        replay = service.apply(
            str(workbook_path), "wp85-controlled-deidentified-source",
            preview.preview_fingerprint, "wp85-test", "controlled-deidentified-source",
        )

        assert preview.source_row_count == 1
        assert receipt.adopted_count == 1
        assert replay.replayed_workbook is True
        assert _order_status(connection, source_row.case_no) == "訂單完成"
    finally:
        connection.close()


def _service(connection):
    workflow = HistoricalOrderAdoptionWorkflow(
        MySqlHistoricalOrderAdoptionRepository(connection), lambda: MySqlUnitOfWork(connection)
    )
    return HistoricalOrderWorkbookImportService(HistoricalOrderWorkbookImportRepository(connection), workflow)


def _seed_controlled_source_roots(connection, source_row):
    with connection.cursor() as cursor:
        cursor.execute("SELECT id FROM clients WHERE case_no=%s", (source_row.case_no,))
        if cursor.fetchone() is None:
            cursor.execute("INSERT INTO clients (case_no,name) VALUES (%s,%s)", (source_row.case_no, source_row.client_name))
            cursor.execute(
                "INSERT INTO orders (case_no,client_id,status,lifecycle_version) VALUES (%s,%s,'洽談中',0)",
                (source_row.case_no, cursor.lastrowid),
            )
        for caregiver in source_row.caregivers:
            cursor.execute("SELECT id FROM staff WHERE name=%s", (caregiver.name,))
            if cursor.fetchone() is None:
                cursor.execute(
                    "INSERT INTO staff (name,identity_card) VALUES (%s,%s)",
                    (caregiver.name, f"W{uuid4().hex[:9]}"),
                )
    connection.commit()


def _seed_case(connection, token, suffix):
    case_no = f"WP85-{suffix}-{token[:10]}"
    client_name = f"wp85-client-{suffix}-{token[:6]}"
    staff_name = f"wp85-staff-{suffix}-{token[:6]}"
    with connection.cursor() as cursor:
        cursor.execute("INSERT INTO clients (case_no,name) VALUES (%s,%s)", (case_no, client_name))
        cursor.execute(
            "INSERT INTO orders (case_no,client_id,status,lifecycle_version) VALUES (%s,%s,'洽談中',0)",
            (case_no, cursor.lastrowid),
        )
        cursor.execute("INSERT INTO staff (name,identity_card) VALUES (%s,%s)", (staff_name, f"W{token[:9]}{suffix[:1]}"))
    connection.commit()
    return (case_no, staff_name)


def _write_workbook(tmp_path, first_case, first_staff, second_case, second_staff, second_staff_two):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "任意來源"
    sheet.append([
        "客戶姓名", "案件編號", "開始日期", "結束日期", "狀態", "月嫂姓名", "月嫂姓名2",
    ])
    sheet.append([
        _client_name(first_case), first_case, date(2025, 1, 2), date(2025, 1, 31), 1, first_staff, None,
    ])
    sheet.append([_client_name(second_case), second_case, None, None, 1, second_staff, second_staff_two, None, None])
    destination = tmp_path / "workbook.xlsx"
    workbook.save(destination)
    return destination


def _write_single_workbook(destination, case_no, staff_name, status):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "任意來源"
    sheet.append(["客戶姓名", "案件編號", "開始日期", "結束日期", "狀態", "月嫂姓名"])
    sheet.append([_client_name(case_no), case_no, date(2025, 1, 2), date(2025, 1, 31), status, staff_name])
    workbook.save(destination)
    return destination


def _write_status_workbook(destination, case_statuses):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "六欄狀態驗收"
    sheet.append(["客戶姓名", "案件編號", "開始日期", "結束日期", "狀態", "月嫂姓名"])
    for case_no, status in case_statuses:
        sheet.append([_client_name(case_no), case_no, None, None, status, None])
    workbook.save(destination)
    return destination


def _client_name(case_no):
    return f"wp85-client-{case_no.split('-')[1]}-{case_no.split('-')[2][:6]}"


def _seed_staff(connection, token, suffix):
    staff_name = f"wp85-staff-{suffix}-{token[:6]}"
    with connection.cursor() as cursor:
        cursor.execute("INSERT INTO staff (name,identity_card) VALUES (%s,%s)", (staff_name, f"W{token[:8]}{suffix[:2]}"))
    connection.commit()
    return staff_name


def _assert_assignment_and_evidence(connection, first_case, second_case):
    assert _count(connection, "case_staff_assignments", first_case) == 1
    assert _count(connection, "case_staff_assignments", second_case) == 1
    assert _count(connection, "historical_order_pairing_evidence", first_case, via_receipt=True) == 1
    assert _count(connection, "historical_order_pairing_evidence", second_case, via_receipt=True) == 1


def _assert_review_alert(connection, review_identity):
    with connection.cursor() as cursor:
        cursor.execute(
            "SELECT definition_code,source_identity,display_snapshot FROM anomaly_current_alerts "
            "WHERE definition_code='HISTORICAL-ORDER-001' AND source_identity=%s",
            (review_identity,),
        )
        alert = cursor.fetchone()
    assert alert is not None
    assert alert["definition_code"] == "HISTORICAL-ORDER-001"
    assert alert["source_identity"] == review_identity
    assert "staff_missing" in str(alert["display_snapshot"])


def _assert_warning_task(connection, review_identity):
    with connection.cursor() as cursor:
        cursor.execute(
            "SELECT occurrence.logical_code,occurrence.field_path,task.tracking_status,task.tracking_version "
            "FROM import_warning_occurrences occurrence "
            "JOIN import_warning_current_tasks task ON task.occurrence_id=occurrence.id "
            "WHERE occurrence.source_receipt_identity=%s",
            (review_identity,),
        )
        rows = cursor.fetchall()
    assert rows == [
        {
            "logical_code": "ORDER-HIST-STAFF-001",
            "field_path": "$staff",
            "tracking_status": "open",
            "tracking_version": 1,
        }
    ]


def _order_status(connection, case_no):
    with connection.cursor() as cursor:
        cursor.execute("SELECT status FROM orders WHERE case_no=%s", (case_no,))
        return cursor.fetchone()["status"]


def _count(connection, table_name, case_no, *, via_receipt=False):
    with connection.cursor() as cursor:
        if via_receipt:
            cursor.execute(
                "SELECT COUNT(*) AS count FROM historical_order_pairing_evidence evidence "
                "JOIN historical_order_adoption_receipts receipt ON receipt.id=evidence.receipt_id WHERE receipt.case_no=%s",
                (case_no,),
            )
        else:
            cursor.execute(f"SELECT COUNT(*) AS count FROM {table_name} WHERE case_no=%s", (case_no,))
        return int(cursor.fetchone()["count"])


def _raise_after_domain_writes(*_args):
    raise RuntimeError("wp85_forced_outbox_failure")
