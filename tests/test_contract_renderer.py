from io import BytesIO
import json
from pathlib import Path

from openpyxl import Workbook, load_workbook

from subsystems.contract_signing.contract_renderer import render_contract_template


def test_renderer_fills_only_declared_snapshot_values(tmp_path):
    template = tmp_path / "template.xlsx"
    workbook = Workbook()
    workbook.active["A1"] = "unchanged"
    workbook.save(template)
    mapping = tmp_path / "mapping.json"
    mapping.write_text(json.dumps({"param_mappings": {"B2": {"db_key": "case_no"}, "C3": {"db_key": "pending"}}}), encoding="utf-8")

    rendered = render_contract_template(template_path=template, mapping_path=mapping, facts={"case_no": "CASE-1"})
    worksheet = load_workbook(BytesIO(rendered)).active

    assert worksheet["A1"].value == "unchanged"
    assert worksheet["B2"].value == "CASE-1"
    assert worksheet["C3"].value is None
