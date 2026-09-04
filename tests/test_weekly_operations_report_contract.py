"""
File: test_weekly_operations_report_contract.py
Description: 驗證營運週報週界、彙總、遮罩、正式工時、strict API 與三分頁 XLSX。
"""

from datetime import date, datetime
from decimal import Decimal
from io import BytesIO

from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from api.dependencies.admin_auth import require_admin
from api.dependencies.operations_reports import get_weekly_operations_report_query
from api.exception_handlers import CorrelationBoundaryMiddleware, install_typed_error_handlers
from api.routes import operations_reports
from shared_kernel.clock import TAIPEI_TIME_ZONE
from subsystems.access.authentication_session import AdminPrincipal
from subsystems.reporting.weekly_operations_report_query import (
    SubsidyFact,
    SubsidyFacts,
    WeeklyCaseFact,
    WeeklyOperationsReportQuery,
    WeeklyServiceFact,
)
from infrastructure.mysql.weekly_operations_report_query_adapter import (
    _CASE_FACTS_SQL,
    _SERVICE_FACTS_SQL,
)


class _Facts:
    def list_case_facts(self, start_date, end_date):
        assert (start_date, end_date) == (date(2026, 8, 20), date(2026, 8, 26))
        return [
            WeeklyCaseFact(
                7, "115000007", datetime(2026, 8, 18, 9), "王小美", "一般市民", None,
                "東區", "服務中", 20, 8, date(2026, 8, 10), date(2026, 9, 4),
            ),
            WeeklyCaseFact(
                8, "115000008", datetime(2026, 8, 19, 10), "林大華", None, "資格不符",
                "北區", None, None, None, None, None,
            ),
        ]

    def list_service_facts(self, start_date, end_date):
        return [
            WeeklyServiceFact(
                31, "115000007", "王小美", "陳月嫂", date(2026, 8, 10), date(2026, 9, 4),
                8, 5, "服務中", "active",
            ),
        ]

    def list_subsidy_facts(self, start_date, end_date):
        assert (start_date, end_date) == (date(2026, 8, 20), date(2026, 8, 26))
        return SubsidyFacts(
            general=(
                SubsidyFact(
                    1, "115000007", "一般市民", date(2026, 8, 10), date(2026, 8, 20),
                    Decimal("40"), Decimal("5"), 20, 12000, 300,
                    "王小美", "陳月嫂", "A123456789", "完整地址",
                ),
            ),
            subsidized=(),
        )


def _query():
    return WeeklyOperationsReportQuery(
        _Facts(),
        lambda: datetime(2026, 8, 23, 12, tzinfo=TAIPEI_TIME_ZONE),
    )


def _app():
    app = FastAPI()
    app.include_router(operations_reports.router)
    app.add_middleware(CorrelationBoundaryMiddleware)
    install_typed_error_handlers(app)
    app.dependency_overrides[require_admin] = lambda: AdminPrincipal(7, "reports", "Reports", "admin")
    app.dependency_overrides[get_weekly_operations_report_query] = _query
    return app


def test_weekly_query_is_redacted_and_uses_official_work_days():
    response = TestClient(_app()).get(
        "/api/v1/operations-reports/weekly",
        params={"start_date": "2026-08-20", "end_date": "2026-08-26"},
    )
    assert response.status_code == 200
    data = response.json()["data"]
    assert data["schema_version"] == "operations-report.v2"
    assert data["period"] == {
        "start_date": "2026-08-20",
        "end_date": "2026-08-26",
        "timezone": "Asia/Taipei",
        "period_label": "2026-08-20 ~ 2026-08-26",
    }
    assert data["summary"]["application_count"] == 2
    assert data["summary"]["general_eligible_count"] == 1
    assert data["summary"]["rejection_unpartitioned_count"] == 1
    assert data["summary"]["promotion_count"] is None
    assert data["service_rows"][0]["weekly_work_days"] == 5
    assert data["service_rows"][0]["weekly_hours"] == 40
    assert data["case_rows"][0]["applicant_name"] == "王小美"
    assert data["subsidy_partitions"][0]["rows"][0]["identity_card"] == "A123456789"
    assert "王小美" in response.text
    assert "A123456789" in response.text
    assert "完整地址" in response.text


def test_operations_report_rejects_inverted_date_range():
    response = TestClient(_app()).get(
        "/api/v1/operations-reports/weekly",
        params={"start_date": "2026-08-26", "end_date": "2026-08-20"},
    )
    assert response.status_code == 400
    assert response.json()["detail"]["error"]["code"] == "weekly_operations_report_invalid"


def test_operations_report_rejects_legacy_week_start_parameter():
    response = TestClient(_app()).get(
        "/api/v1/operations-reports/weekly",
        params={"start_date": "2026-08-20", "end_date": "2026-08-26", "week_start": "2026-08-20"},
    )
    assert response.status_code == 400
    assert response.json()["detail"]["error"]["code"] == "weekly_operations_report_invalid"


def test_weekly_query_retains_missing_application_date_as_typed_quality_issue():
    class MissingDateFacts(_Facts):
        def list_case_facts(self, start_date, end_date):
            return [
                WeeklyCaseFact(
                    9, "OPS96-WEEKLY-D-MISSING-DATE", None, "林大華", "一般市民", None,
                    "北區", "洽談中", None, None, None, None,
                ),
            ]

    def query():
        return WeeklyOperationsReportQuery(
            MissingDateFacts(),
            lambda: datetime(2026, 8, 23, 12, tzinfo=TAIPEI_TIME_ZONE),
        )

    app = _app()
    app.dependency_overrides[get_weekly_operations_report_query] = query
    response = TestClient(app).get(
        "/api/v1/operations-reports/weekly",
        params={"start_date": "2026-08-20", "end_date": "2026-08-26"},
    )
    assert response.status_code == 200
    row = response.json()["data"]["case_rows"][0]
    assert row["application_date"] is None
    assert "application_date_missing" in row["data_quality_codes"]


def test_case_source_keeps_date_scoped_rows_inside_requested_window_only():
    assert "OR c.created_at IS NULL" not in _CASE_FACTS_SQL
    assert "c.created_at >= %s AND c.created_at < %s" in _CASE_FACTS_SQL


def test_weekly_service_source_excludes_unrestarted_historical_overlays():
    assert "JOIN scheduling_generations g" in _SERVICE_FACTS_SQL
    assert "g.effective_marker=1" in _SERVICE_FACTS_SQL
    assert "JOIN staff_schedule ss" in _SERVICE_FACTS_SQL
    assert "ss.effective_marker=1 AND ss.is_work_day=1" in _SERVICE_FACTS_SQL
    assert "historical_order" not in _SERVICE_FACTS_SQL
    assert "historical_stage" not in _SERVICE_FACTS_SQL


def test_weekly_export_has_fixed_three_sheets_and_summary_without_pii():
    response = TestClient(_app()).get(
        "/api/v1/operations-reports/weekly/export",
        params={"start_date": "2026-08-20", "end_date": "2026-08-26"},
    )
    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "2026-08-20_2026-08-26" in response.headers["content-disposition"]
    workbook = load_workbook(BytesIO(response.content), read_only=True, data_only=True)
    assert workbook.sheetnames == ["週報案件受理總表", "補助案件統計表", "每周服務中說明"]
    case_values = list(workbook["週報案件受理總表"].values)
    assert case_values[0][:2] == ("報表期間", "2026-08-20 ~ 2026-08-26")
    # Row 1 是表頭一，Row 2 是表頭二，Row 3 是合計列
    assert "平台序號" in case_values[1]
    assert "一般市民符合" in case_values[2]
    assert "未登錄" in case_values[3]  # 合計列推廣/詢問人次
    workbook_text = " ".join(str(value) for sheet in workbook for row in sheet.values for value in row if value is not None)
    assert "王小美" in workbook_text
    # 補助案件統計表對齊模板：經費統計格式，不包含身分證字號與地址個資
    assert "A123456789" not in workbook_text

    # 每周服務中說明：對齊模板 15 欄
    service_values = list(workbook["每周服務中說明"].values)
    assert service_values[0][0] == "服務總表-案件服務中說明(每周)"
    expected_headers = (
        "週數", "序號", "市府案號", "雇主", "休假模式",
        "休數", "服務開始", "服務結束", "特殊休假",
        "每週起始日 ", "每週結束日 ", "服務時數", "每周工作日數 ", "每周工時", "結案",
    )
    assert service_values[1][:15] == expected_headers


def test_weekly_export_with_promotion_and_inquiry_counts():
    response = TestClient(_app()).get(
        "/api/v1/operations-reports/weekly/export",
        params={
            "start_date": "2026-08-20",
            "end_date": "2026-08-26",
            "promotion_count": 12,
            "inquiry_count": 34,
        },
    )
    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content), read_only=True, data_only=True)
    case_values = list(workbook["週報案件受理總表"].values)
    summary_row = case_values[3]
    # F 欄為推廣次數 (index 5)，G 欄為詢問人次 (index 6)
    assert summary_row[5] == 12
    assert summary_row[6] == 34
    # 最後一欄 (W欄，index 22) 備註保持空白
    for row in case_values[5:]:
        assert row[22] == "" or row[22] is None
