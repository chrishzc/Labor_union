from datetime import date, datetime, timezone
from io import BytesIO

from openpyxl import load_workbook
import pytest

from domains.staff_payables.reconciliation import StaffPayableStatus
from shared_kernel.money import MoneyNTD
from subsystems.staff_payables.accounts_payable_export import (
    AccountsPayableExportWorkflow,
    ArchivedWorkbook,
    ClientRefundExportFact,
    GovernmentOverpaymentReturnExportFact,
    StaffPayableExportFact,
    aggregate_accounts_payable_rows,
    build_accounts_payable_workbook,
)


class _Snapshot:
    def __enter__(self):
        return self

    def __exit__(self, *_):
        return False


class _Source:
    def __init__(self, facts):
        self.facts = facts

    def load(self, target_date):
        assert target_date == date(2026, 8, 31)
        return self.facts


class _Archive:
    def __init__(self, digest=None):
        self.digest = digest
        self.calls = []

    def save(self, year, filename, content, digest):
        self.calls.append((year, filename, content, digest))
        return ArchivedWorkbook("C:/archive/" + filename, self.digest or digest)

    def list(self, year):
        assert year == 2026
        return ()


def _staff(identity, case, amount, *, status=StaffPayableStatus.PAYABLE, account="001"):
    return StaffPayableExportFact(identity, case, 7, "王月嫂", "812", account, MoneyNTD(amount), date(2026, 8, 31), status)


def _refund(identity, amount, *, payable=True, anomaly=False, refund_type="customer_refund"):
    return ClientRefundExportFact(identity, "CASE-R", "林客戶", "004", "998", MoneyNTD(amount), date(2026, 8, 31), payable, anomaly, refund_type)


def _government_return(identity="government-return:1", amount=600):
    return GovernmentOverpaymentReturnExportFact(identity, "overpayment:1", "新竹市政府", "004", "****1234", MoneyNTD(amount), date(2026, 8, 31))


def test_aggregate_merges_same_staff_day_and_excludes_non_payable_or_anomalous_rows():
    rows = aggregate_accounts_payable_rows(
        (_staff("obligation:2", "CASE-2", 200), _staff("obligation:1", "CASE-1", 300), _staff("obligation:3", "CASE-3", 99, status=StaffPayableStatus.COMPLETED)),
        (_refund("refund:1", 120), _refund("refund:2", 88, anomaly=True), _refund("refund:3", 77, payable=False)),
    )

    assert [(item.payment_type, item.amount.amount, item.obligation_identities) for item in rows] == [
        ("client_refund", 120, ("refund:1",)),
        ("staff_payable", 500, ("obligation:1", "obligation:2")),
    ]


def test_aggregate_rejects_inconsistent_staff_bank_identity():
    with pytest.raises(ValueError, match="accounts_payable_export_has_anomaly"):
        aggregate_accounts_payable_rows((_staff("obligation:1", "CASE-1", 100), _staff("obligation:2", "CASE-2", 100, account="002")), ())


def test_aggregate_includes_government_return_as_an_accounting_detail_only():
    rows = aggregate_accounts_payable_rows((), (), (_government_return(),))

    assert [(row.payment_type, row.amount.amount, row.obligation_identities) for row in rows] == [
        ("government_overpayment_return", 600, ("government-return:1",)),
    ]


def test_export_uses_snapshot_writes_hash_verified_archive_and_returns_workbook():
    archive = _Archive()
    workflow = AccountsPayableExportWorkflow(_Source((_staff("obligation:1", "CASE-1", 500),)), _Source((_refund("refund:1", 120),)), _Source(()), archive, _Snapshot, lambda: datetime(2026, 8, 31, 9, tzinfo=timezone.utc))

    receipt = workflow.export(date(2026, 8, 31))

    assert receipt.row_count == 2
    assert receipt.sha256 == archive.calls[0][3]
    assert receipt.filename.endswith(".xlsx")
    assert receipt.workbook_bytes.startswith(b"PK")


def test_export_rejects_archive_hash_mismatch():
    archive = _Archive("wrong")
    workflow = AccountsPayableExportWorkflow(_Source((_staff("obligation:1", "CASE-1", 500),)), _Source(()), _Source(()), archive, _Snapshot, lambda: datetime(2026, 8, 31, 9, tzinfo=timezone.utc))

    with pytest.raises(RuntimeError, match="accounts_payable_archive_failed"):
        workflow.export(date(2026, 8, 31))


def test_export_keeps_the_main_fixed_transfer_columns_for_client_subsidy_return():
    rows = aggregate_accounts_payable_rows((), (_refund("subsidy:1", 120, refund_type="subsidy_return"),))
    workbook = load_workbook(BytesIO(build_accounts_payable_workbook(rows)))
    values = list(workbook.active.values)

    assert values[0] == (
        "月份-銀行代碼-流水號", "銀行名稱", "客戶or服務人員姓名", "銀行帳號", "銀行代號(碼)", "金額", "身分證字號(匯款到永豐才要填)", "案件編號", "匯款日期",
    )
    assert values[1][:-1] == ("8-633-1", "台新銀行", "林客戶", "998", "004", 120, None, "CASE-R")
    assert values[1][-1].date() == date(2026, 8, 31)
