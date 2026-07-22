from datetime import date
from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import load_workbook

from services import accounts_payable_export


class FakeCursor:
    def __init__(self, staff_rows, client_rows):
        self.staff_rows = staff_rows
        self.client_rows = client_rows
        self.executed = []

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc, traceback):
        return False

    def execute(self, sql, params):
        self.executed.append((" ".join(sql.split()), params))

    def fetchall(self):
        sql = self.executed[-1][0]
        return self.staff_rows if "FROM staff_payments sp" in sql else self.client_rows


class FakeConnection:
    def __init__(self, staff_rows, client_rows):
        self.cursor_instance = FakeCursor(staff_rows, client_rows)
        self.closed = False

    def cursor(self):
        return self.cursor_instance

    def close(self):
        self.closed = True


def test_build_accounts_payable_export_is_read_only_and_bank_serials_reset(monkeypatch):
    connection = FakeConnection(
        staff_rows=[
            {
                "source_payment_id": 4,
                "staff_id": 11,
                "case_no": "115000004",
                "transfer_date": date(2026, 7, 20),
                "amount": Decimal("8000.00"),
                "amount_paid": Decimal("8000.00"),
                "payment_status": "paid",
                "recipient_name": "月嫂乙",
                "identity_card": "B200000002",
                "account_no": "002222222222",
                "recipient_bank_code": "812",
            },
            {
                "source_payment_id": 2,
                "staff_id": 11,
                "case_no": "115000002",
                "transfer_date": date(2026, 7, 10),
                "amount": Decimal("10000.00"),
                "amount_paid": Decimal("5000.00"),
                "payment_status": "partially_paid",
                "recipient_name": "月嫂乙",
                "identity_card": "B200000002",
                "account_no": "002222222222",
                "recipient_bank_code": "812",
            },
            {
                "source_payment_id": 1,
                "staff_id": 10,
                "case_no": "115000001",
                "transfer_date": date(2026, 7, 10),
                "amount": Decimal("32000.00"),
                "amount_paid": Decimal("32000.00"),
                "payment_status": "paid",
                "recipient_name": "月嫂甲",
                "identity_card": "A100000001",
                "account_no": "001111111111",
                "recipient_bank_code": "004",
            },
        ],
        client_rows=[
            {
                "source_payment_id": 8,
                "case_no": "115000003",
                "transfer_date": date(2026, 7, 15),
                "amount": Decimal("12000.00"),
                "subsidy_return_refunded": Decimal("12000.00"),
                "subsidy_return_review_status": "review_required",
                "subsidy_return_review_reason": "歷史身分類別需覆核",
                "recipient_name": "客戶丙",
                "account_no": "003333333333",
                "recipient_bank_code": "8220012",
            }
        ],
    )
    monkeypatch.setattr(accounts_payable_export, "get_connection", lambda: connection)

    result = accounts_payable_export.build_accounts_payable_export("2026-07")

    rows = result["payable_rows"]
    assert [row["月份-銀行代碼-流水號"] for row in rows] == [
        "7-31-1",
        "7-31-2",
        "7-633-1",
    ]
    assert rows[0]["案件編號"] == "115000001"
    assert rows[1]["案件編號"] == "115000002,115000004"
    assert rows[1]["金額"] == Decimal("18000.00")
    assert rows[1]["付款狀態"] == "paid,partially_paid"
    assert rows[2]["客戶or服務人員姓名"] == "客戶丙"
    assert rows[2]["身分證字號(匯款到永豐才要填)"] == ""
    assert rows[2]["金額"] == Decimal("12000.00")
    assert rows[2]["付款狀態"] == "paid"
    assert rows[2]["覆核狀態"] == "review_required"
    assert rows[2]["覆核原因"] == "歷史身分類別需覆核"
    assert result["bank_totals"] == {
        "31": Decimal("50000.00"),
        "633": Decimal("12000.00"),
    }

    assert connection.closed is True
    assert len(connection.cursor_instance.executed) == 2
    for sql, _params in connection.cursor_instance.executed:
        assert not any(keyword in sql.upper() for keyword in ("INSERT ", "UPDATE ", "DELETE "))

    staff_sql, staff_params = connection.cursor_instance.executed[0]
    assert "FROM staff_payments sp" in staff_sql
    assert "COALESCE(od.salary_payment_date_1, sp.due_date) BETWEEN %s AND %s" in staff_sql
    assert "od.order_status IN ('訂單成立', '服務中', '訂單完成')" in staff_sql
    assert "settlement_month" not in staff_sql
    assert "payment_status <>" not in staff_sql
    assert "amount_paid" not in staff_sql.split("WHERE", 1)[1]
    assert staff_params == (date(2026, 7, 1), date(2026, 7, 31))

    client_params = connection.cursor_instance.executed[1][1]
    assert client_params == (date(2026, 7, 1), date(2026, 7, 31))

    client_sql = connection.cursor_instance.executed[1][0].upper()
    assert "CLIENT_PAYMENTS" in client_sql
    assert "SUBSIDY_RETURN_RECEIVABLE" in client_sql
    assert "SUBSIDY_REFUND" not in client_sql
    assert "SUBSIDY_SALARY" not in client_sql
    assert "CP.SUBSIDY_RETURN_RECEIVABLE > 0" in client_sql
    assert "SUBSIDY_RETURN_RECEIVABLE -" not in client_sql
    assert "REVIEW_STATUS" not in client_sql.split("WHERE", 1)[1]
    assert "OD.ORDER_STATUS IN ('訂單成立', '服務中', '訂單完成')" in client_sql

    workbook = load_workbook(BytesIO(result["xlsx_bytes"]))
    worksheet = workbook["應付匯款清單"]
    assert worksheet.max_column == 9
    assert [cell.value for cell in worksheet[1]] == list(accounts_payable_export.EXPORT_HEADERS)
    assert all(cell.fill.fgColor.rgb == "00FFFF00" for cell in worksheet[1])
    assert worksheet.cell(row=2, column=1).value == "7-31-1"
    assert worksheet.cell(row=6, column=1).value == "永豐銀行總額"
    assert worksheet.cell(row=7, column=1).value == "台新銀行總額"


def test_staff_rows_group_by_staff_id_and_deduplicate_cases_stably(monkeypatch):
    connection = FakeConnection(
        staff_rows=[
            {
                "source_payment_id": 3,
                "staff_id": 21,
                "case_no": "115000020",
                "transfer_date": date(2026, 8, 15),
                "amount": Decimal("3000.00"),
                "amount_paid": Decimal("0"),
                "payment_status": "pending",
                "recipient_name": "同名月嫂",
                "identity_card": "A123456789",
                "account_no": "111",
                "recipient_bank_code": "004",
            },
            {
                "source_payment_id": 1,
                "staff_id": 21,
                "case_no": "115000010",
                "transfer_date": date(2026, 8, 15),
                "amount": Decimal("5000.00"),
                "amount_paid": Decimal("5000.00"),
                "payment_status": "paid",
                "recipient_name": "同名月嫂",
                "identity_card": "A123456789",
                "account_no": "111",
                "recipient_bank_code": "004",
            },
            {
                "source_payment_id": 2,
                "staff_id": 21,
                "case_no": "115000010",
                "transfer_date": date(2026, 8, 15),
                "amount": Decimal("2000.00"),
                "amount_paid": Decimal("1000.00"),
                "payment_status": "partially_paid",
                "recipient_name": "同名月嫂",
                "identity_card": "A123456789",
                "account_no": "111",
                "recipient_bank_code": "004",
            },
            {
                "source_payment_id": 4,
                "staff_id": 22,
                "case_no": "115000030",
                "transfer_date": date(2026, 8, 15),
                "amount": Decimal("7000.00"),
                "amount_paid": Decimal("7000.00"),
                "payment_status": "paid",
                "recipient_name": "同名月嫂",
                "identity_card": "B123456789",
                "account_no": "222",
                "recipient_bank_code": "812",
            },
        ],
        client_rows=[],
    )
    monkeypatch.setattr(accounts_payable_export, "get_connection", lambda: connection)

    result = accounts_payable_export.build_accounts_payable_export("2026-08")

    rows = result["payable_rows"]
    assert len(rows) == 2
    assert [row["月份-銀行代碼-流水號"] for row in rows] == ["8-31-1", "8-31-2"]
    assert rows[0]["案件編號"] == "115000010,115000020"
    assert rows[0]["金額"] == Decimal("10000.00")
    assert rows[1]["案件編號"] == "115000030"
    assert rows[1]["金額"] == Decimal("7000.00")
    assert rows[0]["客戶or服務人員姓名"] == rows[1]["客戶or服務人員姓名"]
    assert result["bank_totals"] == {
        "31": Decimal("17000.00"),
        "633": Decimal("0"),
    }


def test_zero_subsidy_candidate_is_not_inferred_or_exported(monkeypatch):
    connection = FakeConnection(
        staff_rows=[],
        client_rows=[
            {
                "source_payment_id": 9,
                "case_no": "115000099",
                "transfer_date": date(2026, 9, 5),
                "amount": Decimal("0"),
                "subsidy_return_refunded": Decimal("0"),
                "subsidy_return_review_status": "review_required",
                "subsidy_return_review_reason": "尚未建立正式義務",
                "recipient_name": "零額候選",
                "account_no": "999",
                "recipient_bank_code": "004",
            }
        ],
    )
    monkeypatch.setattr(accounts_payable_export, "get_connection", lambda: connection)

    result = accounts_payable_export.build_accounts_payable_export("2026-09")

    assert result["payable_rows"] == []
    assert result["bank_totals"] == {"31": Decimal("0"), "633": Decimal("0")}
    client_sql = connection.cursor_instance.executed[1][0].upper()
    assert "CP.SUBSIDY_RETURN_RECEIVABLE AS AMOUNT" in client_sql
    assert "CP.SUBSIDY_RETURN_RECEIVABLE > 0" in client_sql
    assert "SUBSIDY_SALARY" not in client_sql


@pytest.mark.parametrize(
    ("target_month", "expected_bounds"),
    [
        ("2025-12", (date(2025, 12, 1), date(2025, 12, 31))),
        ("2026-08", (date(2026, 8, 1), date(2026, 8, 31))),
        ("2028-02", (date(2028, 2, 1), date(2028, 2, 29))),
    ],
)
def test_query_month_bounds_include_past_future_and_exact_edges(
    monkeypatch, target_month, expected_bounds
):
    connection = FakeConnection(staff_rows=[], client_rows=[])
    monkeypatch.setattr(accounts_payable_export, "get_connection", lambda: connection)

    accounts_payable_export.build_accounts_payable_export(target_month)

    assert [params for _sql, params in connection.cursor_instance.executed] == [
        expected_bounds,
        expected_bounds,
    ]


def test_bank_serials_restart_for_each_export(monkeypatch):
    staff_rows = [
        {
            "source_payment_id": 1,
            "staff_id": 31,
            "case_no": "115000031",
            "transfer_date": date(2026, 10, 15),
            "amount": Decimal("1000"),
            "amount_paid": Decimal("0"),
            "payment_status": "pending",
            "recipient_name": "月嫂",
            "identity_card": "A100000001",
            "account_no": "123",
            "recipient_bank_code": "004",
        }
    ]
    connections = iter(
        [
            FakeConnection(staff_rows=staff_rows, client_rows=[]),
            FakeConnection(staff_rows=staff_rows, client_rows=[]),
        ]
    )
    monkeypatch.setattr(accounts_payable_export, "get_connection", lambda: next(connections))

    first = accounts_payable_export.build_accounts_payable_export("2026-10")
    second = accounts_payable_export.build_accounts_payable_export("2026-10")

    assert first["payable_rows"][0]["月份-銀行代碼-流水號"] == "10-31-1"
    assert second["payable_rows"][0]["月份-銀行代碼-流水號"] == "10-31-1"


def test_completed_summary_uses_only_canonical_subsidy_receivable(monkeypatch):
    connection = FakeConnection(staff_rows=[], client_rows=[])
    monkeypatch.setattr(accounts_payable_export, "get_connection", lambda: connection)

    result = accounts_payable_export.build_completed_order_payables_summary("2026-07")

    assert result["summary_rows"] == []
    summary_sql = connection.cursor_instance.executed[0][0].upper()
    assert "CP.SUBSIDY_RETURN_RECEIVABLE" in summary_sql
    assert "SUBSIDY_SALARY" not in summary_sql


def test_invalid_target_month_is_rejected_before_database_access(monkeypatch):
    monkeypatch.setattr(
        accounts_payable_export,
        "get_connection",
        lambda: (_ for _ in ()).throw(AssertionError("database must not be opened")),
    )

    try:
        accounts_payable_export.build_accounts_payable_export("2026-7")
    except ValueError as exc:
        assert "YYYY-MM" in str(exc)
    else:
        raise AssertionError("invalid target_month must be rejected")
