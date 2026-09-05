"""
File: test_weekly_report_batch_workflow.py
Description: 驗證營運週報方案 C 結算批次封存、未結算案件順延、指標儲存與 XLSX D/F/G 跨列合併。
"""

from __future__ import annotations

from argparse import Namespace
from datetime import date, datetime
from io import BytesIO
import os

import openpyxl
import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient

from api.dependencies.admin_auth import require_admin
from api.exception_handlers import CorrelationBoundaryMiddleware, install_typed_error_handlers
from api.routes import operations_reports
from infrastructure.mysql.mysql_adapter import get_connection
from subsystems.access.authentication_session import AdminPrincipal
from subsystems.reporting.weekly_report_batch_service import WeeklyReportBatchService


@pytest.fixture(scope="module")
def weekly_report_database():
    from scripts.bootstrap_disposable_mysql_schema import bootstrap
    from scripts.migrate_weekly_report_batches import DDL_SQL
    from scripts.reset_fake_database import split_sql

    database = os.getenv("LABOR_UNION_TEST_MYSQL_DATABASE", "")
    if not database:
        pytest.fail("weekly report integration requires explicit disposable MySQL configuration")
    bootstrap(Namespace(
        host=os.environ["LABOR_UNION_TEST_MYSQL_HOST"],
        port=int(os.environ["LABOR_UNION_TEST_MYSQL_PORT"]),
        user=os.environ["LABOR_UNION_TEST_MYSQL_USER"],
        password=os.environ["LABOR_UNION_TEST_MYSQL_PASSWORD"],
        database=database,
        confirm_database=database,
    ))
    connection = get_connection()
    try:
        # These reporting tables are not in the canonical fresh assembly yet.
        # Exercise their existing DDL only inside the newly created test database.
        with connection.cursor() as cursor:
            for statement in split_sql(DDL_SQL):
                cursor.execute(statement)
        connection.commit()
        WeeklyReportBatchService(connection).close_batch(
            year=2026, week_code="6-1", promotion_count=20, inquiry_count=10,
            case_nos=[], cutoff_at=datetime(2026, 6, 30, 12),
        )
    finally:
        connection.close()


def _app():
    app = FastAPI()
    app.include_router(operations_reports.router)
    app.add_middleware(CorrelationBoundaryMiddleware)
    install_typed_error_handlers(app)
    app.dependency_overrides[require_admin] = lambda: AdminPrincipal(
        id=1,
        username="admin",
        display_name="管理員",
        role="admin",
        linked_line_user_id=None,
        capabilities=("operations.report.read", "operations.report.manage"),
        is_root=True,
        enabled=True,
        access_control_version=1,
    )
    return app


@pytest.mark.integration
def test_batch_service_close_and_unclosed_cases_flow(weekly_report_database):
    conn = get_connection()
    service = WeeklyReportBatchService(conn)
    test_year = 2027
    w1 = "27-1"
    w2 = "27-2"

    with conn.cursor() as cur:
        # 建立測試案件
        cur.execute(
            """
            INSERT INTO clients (case_no, name, created_at, identity_status)
            VALUES ('TEST-CASE-B1', '張早', '2027-06-03 09:00:00', '一般市民'),
                   ('TEST-CASE-B2', '李午', '2027-06-03 15:00:00', '一般市民')
            """
        )
        cur.execute(
            """
            INSERT INTO orders (case_no, client_id, status, created_at)
            SELECT case_no, id, '訂單成立', created_at FROM clients WHERE case_no IN ('TEST-CASE-B1', 'TEST-CASE-B2')
            """
        )
    conn.commit()

    try:
        # 1. 早上 10:00 結算 27-1 批次，指定或依時間包含 TEST-CASE-B1
        b1 = service.close_batch(
            year=test_year,
            week_code=w1,
            promotion_count=20,
            inquiry_count=10,
            case_nos=["TEST-CASE-B1"],
            cutoff_at=datetime(2027, 6, 3, 10, 0),
        )
        assert b1.week_code == w1
        assert b1.promotion_count == 20
        assert b1.inquiry_count == 10

        # 2. 檢查未結算案件池：TEST-CASE-B1 已被結算，TEST-CASE-B2 仍在未結算池中
        unclosed = service.get_unclosed_cases(test_year)
        unclosed_nos = [c.case_no for c in unclosed]
        assert "TEST-CASE-B1" not in unclosed_nos
        assert "TEST-CASE-B2" in unclosed_nos

        # 3. 下午或下週結算 99-2 批次，納入 TEST-CASE-B2
        b2 = service.close_batch(
            year=test_year,
            week_code=w2,
            promotion_count=15,
            inquiry_count=8,
            case_nos=["TEST-CASE-B2"],
            cutoff_at=datetime(2027, 6, 3, 16, 0),
        )
        assert b2.week_code == w2
        assert b2.case_count == 1

        # 4. 驗證微調指標
        b2_updated = service.update_batch_metrics(b2.id, promotion_count=18, inquiry_count=9)
        assert b2_updated.promotion_count == 18
        assert b2_updated.inquiry_count == 9

        # 5. 驗證指標對應字典
        metrics_map = service.get_weekly_metrics_map(test_year)
        assert metrics_map[w1] == (20, 10)
        assert metrics_map[w2] == (18, 9)

    finally:
        conn.close()


@pytest.mark.integration
def test_batch_api_endpoints(weekly_report_database):
    app = _app()
    client = TestClient(app)
    # 測試 GET /api/v1/operations-reports/weekly/batches
    res = client.get("/api/v1/operations-reports/weekly/batches", params={"year": 2026})
    assert res.status_code == 200
    data = res.json()["data"]
    assert isinstance(data, list)
    assert len(data) > 0
    # 檢查 fixture 明確建立的 6-1 批次，不依賴操作環境或歷史 Excel 資料。
    batch_6_1 = next((b for b in data if b["week_code"] == "6-1"), None)
    assert batch_6_1 is not None
    assert batch_6_1["promotion_count"] == 20
    assert batch_6_1["inquiry_count"] == 10

    # 測試 GET /api/v1/operations-reports/weekly/unclosed-cases
    res_unclosed = client.get("/api/v1/operations-reports/weekly/unclosed-cases")
    assert res_unclosed.status_code == 200
    assert isinstance(res_unclosed.json()["data"], list)


def test_export_merges_dfg_columns_with_saved_weekly_metrics():
    from subsystems.reporting.weekly_operations_report_query import (
        WeeklyOperationsReport, WeeklySummary, WeeklyCaseRow
    )
    from subsystems.reporting.weekly_operations_report_export import export_weekly_operations_report

    case1 = WeeklyCaseRow(
        case_no="CASE-01", applicant_name="張心俞", application_date=date(2026, 6, 2),
        identity_status="一般市民", review_result="general_eligible", order_status="訂單成立",
        service_days=10, service_hours_per_day=8, planned_start_date=date(2026, 6, 10),
        planned_end_date=date(2026, 6, 20), district="板橋區", data_quality_codes=(),
        serial_number=166, month_label="6月", application_date_roc="115/06/02",
        week_code="6-1", general_eligible=1, general_ineligible=0, subsidized_eligible=0,
        subsidized_ineligible=0, order_established=1, negotiating=0, cancelled=0,
        review_rejected=0, service_status="服務中", notes="",
    )
    case2 = WeeklyCaseRow(
        case_no="CASE-02", applicant_name="張育芩", application_date=date(2026, 6, 2),
        identity_status="一般市民", review_result="general_eligible", order_status="訂單成立",
        service_days=10, service_hours_per_day=8, planned_start_date=date(2026, 6, 10),
        planned_end_date=date(2026, 6, 20), district="板橋區", data_quality_codes=(),
        serial_number=167, month_label="", application_date_roc="115/06/02",
        week_code="6-1", general_eligible=1, general_ineligible=0, subsidized_eligible=0,
        subsidized_ineligible=0, order_established=1, negotiating=0, cancelled=0,
        review_rejected=0, service_status="服務中", notes="",
    )
    case3 = WeeklyCaseRow(
        case_no="CASE-03", applicant_name="王喬可", application_date=date(2026, 6, 3),
        identity_status="一般市民", review_result="general_eligible", order_status="訂單成立",
        service_days=10, service_hours_per_day=8, planned_start_date=date(2026, 6, 15),
        planned_end_date=date(2026, 6, 25), district="板橋區", data_quality_codes=(),
        serial_number=168, month_label="", application_date_roc="115/06/03",
        week_code="6-2", general_eligible=1, general_ineligible=0, subsidized_eligible=0,
        subsidized_ineligible=0, order_established=1, negotiating=0, cancelled=0,
        review_rejected=0, service_status="服務中", notes="",
    )

    report = WeeklyOperationsReport(
        schema_version="operations-report.v2",
        start_date=date(2026, 6, 1),
        end_date=date(2026, 6, 30),
        timezone="Asia/Taipei",
        period_label="2026-06-01 ~ 2026-06-30",
        generated_at=datetime.now(),
        source_revision="test",
        summary=WeeklySummary(promotion_count=30, inquiry_count=18, application_count=3,
                              general_eligible_count=3, general_ineligible_count=0,
                              subsidized_eligible_count=0, subsidized_ineligible_count=0,
                              rejection_unpartitioned_count=0, order_established_count=3,
                              negotiating_count=0, cancelled_count=0, incomplete_count=0),
        case_rows=(case1, case2, case3),
        subsidy_partitions=(),
        service_rows=(),
        data_quality_issues=(),
        weekly_metrics={"6-1": (20, 10), "6-2": (10, 8)},
    )

    xlsx_bytes = export_weekly_operations_report(report)
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), data_only=False)
    ws = wb["週報案件受理總表"]

    merged_str = [str(m) for m in ws.merged_cells.ranges]
    assert "D6:D7" in merged_str
    assert "F6:F7" in merged_str
    assert "G6:G7" in merged_str

    assert ws["D6"].value == "6-1"
    assert ws["F6"].value == 20
    assert ws["G6"].value == 10
    assert ws["D8"].value == "6-2"
    assert ws["F8"].value == 10
    assert ws["G8"].value == 8
